import os
import shutil
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

# ---------- util ----------
def salvar_excel(df, destino, sheet_name="Planilha"):
    """Salva e garante fechamento correto do Excel."""
    with pd.ExcelWriter(destino, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

def copiar_arquivo(origem, destino):
    try:
        os.makedirs(os.path.dirname(destino), exist_ok=True)
        shutil.copy(origem, destino)
        print(f"✅ Arquivo copiado de {origem} para {destino} mantendo a formatação.")
    except Exception as e:
        print(f"❌ Erro ao copiar o arquivo: {e}")

# ---------- limpeza ----------
def limpar_linhas(arquivo, coluna, saida=None):
    if arquivo.endswith(".csv"):
        df = pd.read_csv(arquivo, dtype=str, keep_default_na=True)
    elif arquivo.endswith(".xlsx"):
        df = pd.read_excel(arquivo, engine="openpyxl")
    else:
        raise ValueError("Formato de arquivo não suportado.")

    if coluna not in df.columns:
        raise ValueError(f"A coluna '{coluna}' não existe no arquivo.")

    linhas_antes = len(df)
    mask = df[coluna].notna() & (df[coluna].astype(str).str.strip() != "-")
    df_filtrado = df[mask].reset_index(drop=True)
    linhas_depois = len(df_filtrado)

    print(f"Linhas removidas: {linhas_antes - linhas_depois}")

    destino = saida if saida else arquivo
    if destino.endswith(".csv"):
        df_filtrado.to_csv(destino, index=False)
    else:
        salvar_excel(df_filtrado, destino)
    return df_filtrado

def excluir_linhas_total_zero(arquivo, coluna, header=0):
    df = pd.read_excel(arquivo, header=header)
    df[coluna] = pd.to_numeric(df[coluna], errors="coerce").fillna(0)
    df = df[df[coluna] != 0]
    print(f"✅ Linhas com {coluna} = 0 foram excluídas.")
    salvar_excel(df, arquivo)

# ---------- concatenação ----------
def concatenar_colunas(arquivo, col1, col2, col3, col4, col5, col6, col7, col8,  nova_coluna, separador=" ", saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")

    for col in [col1, col2, col3, col4, col5, col6]:
        if col not in df.columns:
            raise ValueError(f"A coluna '{col}' não existe.")

    def limpar_valor(x, cortar=None):
        if pd.isna(x):
            return ""
        try:
            # se for float com .0, transforma em int
            if isinstance(x, float) and x.is_integer():
                return str(int(x))
        except:
            pass
        x_str = str(x)
        if cortar:
            return x_str[:cortar]
        return x_str

    df[nova_coluna] = (
        df[col1].apply(limpar_valor) + separador +
        df[col2].apply(lambda v: limpar_valor(v, cortar=5)) + separador +
        df[col3].apply(limpar_valor) + separador +
        df[col4].apply(limpar_valor) + separador +
        df[col5].apply(limpar_valor) + separador +
        df[col6].apply(limpar_valor) + separador +
        df[col7].apply(limpar_valor) + separador +
        df[col8].apply(limpar_valor)
    ).str.strip()

    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    print(f"✅ Coluna '{nova_coluna}' criada em {destino}")
    return df

def concatenar_colunas2(arquivo, col1, col2, nova_coluna, separador=" ", saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")

    # Verifica se as colunas existem
    if col1 not in df.columns or col2 not in df.columns:
        raise ValueError(f"Colunas '{col1}' ou '{col2}' não existem.")

    # Converte para string, remove .0 e trata NaN -> vazio
    col1_str = df[col1].apply(lambda x: "" if pd.isna(x) else str(x).replace(".0", ""))
    col2_str = df[col2].apply(lambda x: "" if pd.isna(x) else str(x))

    # Concatena e substitui casos de " / " sozinho por vazio
    df[nova_coluna] = (col1_str + separador + col2_str).str.strip()
    df[nova_coluna] = df[nova_coluna].replace({"^/?\s*/?\s*$": ""}, regex=True)

    # Salva
    destino = saida if saida else arquivo
    with pd.ExcelWriter(destino, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name="Planilha")

    print(f"✅ Coluna '{nova_coluna}' criada em {destino}")
    return df

# ---------- procurar/preencher ----------
def procurar_e_preencher2(
    arquivo_base,
    col_chave_base,
    arquivo_ref,
    col_chave_ref,
    col_valor_ref,
    nova_coluna,
    saida=None,
    aba_base=None,   # <<< nova opção
    aba_ref=None     # <<< nova opção
):
    """
    Preenche/cria uma coluna em arquivo_base com valores buscados em arquivo_ref.

    - Se 'nova_coluna' não existir: cria e preenche toda.
    - Se já existir: só preenche células vazias.
    - Se 'aba_base' ou 'aba_ref' forem passados, abre a aba correspondente;
      caso contrário, usa a primeira aba do arquivo.
    """

    # Lê os arquivos (com aba específica, se informado)
    df_base = pd.read_excel(arquivo_base, engine="openpyxl", sheet_name=aba_base)
    df_ref = pd.read_excel(arquivo_ref, engine="openpyxl", sheet_name=aba_ref)

    # Cria o dicionário de referência {chave: valor}
    mapa = dict(zip(df_ref[col_chave_ref], df_ref[col_valor_ref]))

    # Se a coluna não existir, cria
    if nova_coluna not in df_base.columns:
        df_base[nova_coluna] = df_base[col_chave_base].map(mapa)
    else:
        # Só preenche onde estiver vazio
        df_base[nova_coluna] = df_base.apply(
            lambda row: row[nova_coluna]
            if pd.notna(row[nova_coluna]) and str(row[nova_coluna]).strip() != ""
            else mapa.get(row[col_chave_base], row[nova_coluna]),
            axis=1
        )

    # Salva
    destino = saida if saida else arquivo_base
    with pd.ExcelWriter(destino, engine="openpyxl", mode="w") as writer:
        df_base.to_excel(writer, index=False, sheet_name=(aba_base if aba_base else "Planilha"))

    print(f"✅ Coluna '{nova_coluna}' atualizada/criada em {destino} (aba: {aba_base or 'primeira'})")
    return df_base

# ---------- criação/renomeação ----------
def criar_coluna_mapeada(arquivo, coluna_origem, nova_coluna, mapeamento, saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")
    if coluna_origem not in df.columns:
        raise ValueError(f"A coluna '{coluna_origem}' não existe.")

    df[nova_coluna] = df[coluna_origem].map(mapeamento).fillna(df[coluna_origem])

    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    return df

def renomear_cabecalho(arquivo, novos_nomes, saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")
    if len(df.columns) != len(novos_nomes):
        raise ValueError("Número de nomes não bate com número de colunas.")

    df.columns = novos_nomes
    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    return df

def limpar_coluna(arquivo, coluna, saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")
    if coluna not in df.columns:
        raise ValueError(f"A coluna '{coluna}' não existe.")

    df[coluna] = ""
    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    return df

def copiar_colunas_para_outro_arquivo(arquivo_origem, colunas_origem, arquivo_destino, colunas_destino, saida=None):
    df_origem = pd.read_excel(arquivo_origem, engine="openpyxl")
    df_destino = pd.read_excel(arquivo_destino, engine="openpyxl")

    for c in colunas_origem:
        if c not in df_origem.columns:
            raise ValueError(f"A coluna '{c}' não existe no arquivo de origem.")
    for c in colunas_destino:
        if c not in df_destino.columns:
            raise ValueError(f"A coluna '{c}' não existe no arquivo de destino.")

    df_novos = df_origem[colunas_origem].copy()
    df_novos.columns = colunas_destino
    df_novos_completo = pd.DataFrame(columns=df_destino.columns)

    for col in colunas_destino:
        df_novos_completo[col] = df_novos[col]

    df_final = pd.concat([df_destino, df_novos_completo], ignore_index=True)

    destino = saida if saida else arquivo_destino
    salvar_excel(df_final, destino)
    return df_final

def criar_coluna_com_valor(arquivo, nova_coluna, valor, saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")
    df[nova_coluna] = valor
    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    return df

def criar_coluna_condicional(arquivo, coluna_base, nova_coluna, condicoes, saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")
    if coluna_base not in df.columns:
        raise ValueError(f"A coluna '{coluna_base}' não existe.")

    datas = pd.to_datetime(df[coluna_base], format="%d/%m/%Y", errors="coerce")
    anos = datas.dt.year
    df[nova_coluna] = anos.map(condicoes).fillna("")

    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    return df

def criar_coluna_6digitos(arquivo, coluna_origem, nova_coluna, saida=None):

    if arquivo.endswith(".csv"):
        df = pd.read_csv(arquivo, dtype=str)
    elif arquivo.endswith(".xlsx"):
        df = pd.read_excel(arquivo, engine="openpyxl")
    else:
        raise ValueError("Formato de arquivo não suportado. Use CSV ou XLSX.")

    # Garante que a coluna existe
    if coluna_origem not in df.columns:
        raise ValueError(f"A coluna '{coluna_origem}' não existe no arquivo.")

    # Cria a nova coluna (apenas os 6 primeiros dígitos/caracteres)
    df[nova_coluna] = df[coluna_origem].astype(str).str[:6]

    # Define saída
    destino = saida if saida else arquivo
    if destino.endswith(".csv"):
        df.to_csv(destino, index=False)
    elif destino.endswith(".xlsx"):
        with pd.ExcelWriter(destino, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name="Planilha")

    print(f"✅ Coluna '{nova_coluna}' criada no arquivo {destino}")
    return df


def concatenar_colunas3(
    arquivo,
    col1, col2, col3, col4, col5, col6, col7, col8, col9, col10,
    nova_coluna,
    separador="",
    saida=None
):
    df = pd.read_excel(arquivo, engine="openpyxl")

    for col in [col1, col2, col3, col4, col5, col6, col7, col8, col9, col10]:
        if col not in df.columns:
            raise ValueError(f"A coluna '{col}' não existe.")

    def limpar_valor(x, cortar=None):
        if pd.isna(x):
            return ""
        try:
            if isinstance(x, float) and x.is_integer():
                return str(int(x))
        except:
            pass
        x_str = str(x)
        if cortar:
            return x_str[:cortar]
        return x_str

    # formata o Localizador (col8): completa até 4 dígitos, mas não mexe se passar de 4
    def formatar_localizador(valor):
        v = limpar_valor(valor)
        if v.isdigit():
            if len(v) < 4:
                return v.zfill(4)  # completa com zeros à esquerda
            else:
                return v           # mantém do jeito que está
        return v

    df[nova_coluna] = (
        df[col1].apply(limpar_valor) +
        df[col2].apply(lambda v: limpar_valor(v, cortar=5)) +
        df[col3].apply(limpar_valor) +
        df[col4].apply(limpar_valor) +
        df[col5].apply(limpar_valor) +
        df[col6].apply(limpar_valor) +
        df[col7].apply(limpar_valor) +
        df[col8].apply(formatar_localizador) +
        df[col9].apply(limpar_valor) +
        df[col10].apply(limpar_valor)
    ).str.strip()

    destino = saida if saida else arquivo
    salvar_excel(df, destino)
    print(f"✅ Coluna '{nova_coluna}' criada em {destino}")
    return df

def preencher_coluna_por_condicoes(arquivo, coluna_base, coluna_destino, condicoes, saida=None):

    if arquivo.endswith(".csv"):
        df = pd.read_csv(arquivo, dtype=str)
    elif arquivo.endswith(".xlsx"):
        df = pd.read_excel(arquivo, engine="openpyxl")
    else:
        raise ValueError("Formato de arquivo não suportado. Use CSV ou XLSX.")

    if coluna_base not in df.columns:
        raise ValueError(f"A coluna '{coluna_base}' não existe no arquivo.")
    if coluna_destino not in df.columns:
        raise ValueError(f"A coluna '{coluna_destino}' não existe no arquivo.")

    # Cria uma série com os valores mapeados
    novos_valores = df[coluna_base].map(condicoes)

    # Só atualiza as células da coluna_destino que estão vazias
    df[coluna_destino] = df[coluna_destino].where(df[coluna_destino].notna() & (df[coluna_destino] != ""), novos_valores)

    # Salva
    destino = saida if saida else arquivo
    if destino.endswith(".csv"):
        df.to_csv(destino, index=False)
    elif destino.endswith(".xlsx"):
        with pd.ExcelWriter(destino, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name="Planilha")

    print(f"✅ Coluna '{coluna_destino}' preenchida no arquivo {destino}")
    return df

def criar_coluna_segundo_digito(arquivo, coluna_origem, nova_coluna, saida=None):

    if arquivo.endswith(".csv"):
        df = pd.read_csv(arquivo, dtype=str)
    elif arquivo.endswith(".xlsx"):
        df = pd.read_excel(arquivo, engine="openpyxl", dtype=str)
    else:
        raise ValueError("Formato de arquivo não suportado. Use CSV ou XLSX.")

    if coluna_origem not in df.columns:
        raise ValueError(f"A coluna '{coluna_origem}' não existe no arquivo.")

    # Pega apenas o 2º dígito (posição 1, já que Python começa do 0)
    df[nova_coluna] = df[coluna_origem].astype(str).str[1:2]

    # Salva o resultado
    destino = saida if saida else arquivo
    if destino.endswith(".csv"):
        df.to_csv(destino, index=False)
    elif destino.endswith(".xlsx"):
        with pd.ExcelWriter(destino, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name="Planilha")

    print(f"✅ Coluna '{nova_coluna}' criada no arquivo {destino}")
    return df

def procurar_e_preencher(
    arquivo_base,
    col_chave_base,
    arquivo_ref,
    col_chave_ref,
    col_valor_ref,
    nova_coluna,
    saida=None
):
    df_base = pd.read_excel(arquivo_base, engine="openpyxl")
    df_ref = pd.read_excel(arquivo_ref, engine="openpyxl")

    # Cria o dicionário de referência {chave: valor}
    mapa = dict(zip(df_ref[col_chave_ref], df_ref[col_valor_ref]))

    # Se a coluna não existir, cria e preenche toda
    if nova_coluna not in df_base.columns:
        df_base[nova_coluna] = df_base[col_chave_base].map(mapa)

    else:
        # Só preenche onde estiver vazio
        df_base[nova_coluna] = df_base.apply(
            lambda row: row[nova_coluna]
            if pd.notna(row[nova_coluna]) and str(row[nova_coluna]).strip() != ""
            else mapa.get(row[col_chave_base], row[nova_coluna]),
            axis=1
        )

    # Salva
    destino = saida if saida else arquivo_base
    salvar_excel(df_base, destino)
    print(f"✅ Coluna '{nova_coluna}' atualizada/criada em {destino}")
    return df_base


def padronizar_coluna_quatro_digitos(arquivo, coluna, saida=None):
    df = pd.read_excel(arquivo, engine="openpyxl")

    if coluna not in df.columns:
        raise ValueError(f"A coluna '{coluna}' não existe no arquivo.")

    def ajustar(valor):
        if pd.isna(valor) or str(valor).strip() == "":
            return ""   # mantém vazio
        try:
            valor_int = int(float(valor))  # converte para inteiro (remove .0)
            return f"{valor_int:04d}"      # sempre 4 dígitos
        except:
            return str(valor).strip()      # fallback caso não seja número

    df[coluna] = df[coluna].apply(ajustar)

    destino = saida if saida else arquivo
    df.to_excel(destino, index=False, engine="openpyxl")
    print(f"✅ Coluna '{coluna}' padronizada para 4 dígitos em {destino}")
    return df

def gerar_tabela_dinamica(arquivo_base, arquivo_modelo, saida=None):
    # Lê os dados do analítico
    df = pd.read_excel(arquivo_base, engine="openpyxl")

    # Cria a tabela dinâmica: soma valores por agrupamento + ano
    tabela = pd.pivot_table(
        df,
        values="Valor Descentralizado (R$)",
        index=[
            "Grupo", "TED", "SIAFI", "UO descentralizadora",
            "Unidade Descentralizadora", "Unidade Descentralizada",
            "Descrição do objeto", "RP - Descrição", "Politicas",
            "Ação", "PTRES", "PO", "GND", "PI"
        ],
        columns="Exercício Orçamentário",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # Abre o modelo com openpyxl
    wb = openpyxl.load_workbook(arquivo_modelo)
    ws = wb.active  # primeira aba (ou escolha pelo nome)

    # Data D-1
    ws["A2"] = f"Base Siafi: {(datetime.today() - timedelta(days=1)).strftime('%d/%m/%Y')}"

    # Limpa dados antigos da área de preenchimento (linha 4 em diante, sem apagar fórmulas!)
    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=4, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.value = None

    # Escreve os novos dados a partir da linha 4
    for r_idx, row in enumerate(tabela.values.tolist(), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Salva
    destino = saida if saida else arquivo_modelo
    wb.save(destino)
    print(f"✅ Tabela dinâmica gerada e copiada para {destino}")

def preencher_brancos_multiplas_colunas(arquivo, colunas_textos, saida=None):
    import pandas as pd

    # Lê o arquivo
    df = pd.read_excel(arquivo, engine="openpyxl")

    for coluna, valor in colunas_textos.items():
        if coluna not in df.columns:
            raise ValueError(f"A coluna '{coluna}' não existe no arquivo.")
        
        # Garante que tudo é texto e preenche os brancos
        df[coluna] = df[coluna].apply(
            lambda x: str(x).split(".")[0] if pd.notna(x) and str(x).strip() != "" else valor
        )

    # Salva novamente no Excel
    destino = saida if saida else arquivo
    df.to_excel(destino, index=False, engine="openpyxl")

    print(f"✅ Colunas {list(colunas_textos.keys())} preenchidas e convertidas para texto em {destino}")
    return df

def preencher_coluna_com_base_em_outra(arquivo, coluna_origem, coluna_destino, saida=None):

    # Lê o arquivo
    if arquivo.endswith(".csv"):
        df = pd.read_csv(arquivo, dtype=str)
    elif arquivo.endswith(".xlsx"):
        df = pd.read_excel(arquivo, engine="openpyxl")
    else:
        raise ValueError("Formato de arquivo não suportado. Use CSV ou XLSX.")

    # Verifica se as colunas existem
    for col in [coluna_origem, coluna_destino]:
        if col not in df.columns:
            raise ValueError(f"A coluna '{col}' não existe no arquivo.")

    # Preenche apenas as células em branco da coluna_destino
    df[coluna_destino] = df.apply(
        lambda row: row[coluna_origem] if pd.isna(row[coluna_destino]) or str(row[coluna_destino]).strip() == "" else row[coluna_destino],
        axis=1
    )

    # Salva o resultado
    destino = saida if saida else arquivo
    if destino.endswith(".csv"):
        df.to_csv(destino, index=False)
    elif destino.endswith(".xlsx"):
        df.to_excel(destino, index=False, engine="openpyxl")

    print(f"✅ Coluna '{coluna_destino}' preenchida com dados de '{coluna_origem}' onde estava em branco → {destino}")
    return df

def main():
    
    analitico = r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\Analitico TED.xlsx"
    copia_analitico = r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    politicas = r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\Politicas.xlsx"
    copia_politicas = r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Politicas.xlsx"
    ncdispensa = r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\Notas de Crédito - Dispensa de TED.xlsx"
    copiancdispensa = r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    copia_ministro_final = r"X:\Demanda Ministro Compartilhada\Execução na Ponta - Universidades e Institutos\COPIA Analitico TED.xlsx"
    
    copiar_arquivo(analitico, copia_analitico)
    copiar_arquivo(politicas, copia_politicas)
    copiar_arquivo(ncdispensa, copiancdispensa)
    
    df = pd.read_excel(r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx")
    print(df["SIAFI"].unique())
    
    df_limpo = limpar_linhas(
        arquivo=copia_analitico,     # trabalha na CÓPIA
        coluna="SIAFI",
        saida=copia_analitico        # salva na MESMA CÓPIA (limpa de verdade)
    )
    print(df_limpo.head())
    
    ver = pd.read_excel(copia_analitico, engine="openpyxl")
    print("Contagem de '-':", (ver["SIAFI"].astype(str).str.strip() == "-").sum())
    
    excluir_linhas_total_zero(copia_analitico, "Valor Descentralizado (R$)", header=0)
    
    novos_nomes = ["Emissão - Dia",	"Emitente - UG", "Descrição UG","Favorecido Doc","Descrição Favorecido Doc","NC - Evento", "Descrição Evento", "NC - PTRES", "NC", "NC - Plano Interno", "Descrição Plano Interno", "Descrição 2 Plano interno", "NC - Natureza Despesa","Descrição Natureza Despesa", "NC - Fonte Recursos", "Descrição Fonte Recurso","NC - Transferência", "NC - Valor Linha","Doc - Observação"]

    df = renomear_cabecalho(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        novos_nomes=novos_nomes,
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )

    print(df.head())
    
    df = limpar_coluna(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        coluna="NC - Transferência",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )

    print(df.head())
    
    df = concatenar_colunas2(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col1=r"Favorecido Doc",
        col2=r"Descrição Favorecido Doc",
        nova_coluna=r"Favorecido Completo",
        separador=r" / ",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    print(df.tail())
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col_chave_base="NC - PTRES",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Crédito Disponivel Geral.xlsx",
        col_chave_ref="PTRES",
        col_valor_ref="Ação Governo",
        nova_coluna="Ação",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col_chave_base="NC - PTRES",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Crédito Disponivel Geral.xlsx",
        col_chave_ref="PTRES",
        col_valor_ref="Cod PO",
        nova_coluna="PO",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col_chave_base="NC - PTRES",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Crédito Disponivel Geral.xlsx",
        col_chave_ref="PTRES",
        col_valor_ref="Resultado EOF",
        nova_coluna="RP",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )

    
    df = criar_coluna_com_valor(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        nova_coluna="UO descentralizadora",
        valor="26101 / Administração Direta",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    condicoes = {
        2024: "2024",
        2025: "2025"
    }

    df = criar_coluna_condicional(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        coluna_base="Emissão - Dia",    # coluna de datas
        nova_coluna="Exercício Orçamentário",
        condicoes=condicoes,
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )

    print(df[["Emissão - Dia", "Exercício Orçamentário"]].head())
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col_chave_base="Favorecido Doc",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\AGLOMERADO - Reformulado (COM REGIÕES).xlsx",
        col_chave_ref="UG",
        col_valor_ref="U O",
        nova_coluna="UO Descentralizada",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col_chave_base="Favorecido Doc",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\AGLOMERADO - Reformulado (COM REGIÕES).xlsx",
        col_chave_ref="UG",
        col_valor_ref="DESCRIÇÃO ÓRGÃO (Completa)",
        nova_coluna="UO Descrição",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    df = concatenar_colunas2(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        col1=r"UO Descentralizada",
        col2=r"UO Descrição",
        nova_coluna=r"UO Completa",
        separador=r" / ",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )
    
    df = criar_coluna_segundo_digito(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        coluna_origem="NC - Natureza Despesa",
        nova_coluna="GND",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx"
    )

    print(df[["NC - Natureza Despesa", "NC - Natureza Despesa"]].head())

    df = copiar_colunas_para_outro_arquivo(
        arquivo_origem=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Notas de Crédito - Dispensa de TED.xlsx",
        colunas_origem=["Favorecido Completo", "Doc - Observação", "NC - PTRES", "NC - Plano Interno", "NC - Valor Linha", "Exercício Orçamentário", "UO Completa", "Ação", "PO", "RP", "UO descentralizadora","GND"],
        arquivo_destino=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        colunas_destino=["Unidade Descentralizada", "Descrição do objeto","PTRES", "PI","Valor Descentralizado (R$)","Exercício Orçamentário","UO Descentralizada", "Ação", "PO", "RP","UO descentralizadora","GND"],
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )
    
    df = padronizar_coluna_quatro_digitos(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        coluna="PO",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )
    
    df = criar_coluna_6digitos(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        coluna_origem="Unidade Descentralizada",
        nova_coluna="UG",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df[["Unidade Descentralizada", "UG"]].head())
     
    df = concatenar_colunas(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        col1=r"Exercício Orçamentário",
        col2=r"UO descentralizadora",
        col3=r"Ação",
        col4=r"PO",
        col5=r"PI",
        col6=r"RP",
        col7=r"UG",
        col8=r"PTRES",
        nova_coluna=r"Chave",
        separador=r"",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df.head())
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        col_chave_base="Chave",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Politicas.xlsx",
        col_chave_ref="IdGestor",
        col_valor_ref="Políticas",
        nova_coluna="Politicas",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df_final.head())
    
    df_final = procurar_e_preencher(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        col_chave_base="Chave",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Politicas.xlsx",
        col_chave_ref="IdGestor",
        col_valor_ref="Localizador",
        nova_coluna="Localizador",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df_final.head())
    
    mapeamento = {
    2: "Discricionárias",
    3: "PAC",
    6: "Emenda Individual",
    7: "Emenda de Bancada"
    }

    df = criar_coluna_mapeada(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        coluna_origem="RP",
        nova_coluna="RP - Descrição",
        mapeamento=mapeamento,
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df[["RP", "RP - Descrição"]].head())
    
    mapeamento = {
    4: "Investimento",
    3: "Custeio",
    5: "Aquisição",
    
    }

    df = criar_coluna_mapeada(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        coluna_origem="GND",
        nova_coluna="GND - Descrição",
        mapeamento=mapeamento,
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df[["GND", "GND - Descrição"]].head())
    
    condicoes = {
        "Funcionamento EPT": "150016 / Secretaria de Educação Profissional e Tecnológica - SETEC",
        "Adm. SECADI": "157055 / Secretaria de Educação Continuada, Alfabetização de Jovens e Adultos, Diversidade e Inclusão - SECADI",
        "Adm. SESU": "150011 / Secretaria de Educação Superior - SESU",
        "Internacionalização IFES": "150011 / Secretaria de Educação Superior - SESU",
        "Adm. SGA": "150002 / Subsecretaria de Gestão Administrativa - SGA - SAA",
        "Internacionalização IFES": "150011 / Secretaria de Educação Superior - SESU",
        "Adm. SPO" : "150014 / Subsecretaria de Planejamento e Orçamento - SPO"
    }

    df = preencher_coluna_por_condicoes(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        coluna_base="Politicas",
        coluna_destino="Unidade Descentralizadora",
        condicoes=condicoes,
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )

    print(df[["Politicas", "Unidade Descentralizadora"]].head())
    
    
    df_final = procurar_e_preencher2(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        col_chave_base="UG",
        arquivo_ref=r"W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx",
        col_chave_ref=" UG Executora - Código",
        col_valor_ref="UG Executora - Grupo ",
        nova_coluna="Grupo",
        aba_base="Planilha",
        aba_ref="APOIO UG",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )
    
    df = concatenar_colunas3(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        col1=r"Exercício Orçamentário",
        col2=r"UO descentralizadora",
        col3=r"Ação",
        col4=r"PO",
        col5=r"PI",
        col6=r"RP",
        col7=r"GND",
        col8=r"Localizador",
        col9=r"UG",
        col10=r"PTRES",
        nova_coluna=r"Chave2",
        separador=r"",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )
    
    df = preencher_brancos_multiplas_colunas(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        colunas_textos={
            "TED": "Dispensa de TED", 
            "SIAFI": "Dispensa de TED",
        },
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )
    
    df = preencher_coluna_com_base_em_outra(
        arquivo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        coluna_origem="Unidade Descentralizadora",
        coluna_destino="Unidade Gestora da Política",
        saida=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx"
    )
    
    gerar_tabela_dinamica(
        arquivo_base=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\COPIA Analitico TED.xlsx",
        arquivo_modelo=r"W:\B - TED\7 - AUTOMAÇÃO\Dados do Ministro\Planilha modelo.xlsx"
    )
    copiar_arquivo(copia_analitico,copia_ministro_final)
    print('Processo totalmente finalizado!')

if __name__ == "__main__":
    main()       
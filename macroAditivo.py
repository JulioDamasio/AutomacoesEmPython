from openpyxl import load_workbook
from openpyxl.styles import Alignment  # Importe a classe Alignment
import shutil
import os
import pandas as pd
from datetime import datetime

def copiar_arquivo(origem, destino):
    try:
        shutil.copy(origem, destino)
        print(f"Arquivo copiado de {origem} para {destino} mantendo a formatação.")
    except Exception as e:
        print(f"Erro ao copiar o arquivo: {e}")

def abrir_e_desmescle_excel(caminho_arquivo_copia):
    wb = load_workbook(caminho_arquivo_copia)
    sheet = wb.active
    merged_cells = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_cells:
        sheet.unmerge_cells(str(merged_cell_range))
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)
    wb.save(caminho_arquivo_copia)

def adicionar_colunas_excel(caminho_arquivo_copia):
    df = pd.read_excel(caminho_arquivo_copia)
    df['observação'] = ''
    df['data do dia'] = ''
    df.to_excel(caminho_arquivo_copia, index=False)

def apagar_linhas(caminho_arquivo):
        df = pd.read_excel(caminho_arquivo)
        df.drop(df.index[0], inplace=True)
        df.drop(df.tail(1).index, inplace=True)
        df.to_excel(caminho_arquivo, index=False)

def remover_linhas_vazias_nulas(caminho_arquivo_copia):
        df = pd.read_excel(caminho_arquivo_copia)
        df = df.dropna(subset=['SIAFI'], axis=0)
        df.to_excel(caminho_arquivo_copia, index=False)

def remove_duplicates_by_siafi(df):
    # Verifica se há duplicatas na coluna SIAFI e mantém apenas a primeira ocorrência
    df_unique = df.drop_duplicates(subset=['SIAFI'])
    
    return df_unique        

def formatar_data_vigencia(caminho_arquivo_copia):
    # Dicionário para traduzir os meses de inglês para português
    meses_pt = {
        'jan': 'jan',
        'feb': 'fev',
        'mar': 'mar',
        'apr': 'abr',
        'may': 'mai',
        'jun': 'jun',
        'jul': 'jul',
        'aug': 'ago',
        'sep': 'set',
        'oct': 'out',
        'nov': 'nov',
        'dec': 'dez'
    }
    
    # Obter a data atual no formato desejado
    data_do_dia = datetime.now().strftime('%d%b%Y').lower()
    
    # Traduzir a data atual para o formato em português
    dia = data_do_dia[:2]
    mes = meses_pt[data_do_dia[2:5]]
    ano = data_do_dia[5:]
    data_formatada = f"{dia}{mes}{ano}"
    
    def formatar_data(data):
        formatos = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y']
        for fmt in formatos:
            try:
                dt = datetime.strptime(data, fmt)
                return dt.strftime('%d') + meses_pt[dt.strftime('%b').lower()] + dt.strftime('%Y').lower()
            except ValueError:
                pass
        raise ValueError(f"Formato de data desconhecido: {data}")
    
    try:
        # Carrega o arquivo Excel com o Pandas
        df = pd.read_excel(caminho_arquivo_copia)
        
        # Converte a coluna 'Fim da Vigência' para string e depois aplica o formato desejado
        df['Fim da Vigência'] = df['Fim da Vigência'].astype(str).apply(formatar_data)
        
        # Adiciona a data formatada ao DataFrame
        df['data do dia'] = data_formatada
        
        # Salva as modificações de volta no arquivo
        df.to_excel(caminho_arquivo_copia, index=False)
        
        print("Data de vigência formatada com sucesso.")
        
    except Exception as e:
        print(f"Erro ao formatar a data de vigência: {e}")   

def preencher_observacao(caminho_arquivo_copia):
    try:
        # Carrega o arquivo Excel com o Pandas
        df = pd.read_excel(caminho_arquivo_copia)
        
        # Preenche a coluna 'observação' com o texto desejado
        df['observação'] = 'Aditivo de Vigencia'
        
        # Salva as modificações de volta no arquivo
        df.to_excel(caminho_arquivo_copia, index=False)
        
        print("Coluna de observação preenchida com sucesso.")
        
    except Exception as e:
        print(f"Erro ao preencher a coluna de observação: {e}")

def formatar_valores_arquivo(caminho_arquivo):
    try:
        # Carrega o arquivo Excel com o Pandas
        df = pd.read_excel(caminho_arquivo)
        
        # Converte a coluna 'Valor Orçamentário (R$)' para float antes de formatar
        df['Valor Orçamentário (R$)'] = df['Valor Orçamentário (R$)'].replace('[,.]', '', regex=True).astype(float)
        
         # Converte a coluna 'Valor Autorizado (R$)' para float antes de formatar
        df['Valor Autorizado (R$)'] = df['Valor Autorizado (R$)'].replace('[,.]', '', regex=True).astype(float)
        
        # Formata os valores da coluna 'Valor Orçamentário (R$)' para exibir duas casas decimais
        df['Valor Orçamentário (R$)'] = df['Valor Orçamentário (R$)'].map('{:,.2f}'.format)
        
        # Formata os valores da coluna 'Valor Autorizado (R$)' para exibir duas casas decimais
        df['Valor Autorizado (R$)'] = df['Valor Autorizado (R$)'].map('{:,.2f}'.format)
        
        # Salva as modificações de volta no arquivo
        df.to_excel(caminho_arquivo, index=False)
        
        print("Valores formatados com sucesso.")
        
    except Exception as e:
        print(f"Erro ao formatar os valores do arquivo: {e}")

def remover_pontos_virgulas(caminho_arquivo):
    try:
        # Carrega o arquivo Excel com o Pandas
        df = pd.read_excel(caminho_arquivo)
        
        # Remove pontos e vírgulas da coluna 'Valor Orçamentário (R$)'
        df['Valor Orçamentário (R$)'] = df['Valor Orçamentário (R$)'].replace('[,.]', '', regex=True)
        
        # Remove pontos e vírgulas da coluna 'Valor Autorizado (R$)'
        df['Valor Autorizado (R$)'] = df['Valor Autorizado (R$)'].replace('[,.]', '', regex=True)
        
        # Salva as modificações de volta no arquivo
        df.to_excel(caminho_arquivo, index=False)
        
        print("Pontos e vírgulas removidos com sucesso.")
        
    except Exception as e:
        print(f"Erro ao remover pontos e vírgulas: {e}")

def generate_macro_vigencia(output_directory, df):
    tela_inicial = 2  # Inicializa a variável com o número da primeira tela
    
    modelo_xml = f"""<HAScript name="Aditivo de vigência _ Macro 2" description="" timeout="60000" pausetime="300" promptall="true" blockinput="false" author="AugustoCezar" creationdate="13/03/2024 15:20:31" supressclearevents="false" usevars="false" ignorepauseforenhancedtn="true" delayifnotenhancedtn="0" ignorepausetimeforenhancedtn="true">

    <screen name="Tela1" entryscreen="true" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="&gt;incaditivo[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela2" />
        </nextscreens>
    </screen>
    """

    tela = 2  # Inicializa a variável com o número da primeira tela
    
    for index, row in df.iterrows():
        acao = str(row['Ação (Orçamentário)'])
        fim_vigencia = str(row['Fim da Vigência'])
        siafi = str(row['SIAFI'])
        ted = str(row['TED'])
        observacao = str(row['observação'])
        data_do_dia = str(row['data do dia'])

        modelo_xml += f"""
    <screen name="Tela{tela}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="42" optional="false" invertmatch="false" />
            <numinputfields number="1" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="{siafi}[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 1}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 1}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="151" optional="false" invertmatch="false" />
            <numinputfields number="66" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="{ted}[tab]{ted}[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 2}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 2}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="157" optional="false" invertmatch="false" />
            <numinputfields number="1" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="s[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 3}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 3}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="150" optional="false" invertmatch="false" />
            <numinputfields number="38" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="[tab]{fim_vigencia}[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 4}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 4}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="156" optional="false" invertmatch="false" />
            <numinputfields number="1" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="s[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 5}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 5}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="285" optional="false" invertmatch="false" />
            <numinputfields number="130" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 6}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 6}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="64" optional="false" invertmatch="false" />
            <numinputfields number="9" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="{data_do_dia}{data_do_dia}{observacao}[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 7}" />
        </nextscreens>
    </screen>

    <screen name="Tela{tela + 7}" entryscreen="false" exitscreen="false" transient="false">
        <description >
            <oia status="NOTINHIBITED" optional="false" invertmatch="false" />
            <numfields number="73" optional="false" invertmatch="false" />
            <numinputfields number="1" optional="false" invertmatch="false" />
        </description>
        <actions>
            <input value="s[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false" />
        </actions>
        <nextscreens timeout="0" >
            <nextscreen name="Tela{tela + 8}" />
        </nextscreens>
    </screen>
    """

        tela += 8  # Atualiza a variável para o próximo número de tela conforme o seu modelo

    # fecha o modelo
    modelo_xml += """    
</HAScript> 
        """

    # Salva o arquivo
    output_filename = os.path.join(output_directory, "Macro_vigencia.MAC")
    with open(output_filename, "w") as file:
        file.write(modelo_xml)

    print(f"Nova macro de vigência gerada com sucesso. O arquivo está no caminho: {output_filename}")
             
def main():
    caminho_arquivo_original = r'W:\B - TED\7 - AUTOMAÇÃO\macro de vigência\macro de vigência.xlsx'
    caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\macro de vigência\copia macro de vigencia.xlsx'
    
    copiar_arquivo(caminho_arquivo_original, caminho_arquivo_copia)
    abrir_e_desmescle_excel(caminho_arquivo_copia)
    adicionar_colunas_excel(caminho_arquivo_copia)
    apagar_linhas(caminho_arquivo_copia)
    
    # Carrega o arquivo Excel para um DataFrame
    df = pd.read_excel(caminho_arquivo_copia)
    
    remover_linhas_vazias_nulas(caminho_arquivo_copia)
    
    # Carrega o arquivo Excel para um DataFrame
    df = pd.read_excel(caminho_arquivo_copia)
    
    # Remove as linhas duplicadas com base na coluna SIAFI
    df_sem_duplicatas = remove_duplicates_by_siafi(df)
    
    # Salva o DataFrame modificado de volta no arquivo Excel
    df_sem_duplicatas.to_excel(caminho_arquivo_copia, index=False)
    formatar_data_vigencia(caminho_arquivo_copia)
    preencher_observacao(caminho_arquivo_copia)
    formatar_valores_arquivo(caminho_arquivo_copia)
    # remover_pontos_virgulas(caminho_arquivo_copia)
    
    # Carrega o arquivo Excel novamente para obter o DataFrame atualizado
    df = pd.read_excel(caminho_arquivo_copia)
    
    # Chama a função generate_macro_vigencia com o argumento df
    generate_macro_vigencia(r'W:\B - TED\7 - AUTOMAÇÃO\macro de vigência', df)
    
if __name__ == "__main__":
    main()
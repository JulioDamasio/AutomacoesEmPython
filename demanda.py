import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
import pandas as pd
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import win32com.client
from datetime import datetime

def copiar_arquivos(destino):
    origens = [
        r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\Apoio PTRES.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\Despesas liquidadas.xlsx'
    ]
    for origem in origens:
        diretorio_origem, nome_arquivo = os.path.split(origem)
        novo_nome = "COPIA " + nome_arquivo
        shutil.copy(origem, os.path.join(destino, novo_nome))  # Copia para a pasta destino com o novo nome
        print(f"Arquivo {nome_arquivo} copiado como {novo_nome} para {destino}")
        
def criar_arquivos_por_concedente(caminho_arquivo, destino, modelo_formatado, colunas_desejadas,colunas_reordenadas, nomes_finais_colunas,):
    # Carregar o arquivo Excel em um DataFrame
    df = pd.read_excel(caminho_arquivo, engine="openpyxl")

    # Remover a última linha para evitar processamento dessa linha
    df = df.iloc[:-1]

    # Remover espaços em branco dos nomes das colunas
    df.columns = df.columns.str.strip()

    # Garantir que todos os valores da coluna 'Sigla Concedente' sejam strings
    df['Sigla Concedente'] = df['Sigla Concedente'].astype(str).str.strip()

    # Filtrar as linhas onde a coluna 'TED' está vazia
    df_vazio_ted = df[df['TED'].isna() | (df['TED'] == '')]

    # Verificar se há linhas vazias na coluna 'TED'
    if df_vazio_ted.empty:
        print("Não há linhas com TED vazio.")
        return

    # Obter as siglas únicas da coluna 'Sigla Concedente'
    siglas_unicas = df_vazio_ted['Sigla Concedente'].unique()
    
    # Criar um arquivo para cada sigla concedente
    for sigla in siglas_unicas:
        
        # Filtrar as linhas correspondentes à sigla atual
        df_sigla = df_vazio_ted[df_vazio_ted['Sigla Concedente'] == sigla]
        
        # Manter apenas as colunas desejadas, na ordem especificada
        df_sigla = df_sigla[colunas_desejadas]
        
        # reordenar as colunas para a ordem desejada
        df_sigla = df_sigla[colunas_reordenadas]

        # Renomear as colunas conforme o array de novos nomes
        df_sigla = df_sigla.rename(columns=dict(zip(colunas_reordenadas, nomes_finais_colunas)))
        
        # Criar o nome do arquivo de saída
        nome_arquivo = f'Devolução dos Valores de Emendas {sigla}.xlsx'
        caminho_salvar = os.path.join(destino, nome_arquivo)

        # Copiar o modelo de formatação para criar o novo arquivo
        wb = load_workbook(modelo_formatado)
        ws = wb.active  # Usar a aba ativa do modelo

        # Inserir os dados no modelo a partir da linha 9 (ignorando o cabeçalho do DataFrame)
        for i, row in enumerate(df_sigla.itertuples(index=False), start=9):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)
                
        # Salvar o arquivo Excel com o nome gerado
        wb.save(caminho_salvar)        
        
    print(f"Arquivo '{nome_arquivo}' criado com sucesso em {destino}")


# Dicionário com os e-mails para cada sigla concedente
emails_concedentes = {
    "SECADI": ["rodolfo.souza@mec.gov.br", "spo.ted@mec.gov.br"],
    "SETEC": ["fabioibiapina@mec.gov.br", "edsonfonseca@mec.gov.br", "eduardocamara@mec.gov.br", "spo.ted@mec.gov.br"],
    "SESU" : ["aldousalbuquerque@mec.gov.br", "marciomarques@mec.gov.br", "igorsegovia@mec.gov.br", "spo.ted@mec.gov.br"],
    "SEB"  : ["marileusaoliveira@mec.gov.br", "taniabatista@mec.gov.br", "spo.ted@mec.gov.br"]
}

# Caminho da assinatura padrão (imagem)
caminho_assinatura = r"W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\Assinatura.PNG"

def enviar_email(destino, assunto, corpo, anexos):
    try:
        # Criar objeto do Outlook
        outlook = win32com.client.Dispatch("Outlook.Application")
        email = outlook.CreateItem(0)  # Criar novo e-mail

        # Definir a conta remetente
        conta_envio = "spo.ted@mec.gov.br"
        email.SentOnBehalfOfName = conta_envio  # Define explicitamente a conta de envio

        # Definir destinatários
        email.To = "; ".join(destino)

        # Anexar a imagem da assinatura ao e-mail
        if os.path.exists(caminho_assinatura):
            anexo_assinatura = email.Attachments.Add(caminho_assinatura)
            anexo_assinatura.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001E", "assinatura_imagem")

            assinatura_html = '<br><br>'
            assinatura_html += '<img src="cid:assinatura_imagem" width="350px">'  # Insere a imagem com largura ajustada
        else:
            assinatura_html = ""  # Se não houver imagem, não adiciona nada

        # Criar corpo do e-mail em HTML
        corpo_html = f"""
        <p>Prezados(as),</p>
        <p>Encaminhamos, em anexo, uma planilha com duas abas: uma contendo os valores liquidados cujos empenhos ainda não foram cadastrados, e outra com os TEDs cujas vigências expiraram nesta secretaria.</p>

        <p>Em relação aos empenhos com valores liquidados que ainda não foram cadastrados no módulo SPO-TED, ressaltamos que as liberações financeiras ocorrem semanalmente, conforme a disponibilidade de recursos, por meio de lotes. Para garantir o repasse financeiro, esta Subsecretaria adota os seguintes critérios:</p>

        <ul>
            <li>O TED deve estar vigente. Caso seja necessário prorrogar a vigência, as unidades devem iniciar o processo de aditivo de alteração de vigência.</li>
            <li>A despesa deve estar liquidada.</li>
            <li>As Notas de Empenho (NE) devem estar corretamente cadastradas no SIMEC, na aba "Movimentação Financeira – Dados do Empenho". Caso a NE cadastrada esteja incorreta ou haja duplicidade (a mesma NE registrada em TEDs diferentes), o TED não poderá receber os recursos, devido à impossibilidade de identificar corretamente o valor a ser repassado.</li>
        </ul>

        <p>Em relação aos TEDs vencidos com valores liquidados, solicitamos que as providências necessárias sejam tomadas para regularizar a situação. Se o fato gerador da despesa tiver ocorrido dentro da vigência será necessário que seja anexado ao TED documento que ateste a liquidação da despesa dentro do período estabelecido e envie e-mail para spo.ted@mec.gov.br informando a situação e o número do TED. </p>
        
        <p>Reforçamos que o repasse financeiro será realizado conforme os critérios mencionados. Caso, após o cumprimento de todos os requisitos, a unidade não seja contemplada, pedimos que entre em contato conosco pelo e-mail: spo.ted@mec.gov.br.</p>

        <p>Atenciosamente.</p>
        {assinatura_html}  <!-- Insere a assinatura com imagem -->
        """

        # Definir o corpo do e-mail como HTML
        email.HTMLBody = corpo_html

        # Anexar os arquivos
        for anexo in anexos:
            email.Attachments.Add(anexo)

        # Definir o assunto
        email.Subject = assunto

        # Enviar o e-mail
        email.Send()
        print(f"E-mail enviado com sucesso para {', '.join(destino)}")
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")

def enviar_emails_com_planilhas(destino):
    # Percorrer os arquivos na pasta de destino
    for arquivo in os.listdir(destino):
        if arquivo.startswith("Cadastrar Empenho & TEDS Vencidos "):  # Verifica se é um arquivo válido
            sigla = arquivo.replace("Cadastrar Empenho & TEDS Vencidos ", "").replace(".xlsx", "")

            # Verificar se há e-mails para essa sigla
            if sigla in emails_concedentes:
                caminho_arquivo = os.path.join(destino, arquivo)
                destinatarios = emails_concedentes[sigla]

                # Definir assunto e corpo do e-mail
                assunto = f"Despesas liquidadas: empenhos não cadastrados SPO/TED e TEDs com vigência expirada. - {sigla}"

                # Enviar e-mail com a planilha anexada
                enviar_email(destinatarios, assunto, "", [caminho_arquivo])
            else:
                print(f"Não há destinatários cadastrados para a sigla {sigla}")
    
    
def main():
    liquidacao = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2025.xlsx'
    despesas_liquidadas = r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\Despesas liquidadas.xlsx'
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho'
    copia_despesas_liquidadas = r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\COPIA Despesas liquidadas.xlsx'
    copia_apoio_ptres = r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\COPIA Apoio PTRES.xlsx'
    modelo_formatado = r'W:\B - TED\7 - AUTOMAÇÃO\Cadastrar Empenho\Modelo de Formatação.xlsx'
    
    
    # Copiar os arquivos
    copiar_arquivos(destino)
    
    # Definir os novos nomes de colunas e o nome da nova coluna
    novos_nomes_colunas = [
        'Resultado EOF', 'DESCRIÇÃO EOF', 'NE CCor - Ano Emissão', 'Órgão UGE', 'DESCRIÇÃO UGE', 
        'UG Executora', 'DESCRIÇÃO EXECUTORA', 'UGE - UG Setorial Financeira', 'DESCRIÇÃO FINANCEIRA', 
        'Ação Governo', 'PTRES', 'PI', 'NE CCor', 'Grupo Despesa', 'Natureza Despesa Detalhada', 
        'NATUREZA', 'Elemento Despesa', 'ND', 'Fonte Recursos Detalhada', 
        'DESPESAS LIQUIDADAS A PAGAR(CONTROLE EMPENHO)', 'RESTOS A PAGAR PROCESSADOS A PAGAR', 
        'RESTOS A PAGAR NAO PROCES. LIQUIDADOS A PAGAR', 'Total', 'SITUAÇÃO', 'TED', 'SIAFI', 'Vigência', 
        'Estado Atual','Sigla Concedente'
    ]
    
    nome_novas_colunas = ['Sigla Concedente']

    
    # Definir as colunas desejadas e a ordem delas
    colunas_desejadas = ['Resultado EOF', 'NE CCor - Ano Emissão', 'Órgão UGE', 'UG Executora', 'DESCRIÇÃO EXECUTORA', 'Ação Governo',	'PTRES', 'PI','NE CCor','Grupo Despesa','Natureza Despesa Detalhada','Fonte Recursos Detalhada','Total', 'Sigla Concedente']
    
    colunas_reordenadas = ['Sigla Concedente','Resultado EOF','NE CCor - Ano Emissão', 'Órgão UGE','UG Executora','DESCRIÇÃO EXECUTORA','Ação Governo','PTRES', 'PI','NE CCor','Grupo Despesa','Natureza Despesa Detalhada','Fonte Recursos Detalhada','Total']
    
    nomes_finais_colunas = ['Unidade Descentralizadora', 'Resultado Primário (RP)', 'Ano Emissão', 'Órgão UGE', 'UG Unidade Descentralizada', 'Unidade Descentralizada',	'Ação Governo', 'PTRES','PI','Dados do Empenho','Grupo Despesa', 'Natureza Despesa Detalhada','Fonte Recursos Detalhada','Total']
    
    # Chamar a função para criar os arquivos por Sigla Concedente
    criar_arquivos_por_concedente(copia_despesas_liquidadas, destino, modelo_formatado, colunas_desejadas, colunas_reordenadas, nomes_finais_colunas)
    
    # Executar envio de e-mails
    enviar_emails_com_planilhas(destino)        
        
if __name__ == "__main__":
    main()        
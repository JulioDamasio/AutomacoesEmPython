import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

def copiar_arquivos(destino):
    origens = [
        r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\TEDS Auditoria.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\Síntese RCO e Auditorias.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\Relatorio Analitico DSV.xlsx',
    ]
    for origem in origens:
        caminho, nome_arquivo = os.path.split(origem)
        novo_nome = "COPIA " + nome_arquivo
        shutil.copy(origem, os.path.join(destino, novo_nome))  # Copia para a pasta destino com o novo nome
        print(f"Arquivo {nome_arquivo} copiado como {novo_nome} para {destino}")


def remover_linhas_indesejadas(caminho_arquivo):
    try:
        # Carregar a única aba do Excel
        df = pd.read_excel(caminho_arquivo)
        
        # Remover linhas com "Estado Atual" igual a "Arquivado" (ignorando maiúsculas/minúsculas e espaços)
        df = df[df['Estado Atual'].astype(str).str.strip().str.lower() != 'arquivado']

        # Remover linhas onde "SIAFI" está vazio, "-", ou é NaN
        df = df[~(df['SIAFI'].isna() | df['SIAFI'].astype(str).str.strip().isin(['', '-',"'-"]))]

        # Salvar o arquivo sobrescrevendo o original
        df.to_excel(caminho_arquivo, index=False)
        
        print("Linhas removidas com sucesso.")
        
    except Exception as e:
        print(f"Ocorreu um erro: {e}") 

def adicionar_colunas_excel(copia_teds_auditoria, extrato, novas_colunas, valor_padrao=None, salvar_como=None):
    
    xls = pd.read_excel(copia_teds_auditoria, sheet_name=None)

    if extrato not in xls:
        raise ValueError(f"A aba '{extrato}' não foi encontrada no arquivo.")

    df = xls[extrato]

    # Adiciona colunas
    for coluna in novas_colunas:
        if coluna not in df.columns:
            df[coluna] = valor_padrao

    # Atualiza a aba modificada
    xls[extrato] = df

    # Salva o arquivo
    caminho_saida = salvar_como if salvar_como else copia_teds_auditoria
    with pd.ExcelWriter(caminho_saida, engine='openpyxl', mode='w') as writer:
        for aba, tabela in xls.items():
            tabela.to_excel(writer, sheet_name=aba, index=False)

def preencher_colunas(copia_teds_auditoria, copia_relatorio_analitico_DSV, coluna_valores, coluna_destino, chave_coluna):
    try:
        # Carregar os DataFrames
        df_ted = pd.read_excel(copia_teds_auditoria)
        df_valores = pd.read_excel(copia_relatorio_analitico_DSV)
        
        # Corrigir possíveis espaços e caracteres indesejados nas colunas
        df_ted.columns = df_ted.columns.str.strip()
        df_valores.columns = df_valores.columns.str.strip()
        
        # Verificar se as colunas esperadas estão presentes
        if 'TED' not in df_ted.columns:
            raise ValueError(f"A coluna 'TED' não foi encontrada em {copia_teds_auditoria}.")
        if chave_coluna not in df_valores.columns:
            raise ValueError(f"A coluna chave '{chave_coluna}' não foi encontrada em {copia_relatorio_analitico_DSV}.")
        if coluna_valores not in df_valores.columns:
            raise ValueError(f"A coluna de valores '{coluna_valores}' não foi encontrada em {copia_relatorio_analitico_DSV}.")
        
        # Padronizar a chave como string SEM casas decimais e espaços
        df_valores[chave_coluna] = df_valores[chave_coluna].astype(str).str.strip().str.replace('.0', '', regex=False)
        df_ted['TED'] = df_ted['TED'].astype(str).str.strip().str.replace('.0', '', regex=False)
        
        # Elimina duplicatas mantendo o último valor encontrado para cada TED
        df_valores_unicos = df_valores.drop_duplicates(subset=chave_coluna, keep='last')

        # Cria dicionário simples
        lookup_dict = dict(zip(df_valores_unicos[chave_coluna], df_valores_unicos[coluna_valores]))
        
        # Debug: Verificar alguns valores do dicionário e da coluna de destino
        print("Exemplos de lookup_dict:")
        for k, v in list(lookup_dict.items())[:5]:
            print(f"{k}: {v}")
        
        print("Exemplos de TED em df_ted:")
        print(df_ted['TED'].head())
        
        # Preencher a coluna destino com base no lookup_dict
        df_ted[coluna_destino] = df_ted['TED'].map(lookup_dict).fillna(0)
        
        # Debug: Verificar alguns valores preenchidos
        print("Exemplos de valores preenchidos:")
        print(df_ted[[coluna_destino, 'TED']].head())
        
        # Salvar as alterações no arquivo
        df_ted.to_excel(copia_teds_auditoria, index=False)
        
        print(f"Coluna {coluna_destino} preenchida com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna {coluna_destino}: {e}")            

def preencher_coluna_ano(caminho_arquivo, coluna_data='Data da última tramitação', coluna_ano='Ano'):
    try:
        # Carregar o DataFrame
        df = pd.read_excel(caminho_arquivo)
        
        # Verificar se a coluna de data existe
        if coluna_data not in df.columns:
            raise ValueError(f"A coluna '{coluna_data}' não foi encontrada no arquivo.")
        
        # Garantir que a coluna está em formato datetime
        df[coluna_data] = pd.to_datetime(df[coluna_data], errors='coerce', dayfirst=True)

        # Criar nova coluna com o ano
        df[coluna_ano] = df[coluna_data].dt.year

        # Salvar de volta no mesmo arquivo
        df.to_excel(caminho_arquivo, index=False)
        
        print(f"Coluna '{coluna_ano}' preenchida com sucesso com base na coluna '{coluna_data}'.")
    
    except Exception as e:
        print(f"Erro ao preencher a coluna '{coluna_ano}': {e}")        

def preencher_situacao_rco(caminho_arquivo,
                           coluna_fim_vigencia='Fim da Vigência',
                           coluna_estado_atual='Estado Atual',
                           coluna_situacao='Situação RCO'):
    try:
        # Carregar o DataFrame
        df = pd.read_excel(caminho_arquivo)

        # Verificar colunas obrigatórias
        for col in [coluna_fim_vigencia, coluna_estado_atual]:
            if col not in df.columns:
                raise ValueError(f"A coluna '{col}' não foi encontrada no arquivo.")

        # Converter coluna de data
        df[coluna_fim_vigencia] = pd.to_datetime(df[coluna_fim_vigencia], errors='coerce', dayfirst=True)

        # Criar coluna Situação RCO se não existir
        if coluna_situacao not in df.columns:
            df[coluna_situacao] = ''
        
        # Garantir que os valores vazios sejam tratados corretamente
        df[coluna_situacao] = df[coluna_situacao].fillna('').astype(str).str.strip()

        # Data de hoje (sem hora)
        hoje = pd.to_datetime(datetime.today().date())

        # Aplicar Regras SOMENTE se a célula estiver vazia
        # Regra 1: Estado Atual == 'Comprovado no SIAFI.' → 'Comprovado'
        mask1 = (df[coluna_estado_atual] == 'Comprovado no SIAFI.') & (df[coluna_situacao] == '')
        df.loc[mask1, coluna_situacao] = 'Comprovado'

        # Regra 2: Fim da Vigência >= hoje → 'Vigente'
        mask2 = (df[coluna_fim_vigencia] >= hoje) & (df[coluna_situacao] == '')
        df.loc[mask2, coluna_situacao] = 'Vigente'

        # Regra 3: Estado Atual em lista → 'RCO na Descentralizadora'
        estados_descentralizadora = [
            "Aguardando aprovação pela Diretoria",
            "Relatório de cumprimento do objeto aguardando aprovação da Diretoria",
            "Relatório de cumprimento do objeto aguardando aprovação do Representante Legal da Descentralizadora",
            "Relatório de cumprimento do objeto em análise pela Coordenação",
            "Termo em alteração - Descentralizadora",
            "Termo em análise pela Coordenação",
            "Em distribuição pelo Gabinete da Secretaria/Autarquia",
            "Termo aguardando autorização do Representante Legal da Descentralizadora",
            "Termo em análise pela Coordenação"
        ]
        mask3 = df[coluna_estado_atual].isin(estados_descentralizadora) & (df[coluna_situacao] == '')
        df.loc[mask3, coluna_situacao] = 'RCO na Descentralizadora'

        # Regra final (default): Se ainda estiver vazia → 'RCO na Descentralizada'
        df.loc[df[coluna_situacao] == '', coluna_situacao] = 'RCO na Descentralizada'

        # Salvar de volta o arquivo
        df.to_excel(caminho_arquivo, index=False)

        print("Coluna 'Situação RCO' preenchida com sucesso com base em todas as regras, sem sobrescrições.")
    
    except Exception as e:
        print(f"Erro ao preencher a coluna '{coluna_situacao}': {e}")

def preencher_situacao_300_dias(caminho_arquivo):
    try:
        # Carrega o DataFrame
        df = pd.read_excel(caminho_arquivo)

        # Garante que as datas estejam em datetime
        df['Fim da Vigência'] = pd.to_datetime(df['Fim da Vigência'], errors='coerce', dayfirst=True)

        # Cria a nova coluna inicialmente vazia
        df['Situação 300 dias'] = ''

        # Preenchimentos diretos
        df.loc[df['Situação RCO'] == 'Comprovado', 'Situação 300 dias'] = 'Comprovado'
        df.loc[df['Situação RCO'] == 'Vigente', 'Situação 300 dias'] = 'Vigente'

        # Define data atual
        hoje = datetime.now()

        # Filtra os que ainda estão vazios
        filtro_vazio = df['Situação 300 dias'] == ''

        # Soma 300 dias à Fim da Vigência
        fim_vigencia_mais_300 = df.loc[filtro_vazio, 'Fim da Vigência'] + pd.to_timedelta(300, unit='d')

        # Define condição de vencido ou não
        vencido_mais_300 = fim_vigencia_mais_300 < hoje

        df.loc[filtro_vazio & vencido_mais_300, 'Situação 300 dias'] = 'Vencido + 300'
        df.loc[filtro_vazio & ~vencido_mais_300, 'Situação 300 dias'] = 'Vencido - 300'

        # Salva de volta
        df.to_excel(caminho_arquivo, index=False)

        print("Coluna 'Situação 300 dias' preenchida com sucesso.")

    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna: {e}")

def preencher_comparacoes_em_lote(
    arquivo_alvo,
    coluna_destino,
    comparacoes,
    linha_cabecalho_origem=0
):
    # Carrega o arquivo de origem uma única vez
    df_origem = pd.read_excel(arquivo_alvo, header=linha_cabecalho_origem)
    df_origem[coluna_destino] = df_origem[coluna_destino].astype(str)

    for comparacao in comparacoes:
        arquivo_referencia = comparacao["arquivo_referencia"]
        aba_referencia = comparacao["aba_referencia"]
        coluna_referencia = comparacao["coluna_referencia"]
        nova_coluna = comparacao["nova_coluna"]
        linha_cabecalho_referencia = comparacao.get("linha_cabecalho_referencia", 0)

        # Carrega a aba de referência
        df_referencia = pd.read_excel(
            arquivo_referencia,
            sheet_name=aba_referencia,
            header=linha_cabecalho_referencia
        )
        df_referencia[coluna_referencia] = df_referencia[coluna_referencia].astype(str)

        # Pega os valores únicos
        valores_referencia = df_referencia[coluna_referencia].unique()

        # Compara e preenche
        df_origem[nova_coluna] = df_origem[coluna_destino].apply(
            lambda x: "Sim" if x in valores_referencia else "Não"
        )

        print(f"✅ Coluna '{nova_coluna}' preenchida com base na aba '{aba_referencia}'")

    # Salva tudo de uma vez, sem sobrescrever colunas anteriores
    df_origem.to_excel(arquivo_alvo, index=False)
    print("✅ Todas as comparações foram salvas no arquivo com sucesso.")

def preencher_coluna_auditoria(arquivo_excel, nome_coluna_resultado='Auditoria'):
    try:
        # Carrega o Excel
        df = pd.read_excel(arquivo_excel)
        df.columns = df.columns.str.strip()  # remove espaços em branco dos nomes

        # Colunas a verificar
        colunas_verificar = ['Auditoria 2022', 'Auditoria 2023', 'Auditoria 2024 até jan/2024']

        # Apaga a coluna 'Auditoria' se já existir
        if nome_coluna_resultado in df.columns:
            df.drop(columns=[nome_coluna_resultado], inplace=True)
            print(f"🧹 Coluna '{nome_coluna_resultado}' apagada antes de recriar.")

        # Passo 1: se todas forem "Não", então 'Auditoria' = "Não"
        df[nome_coluna_resultado] = df[colunas_verificar].apply(
            lambda row: "Não" if all(str(val).strip().lower() == 'não' for val in row) else "", axis=1
        )

        # Passo 2: preenche os valores em branco com "Sim"
        df[nome_coluna_resultado] = df[nome_coluna_resultado].replace("", "Sim")

        # Estatísticas
        print("\n📊 Quantidade de 'Sim' em cada coluna de auditoria:")
        for col in colunas_verificar:
            count = df[col].astype(str).str.strip().str.lower().eq('sim').sum()
            print(f" - {col}: {count} linhas com 'Sim'")

        total_sim = df[nome_coluna_resultado].str.lower().eq("sim").sum()
        total_nao = df[nome_coluna_resultado].str.lower().eq("não").sum()
        print(f"\n✅ Total de linhas com 'Sim' na coluna '{nome_coluna_resultado}': {total_sim}")
        print(f"❌ Total de linhas com 'Não': {total_nao}")
        print(f"📦 Total geral de linhas: {len(df)}")

        # Salva no mesmo arquivo sobrescrevendo
        with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False)

        print(f"\n✅ Coluna '{nome_coluna_resultado}' preenchida com sucesso.")
    except Exception as e:
        print(f"❌ Erro ao preencher a coluna '{nome_coluna_resultado}': {e}")
        
def preencher_vencidos_ate_fev_2024(arquivo_excel, nome_coluna_data='Fim da Vigência', nome_coluna_resultado='Vencidos até Fev/2024'):
    try:
        #Carregar o DataFrame
        df = pd.read_excel(arquivo_excel)

        #Converter a coluna de datas para datetime
        df[nome_coluna_data] = pd.to_datetime(df[nome_coluna_data], errors='coerce')

         #Definir a data limite (sem horário)
        data_limite = pd.to_datetime('2024-02-29').date()

         #Aplicar apenas a data (ignorar hora)
        df[nome_coluna_resultado] = df[nome_coluna_data].dt.date.apply(
            lambda x: 'Sim' if pd.notnull(x) and x <= data_limite else 'Não'
        )

         #Salvar de volta no mesmo arquivo
        with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False)

        print(f"✅ Coluna '{nome_coluna_resultado}' preenchida corretamente.")
    except Exception as e:
        print(f"❌ Ocorreu um erro ao preencher a coluna '{nome_coluna_resultado}': {e}")


def substituir_codigos_concedente(arquivo_excel, nome_coluna='Descentralizadora', nome_aba=0):
    try:
        # Mapeamento dos códigos para siglas
        mapeamento = {
            '150002': 'SGA',
            '150004': 'STIC',
            '150011': 'SESU',
            '150003': 'SE',
            '150016': 'SETEC',
            '150019': 'SEB',
            '150028': 'SECADI',
            '152389': 'SASE',
            '152390': 'SERES',
            '156570': 'SECADI',
            '156575': 'SEB',
            '157054': 'SASE',
            '157055': 'SECADI'
        }

        # Ler o Excel
        df = pd.read_excel(arquivo_excel, sheet_name=nome_aba)

        # Função para extrair e substituir o código
        def extrair_sigla(valor):
            if pd.isna(valor):
                return valor
            valor_str = str(valor).strip()
            codigo = valor_str.split('/')[0].strip()
            return mapeamento.get(codigo, valor_str)

        # Aplicar a função na coluna
        df[nome_coluna] = df[nome_coluna].apply(extrair_sigla)

        # Salvar de volta
        with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=nome_aba, index=False)

        print("Descentralizadoras substituídos com sucesso.")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def reordenar_e_renomear_colunas(df):
    # Renomeia as colunas
    df = df.rename(columns={
        'Descentralizadora': 'Unidade Descentralizadora',
        'Descentralizada': 'Unidade Descentralizada',
        'Descrição do Termo': 'Descrição do Objeto'
    })

    # Define a nova ordem desejada
    nova_ordem = [
        'TED',
        'SIAFI',
        'Unidade Descentralizadora',
        'Unidade Descentralizada',
        'Coordenação',
        'Descrição do Objeto',
        'Estado Atual',
        'Situação RCO',
        'Início da Vigência',
        'Fim da Vigência',
        'Situação 300 dias',
        'Data da última tramitação',
        'Ano',
        'Total Descentralizado (R$)',
        'Total Repassado',
        'Auditoria 2022',
        'Auditoria 2023',
        'Auditoria 2024 até jan/2024',
        'Auditoria'
    ]

    # Garante que apenas as colunas existentes serão reordenadas
    colunas_existentes = [col for col in nova_ordem if col in df.columns]
    
    return df[colunas_existentes]

def formatar_contabil(value):
    if pd.notnull(value):
        if isinstance(value, (int, float)):
            return "{:,.2f}".format(float(value)).replace(",", "_").replace(".", ",").replace("_", ".")
        else:
            return value  # Mantém o cabeçalho ou outros valores não numéricos
    return None

def formatar_data(data):
    if pd.notnull(data):
        # Verifique se a data não é '-' antes de tentar formatá-la
        if data != '-':
            # Converte a data para o formato desejado (DD/MM/AAAA)
            data_formatada = datetime.strptime(str(data), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
            return data_formatada
    return None

def copiar_dados_sem_cabecalho_para_sintese():
    # Caminhos dos arquivos
    copia_teds_auditoria = r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA TEDS Auditoria.xlsx'
    copia_sintese_auditorias = r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx'

    # Lê os dados da planilha fonte, ignorando o cabeçalho
    df = pd.read_excel(copia_teds_auditoria, header=0)  # lê com cabeçalho
    dados_sem_cabecalho = df.values.tolist()  # remove cabeçalho para escrita

    # Carrega o arquivo de destino com openpyxl
    wb = load_workbook(copia_sintese_auditorias)
    
    # Seleciona a aba 'TEDs GERAL'
    if 'TEDs GERAL' not in wb.sheetnames:
        raise ValueError("A aba 'TEDs GERAL' não foi encontrada no arquivo de destino.")
    ws = wb['TEDs GERAL']

    # Começa a escrever na linha 6 (linha 5 é o cabeçalho)
    start_row = 6
    for i, linha in enumerate(dados_sem_cabecalho, start=start_row):
        for j, valor in enumerate(linha, start=1):  # colunas começam em 1 no openpyxl
            ws.cell(row=i, column=j, value=valor)

    # Salva sobrescrevendo o arquivo de destino
    wb.save(copia_sintese_auditorias)

def atualizar_colunas_por_parametro(
arquivo_excel: str,
    aba_origem: str,
    aba_destino: str,
    coluna_chave_origem: str,
    coluna_chave_destino: str,
    colunas_origem: list,
    colunas_destino: list
):
    try:
        # Ler abas como DataFrame bruto sem cabeçalho
        df_origem_raw = pd.read_excel(arquivo_excel, sheet_name=aba_origem, header=None)
        df_destino_raw = pd.read_excel(arquivo_excel, sheet_name=aba_destino, header=None)

        def detectar_header(df):
            for i in range(min(10, len(df))):
                if df.iloc[i].notna().sum() >= 3:
                    return i
            raise ValueError("Cabeçalho não encontrado.")

        header_origem = detectar_header(df_origem_raw)
        header_destino = detectar_header(df_destino_raw)

        # Criar DataFrames com cabeçalho real
        df_origem = df_origem_raw.iloc[header_origem+1:].copy()
        df_origem.columns = df_origem_raw.iloc[header_origem].str.strip()

        df_destino = df_destino_raw.iloc[header_destino+1:].copy()
        df_destino.columns = df_destino_raw.iloc[header_destino].str.strip()

        # Criar dicionários para mapeamento de colunas
        for col_o, col_d in zip(colunas_origem, colunas_destino):
            dicionario = pd.Series(df_origem[col_o.strip()].values, index=df_origem[coluna_chave_origem.strip()]).to_dict()
            df_destino[col_d.strip()] = df_destino[coluna_chave_destino.strip()].map(dicionario)

        # Carregar planilha com openpyxl
        wb = load_workbook(arquivo_excel)
        ws = wb[aba_destino]

        # Localizar cabeçalho no openpyxl (linhas reais)
        header_row = header_destino + 1  # +1 porque openpyxl é 1-based
        headers = [cell.value for cell in ws[header_row]]

        # Identificar colunas pelo nome
        col_indices = {}
        for col_d in colunas_destino:
            col_name = col_d.strip()
            if col_name in headers:
                col_idx = headers.index(col_name) + 1  # +1 para openpyxl
                col_indices[col_name] = col_idx
            else:
                raise ValueError(f"Coluna '{col_name}' não encontrada na aba destino.")

        # Preencher os dados mantendo a formatação
        for row_idx, (_, row) in enumerate(df_destino.iterrows(), start=header_row + 1):
            chave_valor = row[coluna_chave_destino.strip()]
            for col_d in colunas_destino:
                valor = row[col_d.strip()]
                ws.cell(row=row_idx, column=col_indices[col_d.strip()], value=valor)

        # Salvar mantendo formatação
        wb.save(arquivo_excel)
        print("Colunas atualizadas com sucesso, mantendo a formatação!")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")  
        
def main():
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria'
    copia_teds_auditoria = r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA TEDS Auditoria.xlsx'
    copia_relatorio_analitico_DSV = r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Relatorio Analitico DSV.xlsx'
    copia_sintese_auditorias = r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx'
    
    copiar_arquivos(destino)
    remover_linhas_indesejadas(copia_teds_auditoria)
    
    aba_extrato = "Sheet1"
    novas_colunas = ["Situação RCO", "Situação 300 dias", "Vencidos até Fev/2024", "Data da última tramitação", "Ano", "Auditoria 2022", "Auditoria 2023", "Auditoria 2024 até jan/2024", "Auditoria"]
    
    adicionar_colunas_excel(copia_teds_auditoria, aba_extrato, novas_colunas, valor_padrao=None, salvar_como=copia_teds_auditoria)
    
    preencher_colunas(
        copia_teds_auditoria=copia_teds_auditoria,
        copia_relatorio_analitico_DSV=copia_relatorio_analitico_DSV,
        coluna_valores='Data última tramitação',
        coluna_destino='Data da última tramitação',
        chave_coluna= 'TED'
    )

    preencher_coluna_ano(caminho_arquivo=copia_teds_auditoria)
    preencher_situacao_rco(caminho_arquivo=copia_teds_auditoria)
    preencher_situacao_300_dias(copia_teds_auditoria)
    
    preencher_comparacoes_em_lote(
    arquivo_alvo=r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA TEDS Auditoria.xlsx",
    coluna_destino="TED",
    comparacoes=[
        {
            "arquivo_referencia": r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx",
            "aba_referencia": "Planilha Auditoria 2022",
            "coluna_referencia": "TED",
            "nova_coluna": "Auditoria 2022",
            "linha_cabecalho_referencia": 6
        },
        {
            "arquivo_referencia": r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx",
            "aba_referencia": "Planilha Auditoria 2023",
            "coluna_referencia": "TED",
            "nova_coluna": "Auditoria 2023",
            "linha_cabecalho_referencia": 5
        },
        {
            "arquivo_referencia": r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx",
            "aba_referencia": "Planilha Auditoria 2024",
            "coluna_referencia": "TED",
            "nova_coluna": "Auditoria 2024 até jan/2024",
            "linha_cabecalho_referencia": 5
        }
    ],
    linha_cabecalho_origem=0  # Ajuste conforme necessário
)
    preencher_coluna_auditoria(copia_teds_auditoria)
    #preencher_vencidos_ate_fev_2024(copia_teds_auditoria)
    
    substituir_codigos_concedente(
    arquivo_excel=r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA TEDS Auditoria.xlsx",
    nome_coluna="Descentralizadora",
    nome_aba="Sheet1"  # Ou passe o nome da aba, como "TED"
)
    
    # Lê, reordena e sobrescreve o arquivo
    df = pd.read_excel(copia_teds_auditoria)
    df_reordenado = reordenar_e_renomear_colunas(df)
    df_reordenado.to_excel(copia_teds_auditoria, index=False)
    
    copiar_dados_sem_cabecalho_para_sintese()
    print("Cópia concluída com sucesso.")    
    
    atualizar_colunas_por_parametro(
    arquivo_excel=r'W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx',
    aba_origem="TEDs GERAL",
    aba_destino="Planilha Auditoria 2024",
    coluna_chave_origem="TED",          # Nome da coluna na aba origem para comparar
    coluna_chave_destino="TED",         # Nome da coluna na aba destino para comparar
    colunas_origem=["Estado Atual ", "Situação RCO", "Data da última tramitação "],     # Nomes na aba de origem
    colunas_destino=["Situação do TED SIMEC", "Situação RCO", "Data da útima tramitação"]  # Nomes nas colunas da aba de destino
)
    atualizar_colunas_por_parametro(
    arquivo_excel=r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx",
    aba_origem="TEDs GERAL",
    aba_destino="Planilha Auditoria 2023",
    coluna_chave_origem="TED",          # Nome da coluna na aba origem para comparar
    coluna_chave_destino="TED",         # Nome da coluna na aba destino para comparar
    colunas_origem=["Estado Atual ", "Situação RCO", "Data da última tramitação "],     # Nomes na aba de origem
    colunas_destino=["Situação do TED no SIMEC", "Situação do RCO", "Data da útima tramitação"]  # Nomes nas colunas da aba de destino
)
    atualizar_colunas_por_parametro(
    arquivo_excel=r"W:\B - TED\7 - AUTOMAÇÃO\Auditoria\COPIA Síntese RCO e Auditorias.xlsx",
    aba_origem="TEDs GERAL",
    aba_destino="Planilha Auditoria 2022",
    coluna_chave_origem="TED",          # Nome da coluna na aba origem para comparar
    coluna_chave_destino="TED",         # Nome da coluna na aba destino para comparar
    colunas_origem=["Estado Atual ", "Situação RCO", "Data da última tramitação "],     # Nomes na aba de origem
    colunas_destino=["Situação do TED no SIMEC", "Situação do RCO", "Data da última tramitação SIMEC"]  # Nomes nas colunas da aba de destino
)
    
    copiar_dados_sem_cabecalho_para_sintese()
    print("Processo finalizado com sucesso!")
if __name__ == "__main__":
    main()        

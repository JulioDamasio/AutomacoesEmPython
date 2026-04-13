# -*- coding: utf-8 -*-

import os
import shutil
import locale
import re
from copy import copy
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata
from decimal import Decimal, InvalidOperation

print("Iniciando Processamento aguarde...")

# =============================================================================
# CONFIGURAÇÕES GERAIS
# =============================================================================

BASE_DIR = r"W:\B - TED\7 - AUTOMAÇÃO\Liquidação"

ARQUIVO_EXECUCAO_SIMEC_ORIG = os.path.join(BASE_DIR, "Execução SIMEC DOC NE.xlsx")
ARQUIVO_EXECUCAO_SIMEC_COPIA = os.path.join(BASE_DIR, "COPIA Execução SIMEC DOC NE.xlsx")

ARQUIVO_TED_GERAL_ORIG = os.path.join(BASE_DIR, "TED Liquidação Geral - EXERCÍCIO 2026.xlsx")
ARQUIVO_TED_GERAL_COPIA = os.path.join(BASE_DIR, "COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx")

ARQUIVO_LIMITE_SAQUE_ORIG = os.path.join(BASE_DIR, "Limite de Saque MEC.xlsx")
ARQUIVO_LIMITE_SAQUE_COPIA = os.path.join(BASE_DIR, "COPIA Limite de Saque MEC.xlsx")

ARQUIVO_CONTAS_CONTROLE_ORIG = os.path.join(BASE_DIR, "TED - Contas Cadastro e Controle.xlsx")
ARQUIVO_CONTAS_CONTROLE_COPIA = os.path.join(BASE_DIR, "COPIA TED - Contas Cadastro e Controle.xlsx")

ARQUIVO_NCPF_ORIG = os.path.join(BASE_DIR, "NCPF.xlsx")
ARQUIVO_NCPF_COPIA = os.path.join(BASE_DIR, "COPIA NCPF.xlsx")

ARQUIVO_FINAL_REPASSAR = os.path.join(BASE_DIR, "LIQUIDAÇÃO A REPASSAR.xlsx")

DESTINO_FINANCEIRO_2026 = r"W:\B - TED\4 - FINANCEIRO\4.1 - LIQUIDAÇÃO\2026"

SHEET_PRINCIPAL = "Sheet1"
SHEET_DOC_NE = "DOC NE"
SHEET_TABELA = "Tabela"
SHEET_REPASSAR = "REPASSAR"
SHEET_LIMITE_SAQUE = "Limite Saque"
SHEET_CONTAS_CONTROLE = "Contas Cadastro e Controle"
SHEET_NC_PF = "NC-PF"
SHEET_EMENDAS = "EMENDAS"

MESES_PTBR = {
    1: "01-Janeiro",
    2: "02-Fevereiro",
    3: "03-Março",
    4: "04-Abril",
    5: "05-Maio",
    6: "06-Junho",
    7: "07-Julho",
    8: "08-Agosto",
    9: "09-Setembro",
    10: "10-Outubro",
    11: "11-Novembro",
    12: "12-Dezembro",
}

# =============================================================================
# UTILITÁRIOS
# =============================================================================

def log(msg: str) -> None:
    print(msg)

def copiar_arquivo(origem: str, destino: str) -> None:
    shutil.copy2(origem, destino)

def obter_data_atual() -> datetime.date:
    return datetime.now().date()

def garantir_locale_ptbr() -> None:
    try:
        locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")
    except locale.Error:
        log("Aviso: locale pt_BR.UTF-8 não disponível neste ambiente. Continuando sem locale.")

def obter_ou_criar_aba(wb, nome_aba: str):
    if nome_aba in wb.sheetnames:
        return wb[nome_aba]
    return wb.create_sheet(nome_aba)

def recriar_aba(wb, nome_aba: str):
    if nome_aba in wb.sheetnames:
        wb.remove(wb[nome_aba])
    return wb.create_sheet(nome_aba)

def copiar_aba_por_valores(sheet_origem, sheet_destino) -> None:
    for linha in sheet_origem.iter_rows(values_only=True):
        sheet_destino.append(list(linha))

def desmesclar_todas_as_abas(workbook) -> None:
    for sheet in workbook.worksheets:
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))

def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()

def formatar_gestao_emitente(gestao):
    if pd.notnull(gestao):
        gestao_str = normalizar_texto(gestao)
        if gestao_str == "":
            return ""
        if gestao_str.isdigit():
            return gestao_str.zfill(5)
        try:
            return str(int(float(gestao_str))).zfill(5)
        except (ValueError, TypeError):
            return gestao_str
    return ""

def remove_pontos_zeros(valor):
    if pd.notnull(valor):
        valor_str = normalizar_texto(valor)
        if valor_str == "":
            return ""
        if valor_str.isdigit():
            return valor_str
        try:
            return str(int(float(valor_str))).rstrip(".")
        except (ValueError, TypeError):
            return valor_str
    return ""

def formatar_contabil(value):
    if pd.notnull(value):
        if isinstance(value, (int, float)):
            return "{:,.2f}".format(float(value)).replace(",", "_").replace(".", ",").replace("_", ".")
        return value
    return None

def parse_numero_br(valor):
    """
    Converte valores brasileiros para float.
    Exemplos aceitos:
    '0,00'
    '1.234,56'
    'R$ 1.234,56'
    '  1.234,56  '
    1234.56
    0
    """
    if valor is None:
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if texto == "":
        return None

    # remove quebras, NBSP, etc.
    texto = unicodedata.normalize("NFKC", texto)
    texto = texto.replace("\xa0", " ")
    texto = texto.replace("R$", "")
    texto = texto.replace(" ", "")

    # mantém só dígitos, vírgula, ponto e sinal
    texto = re.sub(r"[^0-9,.\-]", "", texto)

    if texto in ("", "-", "--"):
        return None

    # formato BR: 1.234,56 -> 1234.56
    if "," in texto:
        texto = texto.replace(".", "")
        texto = texto.replace(",", ".")

    try:
        return float(Decimal(texto))
    except (InvalidOperation, ValueError):
        return None

def formatar_data(data):
    if pd.isnull(data):
        return None

    data = str(data).strip()

    if data in ("", "-"):
        return None

    for formato in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(data, formato).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return data

def substituir_valores_grupo_despesa(value):
    if pd.notnull(value):
        try:
            value_int = int(float(value))
            mapa = {3: "C", 4: "D", 5: "E"}
            return mapa.get(value_int, value)
        except (ValueError, TypeError):
            return value
    return value

def remover_quebras_de_linha_texto(valor):
    if isinstance(valor, str):
        return valor.replace("\n", " ").strip()
    return valor

def obter_mes_atual_formatado():
    data_atual = datetime.now()
    return data_atual, MESES_PTBR[data_atual.month]

def linha_deve_ser_descartada_por_zerada(row):
    """
    No arquivo antigo em .xls havia um comportamento em que a primeira linha útil às vezes vinha zerada.
    Agora, com .xlsx, isso nem sempre acontece. Então descartamos só se a linha estiver realmente vazia/zerada.
    """
    if row is None or len(row) == 0:
        return False

    valores = [v for v in row.tolist() if pd.notnull(v)]
    if not valores:
        return False

    marcadores_zerados = {"0", "0,00", "0.00", "0,0", "0.0", 0, 0.0}
    return all(v in marcadores_zerados for v in valores)

# =============================================================================
# ETAPA 1 - EXECUÇÃO SIMEC
# =============================================================================

def processar_exeSimec():
    log("Processando Execução SIMEC DOC NE...")

    copiar_arquivo(ARQUIVO_EXECUCAO_SIMEC_ORIG, ARQUIVO_EXECUCAO_SIMEC_COPIA)

    # Lê preservando o que vier como texto, pois agora o .xlsx já vem formatado visualmente.
    df = pd.read_excel(ARQUIVO_EXECUCAO_SIMEC_COPIA, dtype=str)

    # Só remove a primeira linha se ela realmente for uma linha zerada/lixo do arquivo antigo.
    if not df.empty and linha_deve_ser_descartada_por_zerada(df.iloc[0]):
        df = df.drop(df.index[0]).reset_index(drop=True)

    col_gestao = "Gestão Emitente - NE"
    col_numero_empenho = "Número do Empenho"
    col_ug_emitente = "UG Executora Emitente - NE"
    col_vigencia = "Fim da Vigência"
    col_valor_ne = "Valor da NE"

    df[col_gestao] = df[col_gestao].apply(formatar_gestao_emitente)
    df[col_ug_emitente] = df[col_ug_emitente].apply(remove_pontos_zeros)
    df[col_numero_empenho] = df[col_numero_empenho].astype(str).str.strip()

    # Mantém o número do empenho exatamente como veio, apenas limpando espaços.
    # O campo Dados do Empenho continua na mesma lógica do script anterior.
    df["Dados do Empenho"] = (
        df[col_ug_emitente].astype(str).str.strip()
        + df[col_gestao].astype(str).str.strip().str.zfill(5)
        + df[col_numero_empenho].astype(str).str.strip()
    ).str.replace(".", "", regex=False)

    coluna_dados_empenho = df.pop("Dados do Empenho")
    df.insert(5, "Dados do Empenho", coluna_dados_empenho)

    # Corrige o problema principal do novo .xlsx:
    # Valor da NE vem como texto brasileiro, então não pode usar pd.to_numeric direto.
    if col_valor_ne in df.columns:
        df[col_valor_ne] = df[col_valor_ne].apply(parse_numero_br)
        df[col_valor_ne] = df[col_valor_ne].apply(formatar_contabil)

    if col_vigencia in df.columns:
        df[col_vigencia] = df[col_vigencia].apply(formatar_data)

    df.to_excel(ARQUIVO_EXECUCAO_SIMEC_COPIA, index=False)
    log("Execução SIMEC processada com sucesso.")

# =============================================================================
# ETAPA 2 - TED LIQUIDAÇÃO GERAL
# =============================================================================

def processar_arquivo_ted_geral():
    log("Processando TED Liquidação Geral...")
    garantir_locale_ptbr()

    copiar_arquivo(ARQUIVO_TED_GERAL_ORIG, ARQUIVO_TED_GERAL_COPIA)

    df = pd.read_excel(ARQUIVO_TED_GERAL_COPIA, header=None, dtype=object)
    log("Carregou o arquivo de cópia em um DataFrame.")

    df = df.iloc[12:].copy()
    log("Excluiu as linhas de 1 a 12.")

    # As últimas colunas podem vir como texto formatado em pt-BR.
    for coluna in df.columns[-4:]:
        df[coluna] = df[coluna].apply(parse_numero_br)
    log("Converteu as quatro últimas colunas em números.")

    for coluna in df.columns[-4:]:
        df[coluna] = df[coluna].apply(formatar_contabil)
    log("Formatou as quatro últimas colunas em formato contábil.")

    df.iloc[:, 13] = df.iloc[:, 13].apply(substituir_valores_grupo_despesa)
    log("Substituiu os valores da coluna Grupo Despesa.")

    df["SITUAÇÃO"] = ""

    def determinar_situacao(row):
        ano_atual = datetime.now().year
        valor_coluna = row.iloc[2]
        if pd.notnull(valor_coluna):
            try:
                if int(float(str(valor_coluna).strip())) == ano_atual:
                    return "TRF003"
            except (ValueError, TypeError):
                pass
        return "TRF004"

    df["SITUAÇÃO"] = df.apply(determinar_situacao, axis=1)
    log("Preencheu a coluna SITUAÇÃO.")

    nomes_colunas = [
        "Resultado EOF",
        "DESCRIÇÃO EOF",
        "NE CCor - Ano Emissão",
        "Órgão UGE",
        "DESCRIÇÃO UGE",
        "UG Executora",
        "DESCRIÇÃO EXECUTORA",
        "UGE - UG Setorial Financeira",
        "DESCRIÇÃO FINANCEIRA",
        "Ação Governo",
        "PTRES",
        "PI",
        "NE CCor",
        "Grupo Despesa",
        "Natureza Despesa Detalhada",
        "NATUREZA",
        "Elemento Despesa",
        "ND",
        "Fonte Recursos Detalhada",
        "DESPESAS LIQUIDADAS A PAGAR(CONTROLE EMPENHO)",
        "RESTOS A PAGAR PROCESSADOS A PAGAR",
        "RESTOS A PAGAR NAO PROCES. LIQUIDADOS A PAGAR",
        "Total",
        "SITUAÇÃO",
    ]

    if len(df.columns) >= len(nomes_colunas):
        df.iloc[0] = nomes_colunas
    log("Renomeou o cabeçalho com os nomes das colunas.")

    # Remove linhas cujo Total seja efetivamente zero.
    totais_numericos = df[22].apply(parse_numero_br)
    df = df[(totais_numericos.isna()) | (totais_numericos != 0)]
    log("Removeu as linhas em que o valor da coluna Total é zero.")

    df.to_excel(ARQUIVO_TED_GERAL_COPIA, index=False, header=False)
    log("Reescreveu o arquivo de cópia com as alterações.")

def excluir_linhas_total_zero(arquivo_copia):
    log("Excluindo linhas com Total igual a 0,00...")
    df = pd.read_excel(arquivo_copia, header=None, dtype=object)
    total_num = df[22].apply(parse_numero_br)
    df = df[(total_num.isna()) | (total_num != 0)]
    df.to_excel(arquivo_copia, index=False, header=False)

# =============================================================================
# ETAPA 3 - COPIAR DOC NE PARA O ARQUIVO PRINCIPAL
# =============================================================================

def copyData():
    log("Adicionando colunas auxiliares e copiando DOC NE...")

    wb = openpyxl.load_workbook(ARQUIVO_TED_GERAL_COPIA)
    planilha_principal = wb[SHEET_PRINCIPAL]

    colunas = ["TED", "SIAFI", "Vigência", "Estado Atual", "Vinculação Pagamento"]
    colunas_excel = ["Y", "Z", "AA", "AB", "AC"]

    for coluna_excel, titulo in zip(colunas_excel, colunas):
        planilha_principal[f"{coluna_excel}1"] = titulo

    planilha_doc_ne = recriar_aba(wb, SHEET_DOC_NE)

    wb_doc_ne = openpyxl.load_workbook(ARQUIVO_EXECUCAO_SIMEC_COPIA)
    planilha_dados_doc_ne = wb_doc_ne.active

    copiar_aba_por_valores(planilha_dados_doc_ne, planilha_doc_ne)

    wb.save(ARQUIVO_TED_GERAL_COPIA)
    wb.close()
    wb_doc_ne.close()

    log("Aba DOC NE criada e dados copiados com sucesso.")

def preencher_coluna_vinculacao_pagamento(planilha_principal):
    log("Preenchendo a coluna Vinculação Pagamento...")

    mapa = {
        "2": "400",
        "3": "415",
        "9": "409",
        "6": "405",
        "7": "410",
        "8": "408",
    }

    for linha in planilha_principal.iter_rows(min_row=2):
        resultado_eof = linha[0].value
        if resultado_eof is not None:
            resultado_eof = str(resultado_eof).strip()
            linha[28].value = mapa.get(resultado_eof, linha[28].value)

def preencher_colunas_siafi_vigencia_estado_atual(arquivo_principal, data_atual):
    log("Preenchendo colunas TED, SIAFI, Vigência, Estado Atual e Vinculação Pagamento...")

    wb_principal = openpyxl.load_workbook(arquivo_principal)
    planilha_principal = wb_principal[SHEET_PRINCIPAL]
    planilha_doc_ne = wb_principal[SHEET_DOC_NE]

    ted_dict = {}
    siafi_dict = {}
    vigencia_dict = {}
    estado_atual_dict = {}

    for linha in planilha_doc_ne.iter_rows(min_row=2):
        ne_ccor = linha[5].value           # Dados do Empenho
        estado_atual = linha[6].value      # Estado Atual
        vigencia = linha[7].value          # Fim da Vigência
        siafi = linha[8].value             # SIAFI
        ted = linha[9].value               # TED

        if ne_ccor:
            ted_dict[str(ne_ccor).strip()] = ted
            siafi_dict[str(ne_ccor).strip()] = siafi
            vigencia_dict[str(ne_ccor).strip()] = vigencia
            estado_atual_dict[str(ne_ccor).strip()] = estado_atual

    red_fill = PatternFill(start_color="FFFA5E46", end_color="FFFA5E46", fill_type="solid")
    green_fill = PatternFill(start_color="FF78E6A3", end_color="FF78E6A3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")

    for linha in planilha_principal.iter_rows(min_row=2):
        ne_ccor_sheet1 = linha[12].value
        if ne_ccor_sheet1 is None:
            continue

        chave_ne = str(ne_ccor_sheet1).strip()

        if chave_ne in ted_dict:
            linha[24].value = ted_dict[chave_ne]

        if chave_ne in siafi_dict:
            linha[25].value = siafi_dict[chave_ne]

        if chave_ne in vigencia_dict:
            valor_vigencia = vigencia_dict[chave_ne]
            valor_vigencia_formatado = formatar_data(valor_vigencia)

            if isinstance(valor_vigencia_formatado, str):
                try:
                    vigencia = datetime.strptime(valor_vigencia_formatado, "%d/%m/%Y").date()
                    linha[26].value = vigencia
                    if vigencia < data_atual:
                        linha[26].fill = red_fill
                    elif vigencia > data_atual:
                        linha[26].fill = green_fill
                except ValueError:
                    linha[26].value = valor_vigencia_formatado
                    linha[26].fill = yellow_fill
            else:
                linha[26].value = valor_vigencia_formatado

        if chave_ne in estado_atual_dict:
            linha[27].value = estado_atual_dict[chave_ne]

    preencher_coluna_vinculacao_pagamento(planilha_principal)

    wb_principal.save(arquivo_principal)
    wb_principal.close()
    log("Alterações salvas no arquivo principal.")

# =============================================================================
# ETAPA 4 - TABELA
# =============================================================================

def criar_tabela(arquivo_principal):
    log("Criando aba Tabela...")

    wb = load_workbook(arquivo_principal)
    planilha_tabela = recriar_aba(wb, SHEET_TABELA)
    planilha_principal = wb[SHEET_PRINCIPAL]

    colunas_selecionadas = [
        "Resultado EOF",
        "Vinculação Pagamento",
        "UGE - UG Setorial Financeira",
        "UG Executora",
        "TED",
        "Ação Governo",
        "Vigência",
        "SITUAÇÃO",
        "SIAFI",
        "Fonte Recursos Detalhada",
        "Grupo Despesa",
        "Total",
    ]

    df = pd.DataFrame(planilha_principal.values)
    df.columns = df.iloc[0]
    df = df[1:]

    df = df[colunas_selecionadas].copy()

    data_atual = datetime.now().date()
    df["Vigência"] = pd.to_datetime(df["Vigência"], format="%d/%m/%Y", errors="coerce").dt.date
    df_filtrado = df[(df["Vigência"].notnull()) & (df["Vigência"] > data_atual)]

    for row in dataframe_to_rows(df_filtrado, index=False, header=True):
        planilha_tabela.append(row)

    wb.save(arquivo_principal)
    wb.close()

    log(f"Tabela criada com sucesso na aba '{SHEET_TABELA}'.")

def converter_coluna_total_para_numero(arquivo_principal):
    log("Convertendo a coluna Total para número na aba Tabela...")

    wb = openpyxl.load_workbook(arquivo_principal)
    sheet = wb[SHEET_TABELA]

    for cell in sheet["L"][1:]:
        if cell.value is not None:
            valor = parse_numero_br(cell.value)
            if valor is not None:
                cell.value = valor

    wb.save(arquivo_principal)
    wb.close()

def agrupar_e_somar_total(arquivo_principal):
    log("Agrupando e somando Totais na aba Tabela...")

    wb = load_workbook(arquivo_principal)
    planilha_tabela = wb[SHEET_TABELA]

    data = planilha_tabela.values
    colunas = next(data)
    df = pd.DataFrame(data, columns=colunas)

    colunas_para_agrupar = [
        "Resultado EOF",
        "Vinculação Pagamento",
        "UGE - UG Setorial Financeira",
        "UG Executora",
        "TED",
        "Ação Governo",
        "Vigência",
        "SITUAÇÃO",
        "SIAFI",
        "Fonte Recursos Detalhada",
        "Grupo Despesa",
    ]

    df["Total"] = df["Total"].apply(parse_numero_br)
    df_agrupado = df.groupby(colunas_para_agrupar, as_index=False)["Total"].sum()

    planilha_tabela.delete_rows(1, planilha_tabela.max_row)

    for row in dataframe_to_rows(df_agrupado, index=False, header=True):
        planilha_tabela.append(row)

    for row_idx in range(2, planilha_tabela.max_row + 1):
        cell_total = planilha_tabela.cell(row=row_idx, column=12)
        if isinstance(cell_total.value, (int, float)):
            cell_total.value = formatar_contabil(cell_total.value)

    wb.save(arquivo_principal)
    wb.close()

    log("Agrupamento, soma e formatação contábil dos Totais concluídos.")

# =============================================================================
# ETAPA 5 - LIMITE DE SAQUE
# =============================================================================

def processar_limite_saque(caminho_arquivo_original):
    log("Processando Limite de Saque...")

    copiar_arquivo(caminho_arquivo_original, ARQUIVO_LIMITE_SAQUE_COPIA)

    df = pd.read_excel(ARQUIVO_LIMITE_SAQUE_COPIA, dtype=object)

    if not df.empty:
        df = df.drop(df.index[-1])

    df.columns = [
        "Órgão UGE",
        "Descrição UGE",
        "UG Executora",
        "Descrição UG",
        "Órgão UGE - Gestão",
        "Vinculação Pagamento",
        "Fonte Recursos Detalhada",
        "LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)",
    ]

    df["UG Executora"] = df["UG Executora"].apply(remove_pontos_zeros).apply(formatar_gestao_emitente)
    df["Vinculação Pagamento"] = df["Vinculação Pagamento"].apply(remove_pontos_zeros).astype(str).fillna("")
    df["UG Executora"] = df["UG Executora"].astype(str).fillna("")
    df["Fonte Recursos Detalhada"] = df["Fonte Recursos Detalhada"].astype(str).fillna("")

    df["Vinculação+UG+FONTE"] = (
        df["Vinculação Pagamento"] + df["UG Executora"] + df["Fonte Recursos Detalhada"]
    )

    df = df[
        [
            "Órgão UGE",
            "Descrição UGE",
            "UG Executora",
            "Descrição UG",
            "Órgão UGE - Gestão",
            "Vinculação Pagamento",
            "Fonte Recursos Detalhada",
            "Vinculação+UG+FONTE",
            "LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)",
        ]
    ]

    df.to_excel(ARQUIVO_LIMITE_SAQUE_COPIA, index=False)
    log("Limite de Saque processado com sucesso.")

def copiar_dados_para_limite_saque():
    log("Copiando dados de Limite de Saque para o arquivo principal...")

    wb = openpyxl.load_workbook(ARQUIVO_TED_GERAL_COPIA)
    planilha_limite_saque = recriar_aba(wb, SHEET_LIMITE_SAQUE)

    wb_limite = openpyxl.load_workbook(ARQUIVO_LIMITE_SAQUE_COPIA)
    planilha_dados_limite = wb_limite.active

    copiar_aba_por_valores(planilha_dados_limite, planilha_limite_saque)

    wb.save(ARQUIVO_TED_GERAL_COPIA)
    wb.close()
    wb_limite.close()

# =============================================================================
# ETAPA 6 - CONTAS CADASTRO E CONTROLE
# =============================================================================

def copiarArquivoSemMesclagem(caminho_arquivo_original):
    log("Copiando arquivo Contas Cadastro e Controle sem mesclagens...")

    wb_original = load_workbook(caminho_arquivo_original)
    desmesclar_todas_as_abas(wb_original)
    wb_original.save(caminho_arquivo_original)
    wb_original.close()

    copiar_arquivo(caminho_arquivo_original, ARQUIVO_CONTAS_CONTROLE_COPIA)

    wb_copia = load_workbook(ARQUIVO_CONTAS_CONTROLE_COPIA)
    desmesclar_todas_as_abas(wb_copia)

    for sheet in wb_copia.worksheets:
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value)[2:]

    wb_copia.save(ARQUIVO_CONTAS_CONTROLE_COPIA)
    wb_copia.close()

    log("Arquivo copiado sem mesclagem e com ajuste da Conta Corrente.")

def copiarRepasse():
    log("Copiando Contas Cadastro e Controle para o arquivo principal...")

    arquivo_origem = openpyxl.load_workbook(ARQUIVO_CONTAS_CONTROLE_COPIA)
    arquivo_destino = openpyxl.load_workbook(ARQUIVO_TED_GERAL_COPIA)

    aba_destino = recriar_aba(arquivo_destino, SHEET_CONTAS_CONTROLE)
    aba_origem = arquivo_origem.active

    for row in aba_origem.iter_rows(min_row=1, values_only=True):
        nova_linha = [0 if valor is None else valor for valor in row]
        aba_destino.append(nova_linha)

    yellow_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
    font = Font(bold=True, color="000000")

    for row in aba_destino.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.fill = yellow_fill
            cell.font = font

    arquivo_destino.save(ARQUIVO_TED_GERAL_COPIA)
    arquivo_origem.close()
    arquivo_destino.close()

    log("Aba Contas Cadastro e Controle copiada com sucesso.")

# =============================================================================
# ETAPA 7 - ABA REPASSAR
# =============================================================================

def criar_aba_repassar(wb):
    return recriar_aba(wb, SHEET_REPASSAR)

def processoFinal():
    log("Criando aba REPASSAR e adicionando colunas auxiliares...")

    wb = openpyxl.load_workbook(ARQUIVO_TED_GERAL_COPIA)
    planilha_tabela = wb[SHEET_TABELA]
    planilha_repassar = criar_aba_repassar(wb)

    for row in planilha_tabela.iter_rows(values_only=True):
        planilha_repassar.append(row)

    for col_idx in range(13, 21):
        for row_idx in range(1, planilha_repassar.max_row + 1):
            planilha_repassar.cell(row=row_idx, column=col_idx, value=None)

    cabecalhos = [
        "Vinculação+UG+FONTE",
        "LIMITE DE SAQUE",
        "SOMASE UG+FONTE",
        "CONDICIONAL LIMITE DE SAQUE",
        "SOMASE SIAFI",
        "CONTAS REPASSAR",
        "CONDICIONAL CONTAS A REPASSAR",
        "SIMEC (NC-PF)",
        "CONDICIONAL SIMEC",
    ]

    for col_idx, cabecalho in enumerate(cabecalhos, start=13):
        planilha_repassar.cell(row=1, column=col_idx, value=cabecalho)

    wb.save(ARQUIVO_TED_GERAL_COPIA)
    wb.close()

    log("Dados copiados para a aba REPASSAR com sucesso.")

def preencher_coluna_ug_fonte():
    log("Preenchendo Vinculação+UG+FONTE e LIMITE DE SAQUE...")

    wb = openpyxl.load_workbook(ARQUIVO_TED_GERAL_COPIA)

    planilha_repassar = wb[SHEET_REPASSAR]
    planilha_limite = wb[SHEET_LIMITE_SAQUE]

    data_repassar = planilha_repassar.values
    colunas_repassar = next(data_repassar)
    df_repassar = pd.DataFrame(data_repassar, columns=colunas_repassar)

    df_repassar["UG Executora"] = df_repassar["UG Executora"].astype(str).str.strip()
    df_repassar["Fonte Recursos Detalhada"] = df_repassar["Fonte Recursos Detalhada"].astype(str).str.strip()
    df_repassar["Vinculação Pagamento"] = df_repassar["Vinculação Pagamento"].astype(str).str.strip()
    df_repassar["Vinculação+UG+FONTE"] = (
        df_repassar["Vinculação Pagamento"] + df_repassar["UG Executora"] + df_repassar["Fonte Recursos Detalhada"]
    )

    for r_idx, ug_fonte in enumerate(df_repassar["Vinculação+UG+FONTE"], start=2):
        planilha_repassar.cell(row=r_idx, column=13, value=ug_fonte)

    data_limite = planilha_limite.values
    colunas_limite = next(data_limite)
    df_limite = pd.DataFrame(data_limite, columns=colunas_limite)

    df_limite["UG Executora"] = df_limite["UG Executora"].astype(str).str.strip()
    df_limite["Fonte Recursos Detalhada"] = df_limite["Fonte Recursos Detalhada"].astype(str).str.strip()
    df_limite["Vinculação Pagamento"] = df_limite["Vinculação Pagamento"].astype(str).str.strip()
    df_limite["Chave_Limite"] = (
        df_limite["Vinculação Pagamento"] + df_limite["UG Executora"] + df_limite["Fonte Recursos Detalhada"]
    )

    dicionario_limite = (
        df_limite.set_index("Chave_Limite")["LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)"].to_dict()
    )

    for r_idx in range(2, planilha_repassar.max_row + 1):
        chave = planilha_repassar.cell(row=r_idx, column=13).value
        limite_saque = dicionario_limite.get(chave, 0)
        planilha_repassar.cell(row=r_idx, column=14, value=limite_saque)

    wb.save(ARQUIVO_TED_GERAL_COPIA)
    wb.close()

    log("Coluna LIMITE DE SAQUE preenchida corretamente.")

def calcular_e_salvar_somase_ug_fonte(arquivo_path):
    log("Calculando SOMASE UG+FONTE e SOMASE SIAFI...")

    wb = openpyxl.load_workbook(arquivo_path)
    planilha_repassar = wb[SHEET_REPASSAR]

    ug_fonte_totals = {}
    siafi_totals = {}

    for r_idx in range(2, planilha_repassar.max_row + 1):
        total_valor = parse_numero_br(planilha_repassar.cell(row=r_idx, column=12).value)
        if total_valor is None:
            total_valor = 0.0

        planilha_repassar.cell(row=r_idx, column=12, value=total_valor)

        ug_fonte = planilha_repassar.cell(row=r_idx, column=13).value
        siafi = planilha_repassar.cell(row=r_idx, column=9).value

        ug_fonte_totals[ug_fonte] = ug_fonte_totals.get(ug_fonte, 0.0) + total_valor
        siafi_totals[siafi] = siafi_totals.get(siafi, 0.0) + total_valor

    for r_idx in range(2, planilha_repassar.max_row + 1):
        ug_fonte = planilha_repassar.cell(row=r_idx, column=13).value
        siafi = planilha_repassar.cell(row=r_idx, column=9).value

        planilha_repassar.cell(row=r_idx, column=15, value=ug_fonte_totals.get(ug_fonte, 0.0))
        planilha_repassar.cell(row=r_idx, column=17, value=siafi_totals.get(siafi, 0.0))

    wb.save(arquivo_path)
    wb.close()

def preencherCondicionalLimite(arquivo_path):
    log("Preenchendo CONDICIONAL LIMITE DE SAQUE...")

    wb = openpyxl.load_workbook(arquivo_path)
    planilha_repassar = wb[SHEET_REPASSAR]

    for r_idx in range(2, planilha_repassar.max_row + 1):
        limite_saque = parse_numero_br(planilha_repassar.cell(row=r_idx, column=14).value)
        somase_ug = parse_numero_br(planilha_repassar.cell(row=r_idx, column=15).value)

        if limite_saque is not None and somase_ug is not None:
            if limite_saque >= somase_ug:
                planilha_repassar.cell(row=r_idx, column=16, value="NÃO")
            else:
                planilha_repassar.cell(row=r_idx, column=16, value="REPASSAR")

    wb.save(arquivo_path)
    wb.close()

def preencher_coluna_contas_repassar(arquivo_path):
    log("Preenchendo CONTAS REPASSAR...")

    wb = openpyxl.load_workbook(arquivo_path)

    planilha_repassar = wb[SHEET_REPASSAR]
    planilha_contas = wb[SHEET_CONTAS_CONTROLE]

    mapa_contas = {}
    for row in planilha_contas.iter_rows(min_row=3):
        conta_corrente = row[2].value
        valores_firmados = row[5].value
        if conta_corrente is not None:
            mapa_contas[str(conta_corrente).strip()] = valores_firmados

    for r_idx in range(2, planilha_repassar.max_row + 1):
        siafi = planilha_repassar.cell(row=r_idx, column=9).value
        if siafi is not None:
            planilha_repassar.cell(row=r_idx, column=18, value=mapa_contas.get(str(siafi).strip()))

    wb.save(arquivo_path)
    wb.close()

def preencherCondicional(arquivo_path):
    log("Preenchendo CONDICIONAL CONTAS A REPASSAR...")

    wb = openpyxl.load_workbook(arquivo_path)
    planilha_repassar = wb[SHEET_REPASSAR]

    for r_idx in range(2, planilha_repassar.max_row + 1):
        p = parse_numero_br(planilha_repassar.cell(row=r_idx, column=17).value)
        q = parse_numero_br(planilha_repassar.cell(row=r_idx, column=18).value)

        if q is None:
            condicional = "SEM SALDO"
        else:
            if p is None:
                condicional = "SEM SALDO"
            else:
                condicional = "COM SALDO" if p <= q else "SEM SALDO"

        planilha_repassar.cell(row=r_idx, column=19, value=condicional)

    wb.save(arquivo_path)
    wb.close()

# =============================================================================
# ETAPA 8 - NCPF / NC-PF
# =============================================================================

def normalizar_ted(valor):
    """
    Normaliza TED para comparação segura entre abas/arquivos.
    Ex.: 16896, 16896.0, '16896 ', '16896\n' -> '16896'
    """
    if valor is None:
        return None

    texto = str(valor).strip()
    texto = remover_quebras_de_linha_texto(texto).strip()

    if texto == "" or texto.lower() == "none":
        return None

    # tenta converter caso venha como float/string numérica
    try:
        numero = float(texto.replace(",", "."))
        if numero.is_integer():
            return str(int(numero))
    except Exception:
        pass

    return texto

def processar_e_copiar_ncpf(file_path):
    log("Processando NCPF...")

    wb_ncpf = openpyxl.load_workbook(file_path, data_only=True)
    sheet_ncpf = wb_ncpf.active

    wb_copia = Workbook()
    sheet_copia = wb_copia.active
    sheet_copia.title = "NCPF"

    for row in sheet_ncpf.iter_rows(values_only=True):
        valores_limpos = [
            remover_quebras_de_linha_texto(valor) if isinstance(valor, str) else valor
            for valor in row
        ]
        sheet_copia.append(valores_limpos)

    sheet_copia["D1"] = "NC-PF"

    for linha in range(2, sheet_copia.max_row + 1):
        ted_value = sheet_copia[f"A{linha}"].value
        total_descentralizado = sheet_copia[f"B{linha}"].value
        total_repassado = sheet_copia[f"C{linha}"].value

        ted_normalizado = normalizar_ted(ted_value)
        valor_b = parse_numero_br(total_descentralizado)
        valor_c = parse_numero_br(total_repassado)

        sheet_copia[f"A{linha}"] = ted_normalizado

        if valor_b is not None and valor_c is not None:
            resultado = round(valor_b - valor_c, 2)
            sheet_copia[f"D{linha}"] = resultado
        else:
            sheet_copia[f"D{linha}"] = None
            log(
                f"[NCPF][LINHA {linha}] Falha ao converter | "
                f"B bruto={repr(total_descentralizado)} -> {valor_b} | "
                f"C bruto={repr(total_repassado)} -> {valor_c}"
            )

    wb_copia.save(ARQUIVO_NCPF_COPIA)
    wb_copia.close()
    wb_ncpf.close()

    log("NCPF processado com sucesso.")

def copiar_colunas_para_aba_nc_pf(arquivo_origem, arquivo_destino):
    log("Copiando dados de NCPF para a aba NC-PF...")

    wb_origem = openpyxl.load_workbook(arquivo_origem, data_only=True)
    sheet_origem = wb_origem.active

    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_destino = recriar_aba(wb_destino, SHEET_NC_PF)

    sheet_destino.append(["TED", "NC-PF"])

    for row_idx, row in enumerate(
        sheet_origem.iter_rows(min_row=2, max_row=sheet_origem.max_row, values_only=True),
        start=2
    ):
        ted_value = normalizar_ted(row[0] if len(row) > 0 else None)
        nc_pf_value = row[3] if len(row) > 3 else None

        if ted_value is None and nc_pf_value is None:
            continue

        sheet_destino.append([ted_value, nc_pf_value])

        if row_idx <= 10:
            log(f"[NC-PF][LINHA {row_idx}] TED={repr(ted_value)} | NC-PF={repr(nc_pf_value)}")

    wb_destino.save(arquivo_destino)
    wb_origem.close()
    wb_destino.close()

    log("Dados copiados para a aba NC-PF com sucesso.")

def preencher_coluna_simec(arquivo_destino, aba_repassar, aba_nc_pf):
    log("Preenchendo coluna SIMEC (NC-PF)...")

    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_repassar = wb_destino[aba_repassar]
    sheet_nc_pf = wb_destino[aba_nc_pf]

    ted_nc_pf_map = {}

    # monta mapa TED -> NC-PF com dados já normalizados
    for row in sheet_nc_pf.iter_rows(min_row=2, values_only=True):
        ted_value = normalizar_ted(row[0] if len(row) > 0 else None)
        nc_pf_value = parse_numero_br(row[1] if len(row) > 1 else None)

        if ted_value is not None:
            ted_nc_pf_map[ted_value] = nc_pf_value

    # coluna E = índice 4 | coluna T = índice 19
    for row in sheet_repassar.iter_rows(min_row=2):
        ted_repassar = normalizar_ted(row[4].value)

        if ted_repassar is None:
            continue

        nc_pf_value = ted_nc_pf_map.get(ted_repassar)

        if nc_pf_value is not None:
            row[19].value = round(nc_pf_value, 2)
        else:
            row[19].value = None

    wb_destino.save(arquivo_destino)
    wb_destino.close()

    log("Coluna SIMEC (NC-PF) preenchida com sucesso.")

# =============================================================================
# ETAPA 9 - CONDICIONAL SIMEC / FORMATAÇÃO
# =============================================================================

def aplicar_logica_condicional_simec(sheet):
    casas_decimais = 2
    yellow_fill = PatternFill(start_color="E4BB02", end_color="E4BB02", fill_type="solid")

    for row in sheet.iter_rows(min_row=2):
        somase_siafi = parse_numero_br(row[16].value)   # coluna 17
        simec_nc_pf = parse_numero_br(row[19].value)    # coluna 20
        valor_k = parse_numero_br(row[11].value)        # coluna 12
        valor_m = parse_numero_br(row[13].value)        # coluna 14

        # pinta a linha se K < M
        # se sua regra correta for K > M, troque aqui
        if valor_k is not None and valor_m is not None:
            if valor_k < valor_m:
                for cell in row:
                    cell.fill = yellow_fill

        # preenche condicional simec
        if somase_siafi is not None and simec_nc_pf is not None:
            somase_siafi = round(somase_siafi, casas_decimais)
            simec_nc_pf = round(simec_nc_pf, casas_decimais)

            if somase_siafi == simec_nc_pf:
                row[20].value = "REPASSAR"
            elif somase_siafi > simec_nc_pf:
                row[20].value = "NÃO"
            else:
                row[20].value = "REPASSAR"
        else:
            row[20].value = None

    # formata só no final
    colunas_a_formatar = [10, 11, 13, 14, 15, 16, 17, 18, 19]
    for coluna_index in colunas_a_formatar:
        for row in sheet.iter_rows(min_row=2):
            cell = row[coluna_index]
            valor_numerico = parse_numero_br(cell.value)

            if valor_numerico is not None:
                cell.value = formatar_contabil(valor_numerico)

def preencher_condicional_simec(arquivo_destino, aba_repassar):
    log("Preenchendo CONDICIONAL SIMEC na aba REPASSAR...")

    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_repassar = wb_destino[aba_repassar]

    aplicar_logica_condicional_simec(sheet_repassar)

    wb_destino.save(arquivo_destino)
    wb_destino.close()

# =============================================================================
# ETAPA 10 - EMENDAS
# =============================================================================

def mover_linhas_emendas(arquivo_destino):
    log("Movendo linhas de emendas para a aba EMENDAS...")

    wb = openpyxl.load_workbook(arquivo_destino)

    if SHEET_REPASSAR not in wb.sheetnames:
        log("Aba REPASSAR não encontrada no arquivo.")
        wb.close()
        return

    sheet_repassar = wb[SHEET_REPASSAR]
    sheet_emendas = obter_ou_criar_aba(wb, SHEET_EMENDAS)

    df_repassar = pd.read_excel(arquivo_destino, sheet_name=SHEET_REPASSAR)

    if "Resultado EOF" not in df_repassar.columns:
        log("Coluna 'Resultado EOF' não encontrada na aba REPASSAR.")
        wb.close()
        return

    df_emendas = df_repassar[df_repassar["Resultado EOF"].isin([6, 7, 8])]

    if df_emendas.empty:
        log("Nenhuma linha para mover para EMENDAS.")
        wb.close()
        return

    if sheet_emendas.max_row == 1 and sheet_emendas.cell(row=1, column=1).value is None:
        for col_idx, col_name in enumerate(df_emendas.columns, start=1):
            sheet_emendas.cell(row=1, column=col_idx, value=col_name)

    for _, row in df_emendas.iterrows():
        nova_linha_emendas = sheet_emendas.max_row + 1
        linha_original = row.name + 2

        for col_idx, value in enumerate(row, start=1):
            cell_origem = sheet_repassar.cell(row=linha_original, column=col_idx)
            cell_destino = sheet_emendas.cell(row=nova_linha_emendas, column=col_idx, value=value)

            if cell_origem.fill:
                cell_destino.fill = PatternFill(
                    start_color=cell_origem.fill.start_color.rgb,
                    end_color=cell_origem.fill.end_color.rgb,
                    fill_type=cell_origem.fill.fill_type,
                )

    linhas_para_remover = sorted([idx + 2 for idx in df_emendas.index], reverse=True)
    for linha in linhas_para_remover:
        sheet_repassar.delete_rows(linha)

    wb.save(arquivo_destino)
    wb.close()

    log("Linhas copiadas para EMENDAS e removidas de REPASSAR.")

def preencher_condicional_simec2(arquivo_destino, aba_emendas):
    log("Preenchendo CONDICIONAL SIMEC na aba EMENDAS...")

    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_emendas = wb_destino[aba_emendas]

    aplicar_logica_condicional_simec(sheet_emendas)

    wb_destino.save(arquivo_destino)
    wb_destino.close()

# =============================================================================
# ETAPA 11 - CÓPIAS FINAIS
# =============================================================================

def copiar_arquivo_liquidacao():
    log("Copiando arquivo final LIQUIDAÇÃO A REPASSAR.xlsx...")
    try:
        shutil.copy(ARQUIVO_TED_GERAL_COPIA, ARQUIVO_FINAL_REPASSAR)
        log(f"Arquivo copiado com sucesso para {ARQUIVO_FINAL_REPASSAR}")
    except Exception as e:
        log(f"Ocorreu um erro ao copiar o arquivo: {e}")


def copiar_arquivo_liquidacao_para_financeiro():
    log("Copiando arquivo final para a pasta do financeiro...")

    try:
        data_atual, mes_atual_nome = obter_mes_atual_formatado()
        pasta_mes = os.path.join(DESTINO_FINANCEIRO_2026, mes_atual_nome)

        if not os.path.exists(pasta_mes):
            os.makedirs(pasta_mes)
            log(f"Pasta do mês criada: {pasta_mes}")

        nome_arquivo = f"LIQUIDAÇÃO A REPASSAR {data_atual.strftime('%d-%m-%y')}.xlsx"
        destino_final = os.path.join(pasta_mes, nome_arquivo)

        shutil.copy(ARQUIVO_FINAL_REPASSAR, destino_final)
        log(f"Arquivo copiado com sucesso para {destino_final}")
    except Exception as e:
        log(f"Ocorreu um erro ao copiar o arquivo: {e}")


def copiar_arquivo_liquidacao_para_base_automacao():
    log("Copiando arquivo final para a pasta mensal da automação...")

    try:
        data_atual, mes_atual_nome = obter_mes_atual_formatado()
        destino_base = BASE_DIR
        pasta_mes = os.path.join(destino_base, mes_atual_nome)

        if not os.path.exists(pasta_mes):
            os.makedirs(pasta_mes)
            log(f"Pasta do mês criada: {pasta_mes}")

        nome_arquivo = f"LIQUIDAÇÃO A REPASSAR {data_atual.strftime('%d-%m-%y')}.xlsx"
        destino_final = os.path.join(pasta_mes, nome_arquivo)

        shutil.copy(ARQUIVO_FINAL_REPASSAR, destino_final)
        log(f"Arquivo copiado com sucesso para {destino_final}")
    except Exception as e:
        log(f"Ocorreu um erro ao copiar o arquivo: {e}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    # 1) Execução SIMEC
    processar_exeSimec()

    # 2) TED Liquidação Geral
    processar_arquivo_ted_geral()
    excluir_linhas_total_zero(ARQUIVO_TED_GERAL_COPIA)

    # 3) DOC NE no arquivo principal
    copyData()
    preencher_colunas_siafi_vigencia_estado_atual(
        ARQUIVO_TED_GERAL_COPIA,
        obter_data_atual(),
    )

    # 4) Tabela
    criar_tabela(ARQUIVO_TED_GERAL_COPIA)
    converter_coluna_total_para_numero(ARQUIVO_TED_GERAL_COPIA)
    agrupar_e_somar_total(ARQUIVO_TED_GERAL_COPIA)

    # 5) Limite de Saque
    processar_limite_saque(ARQUIVO_LIMITE_SAQUE_ORIG)
    copiar_dados_para_limite_saque()

    # 6) Contas Cadastro e Controle
    copiarArquivoSemMesclagem(ARQUIVO_CONTAS_CONTROLE_ORIG)
    copiarRepasse()

    # 7) REPASSAR
    processoFinal()
    preencher_coluna_ug_fonte()
    calcular_e_salvar_somase_ug_fonte(ARQUIVO_TED_GERAL_COPIA)
    preencherCondicionalLimite(ARQUIVO_TED_GERAL_COPIA)
    preencher_coluna_contas_repassar(ARQUIVO_TED_GERAL_COPIA)
    preencherCondicional(ARQUIVO_TED_GERAL_COPIA)

    # 8) NCPF / NC-PF
    processar_e_copiar_ncpf(ARQUIVO_NCPF_ORIG)
    copiar_colunas_para_aba_nc_pf(ARQUIVO_NCPF_COPIA, ARQUIVO_TED_GERAL_COPIA)
    preencher_coluna_simec(ARQUIVO_TED_GERAL_COPIA, SHEET_REPASSAR, SHEET_NC_PF)

    # 9) Condicional SIMEC
    preencher_condicional_simec(ARQUIVO_TED_GERAL_COPIA, SHEET_REPASSAR)

    # 10) Emendas
    mover_linhas_emendas(ARQUIVO_TED_GERAL_COPIA)
    preencher_condicional_simec2(ARQUIVO_TED_GERAL_COPIA, SHEET_EMENDAS)

    # 11) Cópias finais
    copiar_arquivo_liquidacao()
    copiar_arquivo_liquidacao_para_financeiro()
    copiar_arquivo_liquidacao_para_base_automacao()

    log('Processo Totalmente Finalizado. O seu arquivo final se encontra em "W:\\B - TED\\7 - AUTOMAÇÃO\\Liquidação\\LIQUIDAÇÃO A REPASSAR.xlsx"')

if __name__ == "__main__":
    main()
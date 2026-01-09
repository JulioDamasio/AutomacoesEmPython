import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
import pandas as pd
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

def copiar_arquivo(origem, destino):
    try:
        shutil.copy(origem, destino)
        print(f"Arquivo copiado de {origem} para {destino} mantendo a formatação.")
    except Exception as e:
        print(f"Erro ao copiar o arquivo: {e}")

def copiar_arquivos(destino):
    origens = [
        r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\TEDS para Finalizar.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\TED - Contas Cadastro e Controle.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\PFs desde 2013 TG.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\Ncs desde 2013 TG.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\Pfs desde 2013 SIMEC.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\Ncs desde 2013 SIMEC.xlsx'
    ]
    for origem in origens:
        caminho, nome_arquivo = os.path.split(origem)
        novo_nome = "COPIA " + nome_arquivo
        shutil.copy(origem, os.path.join(destino, novo_nome))  # Copia para a pasta destino com o novo nome
        print(f"Arquivo {nome_arquivo} copiado como {novo_nome} para {destino}")

def copiar_e_apagar_linhas(arquivo_origem, arquivo_destino):
    try:
        # Copiar o arquivo para o destino
        shutil.copyfile(arquivo_origem, arquivo_destino)
        
        # Carregar o arquivo copiado com o pandas, ignorando as 12 primeiras linhas
        df = pd.read_excel(arquivo_destino, skiprows=12)

        # Renomear as colunas
        novo_cabecalho = [
            "Resultado EOF", "Descrição EOF", "NE CCor - Ano Emissão", "Órgão UGE", "Descrição UGE",
            "UG Executora", "Descrição UG", "UGE - UG Setorial Financeira", "Descrição Setorial",
            "Ação Governo", "Descrição Ação", "Função PO", "SubFunção PO", "Programa PO", "Cod PO",
            "Descrição PO", "PTRES", "PI", "NE CCor", "Grupo Despesa", "Natureza Despesa Detalhada",
            "Descrição Natureza", "Elemento Despesa", "Descrição Elemento", "Fonte Recursos Detalhada",
            "DESTAQUE RECEBIDO", "CREDITO DISPONIVEL", "DESPESAS EMPENHADAS (CONTROLE EMPENHO)",
            "DESPESAS PAGAs (CONTROLE EMPENHO)", "Total"
        ]
        df.columns = novo_cabecalho

        # Salvar o DataFrame de volta no arquivo Excel
        df.to_excel(arquivo_destino, index=False)

        print("Primeiras 12 linhas apagadas e cabeçalho renomeado com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao copiar e apagar linhas: {e}")        

def criar_tabela_dinamica(ted_para_finalizar, copia_ted_para_finalizar):
    try:
        # Fazendo uma cópia do arquivo de origem
        df = pd.read_excel(ted_para_finalizar, header=[0, 1])  # Lendo as duas primeiras linhas como cabeçalho
        df = df.drop(df.columns[0], axis=1)  # Excluindo a primeira coluna

        # Renomeando as colunas
        df.columns = ['Estado Atual', 'SIAFI', 'TED', 'Valor Orçamentário (R$)']
        
        # Salvando o DataFrame modificado no arquivo de destino
        df.to_excel(copia_ted_para_finalizar, index=False)
        
        # Lendo o arquivo recém-criado
        df_copia = pd.read_excel(copia_ted_para_finalizar)
        
        # Criando a tabela dinâmica
        pivot_table = df_copia.groupby(['SIAFI','TED'])['Valor Orçamentário (R$)'].sum().reset_index()
        
        # Salvando a tabela dinâmica
        with pd.ExcelWriter(copia_ted_para_finalizar, mode='a', engine='openpyxl') as writer:
            pivot_table.to_excel(writer, sheet_name='TED', index=False)

        print("Tabela dinâmica criada com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def remover_primeira_aba(arquivo):
    try:
        # Carregar o workbook
        wb = load_workbook(arquivo)
        
        # Remover a primeira aba
        if len(wb.sheetnames) > 1:
            first_sheet = wb.sheetnames[0]
            del wb[first_sheet]
        
        # Salvar o workbook com a aba removida
        wb.save(arquivo)
        print("Primeira aba removida com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro ao remover a primeira aba: {e}")

def preencher_colunas_SIMEC(copia_ted_para_finalizar, copia_arquivo_valores, coluna_valores, coluna_destino, chave_coluna='SIAFI'):
    try:
        # Carregar os DataFrames
        df_ted = pd.read_excel(copia_ted_para_finalizar)
        df_valores = pd.read_excel(copia_arquivo_valores)
        
        # Corrigir possíveis espaços e caracteres indesejados nas colunas
        df_ted.columns = df_ted.columns.str.strip()
        df_valores.columns = df_valores.columns.str.strip()
        
        # Verificar se as colunas esperadas estão presentes
        if chave_coluna not in df_ted.columns:
            raise ValueError(f"A coluna chave '{chave_coluna}' não foi encontrada em {copia_ted_para_finalizar}.")
        if coluna_valores not in df_valores.columns:
            raise ValueError(f"A coluna de valores '{coluna_valores}' não foi encontrada em {copia_arquivo_valores}.")
        
        # Criar dicionário de lookup para chave_coluna e valores
        lookup_dict = df_valores.groupby(chave_coluna)[coluna_valores].sum().to_dict()
        
        # Preencher a coluna destino com base no lookup_dict
        df_ted[coluna_destino] = df_ted[chave_coluna].map(lookup_dict).fillna(0)
        
        # Salvar as alterações no arquivo
        df_ted.to_excel(copia_ted_para_finalizar, index=False)
        
        print(f"Coluna {coluna_destino} preenchida com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna {coluna_destino}: {e}")

def copiar_linhas_ano(arquivo_origem, arquivo_destino, coluna_data_origem, coluna_data_destino, ano=2025, linha_cabecalho=0):
    try:
        # Suprimir avisos do openpyxl
        warnings.simplefilter(action='ignore', category=UserWarning)
        
        # Carregar os dados do arquivo de origem
        df_origem = pd.read_excel(arquivo_origem, header=linha_cabecalho)
        
        # Verificar se a coluna de data existe no DataFrame de origem
        if coluna_data_origem not in df_origem.columns:
            raise ValueError(f"A coluna '{coluna_data_origem}' não foi encontrada no arquivo de origem.")
        
        # Filtrar as linhas que contêm o ano especificado na coluna de datas
        df_origem[coluna_data_origem] = pd.to_datetime(df_origem[coluna_data_origem], errors='coerce', dayfirst=True)
        df_2025 = df_origem[df_origem[coluna_data_origem].dt.year == ano]

        # Carregar os dados do arquivo de destino
        try:
            df_destino = pd.read_excel(arquivo_destino)
        except FileNotFoundError:
            # Se o arquivo não existir, criar um novo DataFrame vazio
            df_destino = pd.DataFrame()

        # Verificar se a coluna de data existe no DataFrame de destino
        if coluna_data_destino not in df_destino.columns and not df_destino.empty:
            raise ValueError(f"A coluna '{coluna_data_destino}' não foi encontrada no arquivo de destino.")

        if not df_destino.empty:
            # Remover as linhas existentes no destino que têm o ano especificado
            df_destino[coluna_data_destino] = pd.to_datetime(df_destino[coluna_data_destino], errors='coerce', dayfirst=True)
            df_destino = df_destino[df_destino[coluna_data_destino].dt.year != ano]

        # Adicionar as linhas filtradas ao DataFrame de destino
        df_destino = pd.concat([df_destino, df_2025], ignore_index=True)

        # Salvar o DataFrame atualizado no arquivo de destino
        df_destino.to_excel(arquivo_destino, index=False)
        
        print(f"As linhas com o ano {ano} foram copiadas e adicionadas/atualizadas no arquivo de destino...")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


def preencher_coluna_valor_nc_siafi(copia_ted_para_finalizar, copia_ncs_siafi):
    try:
        # Carregar os DataFrames
        df_ted = pd.read_excel(copia_ted_para_finalizar)
        df_ncs = pd.read_excel(copia_ncs_siafi)
        
        # Criar dicionário de lookup para SIAFI e valores
        lookup_dict = df_ncs.groupby('Nc- Transferencia')['Valor Absoluto'].sum().to_dict()
        
        # Preencher a coluna VALOR NC SIMEC com base no lookup_dict
        df_ted['VALOR NC SIAFI'] = df_ted['SIAFI'].map(lookup_dict).fillna(0)
        
        # Salvar as alterações no arquivo
        df_ted.to_excel(copia_ted_para_finalizar, index=False)
        
        print("Coluna VALOR NC SIAFI preenchida com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna VALOR NC SIMEC: {e}")

def preencher_colunas_SIMEC(copia_ted_para_finalizar, copia_arquivo_valores, coluna_valores, coluna_destino, chave_coluna):
    try:
        # Carregar os DataFrames
        df_ted = pd.read_excel(copia_ted_para_finalizar)
        df_valores = pd.read_excel(copia_arquivo_valores)
        
        # Corrigir possíveis espaços e caracteres indesejados nas colunas
        df_ted.columns = df_ted.columns.str.strip()
        df_valores.columns = df_valores.columns.str.strip()
        
        # Verificar se as colunas esperadas estão presentes
        if 'SIAFI' not in df_ted.columns:
            raise ValueError(f"A coluna 'SIAFI' não foi encontrada em {copia_ted_para_finalizar}.")
        if chave_coluna not in df_valores.columns:
            raise ValueError(f"A coluna chave '{chave_coluna}' não foi encontrada em {copia_arquivo_valores}.")
        if coluna_valores not in df_valores.columns:
            raise ValueError(f"A coluna de valores '{coluna_valores}' não foi encontrada em {copia_arquivo_valores}.")
        
        # Corrigir possíveis espaços e caracteres indesejados nas chaves
        df_valores[chave_coluna] = df_valores[chave_coluna].astype(str).str.strip()
        df_ted['SIAFI'] = df_ted['SIAFI'].astype(str).str.strip()
        
        # Criar dicionário de lookup para chave_coluna e valores
        lookup_dict = df_valores.groupby(chave_coluna)[coluna_valores].sum().to_dict()
        
        # Debug: Verificar alguns valores do dicionário e da coluna de destino
        print("Exemplos de lookup_dict:")
        for k, v in list(lookup_dict.items())[:5]:
            print(f"{k}: {v}")
        
        print("Exemplos de SIAFI em df_ted:")
        print(df_ted['SIAFI'].head())
        
        # Preencher a coluna destino com base no lookup_dict
        df_ted[coluna_destino] = df_ted['SIAFI'].map(lookup_dict).fillna(0)
        
        # Debug: Verificar alguns valores preenchidos
        print("Exemplos de valores preenchidos:")
        print(df_ted[[coluna_destino, 'SIAFI']].head())
        
        # Salvar as alterações no arquivo
        df_ted.to_excel(copia_ted_para_finalizar, index=False)
        
        print(f"Coluna {coluna_destino} preenchida com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna {coluna_destino}: {e}")
        
def adicionar_colunas(arquivo):
    try:
        # Carregar o workbook
        wb = load_workbook(arquivo)
        
        # Selecionar a aba TED
        ws = wb['TED']
        
        # Adicionar novas colunas com os nomes especificados
        ws.cell(row=1, column=1, value='SIAFI')
        ws.cell(row=1, column=2, value='TED')
        ws.cell(row=1, column=3, value='VALOR ORÇAMENTÁRIO SIMEC')
        ws.cell(row=1, column=4, value='VALOR NC SIMEC')
        ws.cell(row=1, column=5, value='VALOR NC SIAFI')
        ws.cell(row=1, column=6, value='VALOR PF SIMEC')
        ws.cell(row=1, column=7, value='VALOR PF SIAFI')
        ws.cell(row=1, column=8, value='NC - PF SIMEC')
        ws.cell(row=1, column=9, value='NC - PF SIAFI')
        ws.cell(row=1, column=10, value='ORÇAMENTÁRIO - PF SIMEC')
        ws.cell(row=1, column=11, value='VALORES FIRMADOS')
        ws.cell(row=1, column=12, value='A REPASSAR')
        ws.cell(row=1, column=13, value='A COMPROVAR')
        ws.cell(row=1, column=14, value='COMPROVADO')
        ws.cell(row=1, column=15, value='NÃO REPASSADO/ DEVOLVIDO')
        ws.cell(row=1, column=16, value='EQUIVALÊNCIA NC')
        ws.cell(row=1, column=17, value='EQUIVALÊNCIA PF')
        ws.cell(row=1, column=18, value='EQUIVALÊNCIA A COMPROVAR')
        ws.cell(row=1, column=19, value='EQUIVALÊNCIA A REPASSAR')
        
        # Salvar as alterações
        wb.save(arquivo)
        print("Novas colunas adicionadas com sucesso à aba TED...")
    except Exception as e:
        print(f"Ocorreu um erro ao adicionar colunas: {e}")          

def processar_arquivo_contas_cadastro(arquivo):
    try:
        # Carregar o DataFrame com a primeira linha como cabeçalho
        df = pd.read_excel(arquivo, header=None)
        
        # Definir o cabeçalho manualmente
        df.columns = ['UG Executora', 'Descrição UG', 'SIAFI', 'Transferência - Dian Final Vigência', 
                      'VALORES FIRMADOS', 'A REPASSAR', 'A COMPROVAR', 'COMPROVADO', 
                      'VALOR NÃO REPASSADO/DEVOLVIDO']
        
        # Remover as duas linhas indesejadas
        df = df.drop(index=[0, 1])
        
        # Resetar o índice após a exclusão das linhas
        df = df.reset_index(drop=True)
        
        # Remover os caracteres 'ED' da coluna 'SIAFI'
        df['SIAFI'] = df['SIAFI'].astype(str).str.replace('ED', '', regex=False)
        
        # Salvar o DataFrame modificado no arquivo
        df.to_excel(arquivo, index=False)
        
        print("Arquivo processado com sucesso...")
    except Exception as e:
        print(f"Ocorreu um erro ao processar o arquivo: {e}")

def subtrair_e_preencher(arquivo_origem, coluna1, coluna2, coluna_resultado):
    # Carregar o DataFrame do arquivo Excel
    df = pd.read_excel(arquivo_origem)

    # Garantir que as colunas sejam do tipo float
    df[coluna1] = pd.to_numeric(df[coluna1], errors='coerce')
    df[coluna2] = pd.to_numeric(df[coluna2], errors='coerce')

    # Subtrair os valores das duas colunas e preencher a terceira coluna
    df[coluna_resultado] = df[coluna1] - df[coluna2]

    # Salvar as alterações no DataFrame de volta para o arquivo Excel
    df.to_excel(arquivo_origem, index=False)
    
def comparar(arquivo_origem, comparacoes):
    # Carregar o DataFrame do arquivo Excel
    df = pd.read_excel(arquivo_origem)

    # Garantir que as colunas sejam do tipo float, lidando com erros de conversão
    def to_float(value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return None  # Retornar None para valores que não podem ser convertidos

    for coluna1, coluna2, coluna_resultado in comparacoes:
        df[coluna1] = df[coluna1].apply(to_float)
        df[coluna2] = df[coluna2].apply(to_float)

        # Arredondar os valores das colunas para 2 casas decimais
        df[coluna1] = df[coluna1].round(2)
        df[coluna2] = df[coluna2].round(2)

        # Comparar os valores das duas colunas com precisão de 2 casas decimais
        df[coluna_resultado] = df.apply(
            lambda row: (row[coluna1] == row[coluna2])
            if pd.notnull(row[coluna1]) and pd.notnull(row[coluna2])
            else False, axis=1
        )

    # Salvar as alterações no DataFrame de volta para o arquivo Excel
    df.to_excel(arquivo_origem, index=False)

    # Carregar o arquivo Excel com openpyxl para aplicar a formatação
    wb = load_workbook(arquivo_origem)

    # Salvar o arquivo Excel com as formatações aplicadas
    wb.save(arquivo_origem)

    print(f"Arquivo salvo com formatação: {arquivo_origem}")

def formatar_colunas_resultado(arquivo_origem, colunas_resultado):
    try:
        # Carregar o DataFrame do arquivo Excel
        df = pd.read_excel(arquivo_origem)

        # Carregar o arquivo Excel com openpyxl para aplicar a formatação
        wb = load_workbook(arquivo_origem)
        ws = wb.active

        # Definir cores para Verdadeiro e Falso
        verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        vermelho = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for coluna_resultado in colunas_resultado:
            if coluna_resultado in df.columns:
                col_idx = df.columns.get_loc(coluna_resultado) + 1  # +1 porque openpyxl usa índices baseados em 1
                
                for row in range(2, len(df) + 2):  # +2 para considerar o cabeçalho e basear em índice 1
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        cell.value = "Verdadeiro"
                        cell.fill = verde
                    else:
                        cell.value = "Falso"
                        cell.fill = vermelho

        # Salvar o arquivo Excel com as formatações aplicadas
        wb.save(arquivo_origem)

        print(f"Arquivo salvo com formatação: {arquivo_origem}")

    except Exception as e:
        print(f"Ocorreu um erro ao formatar o arquivo: {e}")

def formatar_valores_monetarios(copia_ted_para_finalizar, colunas_monetarias):
    try:
        df = pd.read_excel(copia_ted_para_finalizar)
        for coluna in colunas_monetarias:
            if coluna in df.columns:
                df[coluna] = df[coluna].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        df.to_excel(copia_ted_para_finalizar, index=False)
        print(f"Arquivo salvo com formatação monetária: {copia_ted_para_finalizar}")
    except Exception as e:
        print(f"Ocorreu um erro ao formatar valores monetários: {e}")

def main():
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar'
    
    NcTG = r'W:\B - TED\7 - AUTOMAÇÃO\NC e PF\NC funcionando - EXERCÍCIO 2026.xlsx'
    pfTG = r'W:\B - TED\7 - AUTOMAÇÃO\NC e PF\PF Legado - Exercício 2026.xlsx'
    ncSIMEC = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Ncs desde 2013.xlsx'
    pfSIMEC = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Pfs desde 2013.xlsx'
    
    ted_para_finalizar = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\TEDS para Finalizar.xlsx'
    ted_contas_cadastro_controle = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\TED - Contas Cadastro e Controle.xlsx'
    pfs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\Pfs desde 2013 SIMEC.xlsx'
    ncs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\Ncs desde 2013 SIMEC.xlsx'
    ncs_siafi = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\Ncs desde 2013 TG.xlsx'
    pfs_siafi = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\PFs desde 2013 TG.xlsx'
    
    copia_ted_para_finalizar = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA TEDS para Finalizar.xlsx'
    copia_ted_contas_cadastro_controle = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA TED - Contas Cadastro e Controle.xlsx'
    copia_pfs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA Pfs desde 2013 SIMEC.xlsx'
    copia_ncs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA Ncs desde 2013 SIMEC.xlsx'
    copia_ncs_siafi = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA Ncs desde 2013 TG.xlsx'
    copia_pfs_siafi = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA PFs desde 2013 TG.xlsx'
    
    copiar_arquivos(destino)
    criar_tabela_dinamica(ted_para_finalizar, copia_ted_para_finalizar)
    remover_primeira_aba(copia_ted_para_finalizar)
    adicionar_colunas(copia_ted_para_finalizar)
    processar_arquivo_contas_cadastro(copia_ted_contas_cadastro_controle)
    
    # Exemplo de uso
    copiar_linhas_ano(
        arquivo_origem = NcTG,
        arquivo_destino = copia_ncs_siafi,
        coluna_data_origem = 'Emissão - Dia',
        coluna_data_destino = 'Emissão - Dia',
        linha_cabecalho=0
        )
    
    copiar_linhas_ano(
        arquivo_origem = pfTG,
        arquivo_destino = copia_pfs_siafi,
        coluna_data_origem = 'Emissão - Dia',
        coluna_data_destino = 'Emissão dia',
        linha_cabecalho = 5)
    
    copiar_linhas_ano(
        arquivo_origem = ncSIMEC,
        arquivo_destino = copia_ncs_simec,
        coluna_data_origem = 'Data de Emissão da NC',
        coluna_data_destino = 'Data de Emissão da NC',
        linha_cabecalho = 0)
    
    copiar_linhas_ano(
        arquivo_origem = pfSIMEC,
        arquivo_destino = copia_pfs_simec,
        coluna_data_origem = 'Data de Emissão Doc. PF',
        coluna_data_destino = 'Data de Emissão Doc. PF',
        linha_cabecalho = 0)  
    
    # VALOR NC SIMEC
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ncs_simec,
        coluna_valores='Valor Total NC',
        coluna_destino='VALOR NC SIMEC',
        chave_coluna= 'SIAFI'
    )

    # VALOR PF SIMEC
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_pfs_simec,
        coluna_valores='Valor Doc. PF (R$)',
        coluna_destino='VALOR PF SIMEC',
        chave_coluna= 'SIAFI'
    )
    
    # VALOR NC SIAFI
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ncs_siafi,
        coluna_valores='Valor Absoluto',
        coluna_destino='VALOR NC SIAFI',
        chave_coluna='Nc- Transferencia'
    )

    # VALOR PF SIAFI
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_pfs_siafi,
        coluna_valores='Repassado Absoluto',
        coluna_destino='VALOR PF SIAFI',
        chave_coluna='Inscrição (6digitos)'
    )

    # VALORES FIRMADOS - TED CONTAS CADASTRO E CONTROLE
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ted_contas_cadastro_controle,
        coluna_valores='VALORES FIRMADOS',
        coluna_destino='VALORES FIRMADOS',
        chave_coluna='SIAFI'
    )
    
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ted_contas_cadastro_controle,
        coluna_valores='A REPASSAR',
        coluna_destino='A REPASSAR',
        chave_coluna='SIAFI'    
    )
    
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ted_contas_cadastro_controle,
        coluna_valores='A COMPROVAR',
        coluna_destino='A COMPROVAR',
        chave_coluna='SIAFI'                           
    )
    
    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ted_contas_cadastro_controle,
        coluna_valores='COMPROVADO',
        coluna_destino='COMPROVADO',
        chave_coluna='SIAFI'                           
    )

    preencher_colunas_SIMEC(
        copia_ted_para_finalizar=copia_ted_para_finalizar,
        copia_arquivo_valores=copia_ted_contas_cadastro_controle,
        coluna_valores='VALOR NÃO REPASSADO/DEVOLVIDO',
        coluna_destino='NÃO REPASSADO/ DEVOLVIDO',
        chave_coluna='SIAFI'                           
    )
    
    copia_ted_para_finalizar = r'W:\B - TED\7 - AUTOMAÇÃO\Teds para finalizar\COPIA TEDS para Finalizar.xlsx'
    
    # Chamar a função subtrair_e_preencher
    subtrair_e_preencher(
        arquivo_origem=copia_ted_para_finalizar,
        coluna1='VALOR NC SIMEC',
        coluna2='VALOR PF SIMEC',
        coluna_resultado='NC - PF SIMEC'
    )
    
    # Chamar a função subtrair_e_preencher
    subtrair_e_preencher(
        arquivo_origem=copia_ted_para_finalizar,
        coluna1='VALOR NC SIAFI',
        coluna2='VALOR PF SIAFI',
        coluna_resultado='NC - PF SIAFI'
    )
    
    # Chamar a função subtrair_e_preencher
    subtrair_e_preencher(
        arquivo_origem=copia_ted_para_finalizar,
        coluna1='VALOR ORÇAMENTÁRIO SIMEC',
        coluna2='VALOR PF SIMEC',
        coluna_resultado='ORÇAMENTÁRIO - PF SIMEC'
    )
        
    comparacoes = [('VALOR NC SIMEC', 'VALOR NC SIAFI', 'EQUIVALÊNCIA NC'), ('VALOR PF SIMEC', 'VALOR PF SIAFI', 'EQUIVALÊNCIA PF'), ('VALOR PF SIMEC','A COMPROVAR', 'EQUIVALÊNCIA A COMPROVAR'), ('ORÇAMENTÁRIO - PF SIMEC', 'A REPASSAR','EQUIVALÊNCIA A REPASSAR')]

    colunas_formatar = ['VALOR ORÇAMENTÁRIO SIMEC', 'VALOR NC SIMEC', 'VALOR NC SIAFI', 'VALOR PF SIMEC', 'VALOR PF SIAFI', 'NC - PF SIMEC', 'NC - PF SIAFI', 'VALORES FIRMADOS', 'A REPASSAR', 'A COMPROVAR', 'COMPROVADO', 'NÃO REPASSADO/ DEVOLVIDO']
    
    colunas_resultado = ['EQUIVALÊNCIA NC','EQUIVALÊNCIA PF','EQUIVALÊNCIA A COMPROVAR','EQUIVALÊNCIA A REPASSAR']

    colunas_monetarias = [
    'VALOR ORÇAMENTÁRIO SIMEC','VALOR NC SIMEC','VALOR NC SIAFI','VALOR PF SIMEC','VALOR PF SIAFI','NC - PF SIMEC','NC - PF SIAFI','ORÇAMENTÁRIO - PF SIMEC','VALORES FIRMADOS','A REPASSAR','A COMPROVAR','COMPROVADO','NÃO REPASSADO/ DEVOLVIDO']
    
    comparar(copia_ted_para_finalizar, comparacoes)
    formatar_valores_monetarios(os.path.join(destino, 'COPIA TEDS para Finalizar.xlsx'), colunas_monetarias)
    formatar_colunas_resultado(copia_ted_para_finalizar, colunas_resultado)   
    
if __name__ == "__main__":
    main()
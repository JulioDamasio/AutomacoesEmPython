import shutil
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

def formatar_contabil(value):
    if pd.notnull(value):
        if isinstance(value, (int, float)):
            return "{:,.2f}".format(float(value)).replace(",", "_").replace(".", ",").replace("_", ".")
        else:
            return value  # Mantém o cabeçalho ou outros valores não numéricos
    return None

def copiar_e_renomear_arquivo(arquivo_origem):
    try:
        # Obtém o caminho do diretório do arquivo de origem
        diretorio_origem = os.path.dirname(arquivo_origem)
        
        # Define o novo nome do arquivo
        novo_nome_arquivo = "CONFORMIDADE ARQUIVO FINAL.xlsx"
        
        # Cria o caminho completo para o novo arquivo
        caminho_novo_arquivo = os.path.join(diretorio_origem, novo_nome_arquivo)
        
        # Copia o arquivo de origem para o novo local com o novo nome
        shutil.copy(arquivo_origem, caminho_novo_arquivo)
        
        print(f"Arquivo copiado e renomeado com sucesso para {caminho_novo_arquivo}")
        
    except Exception as e:
        print(f"Não foi possível copiar e renomear o arquivo: {e}")
        
def copiar_arquivo(arquivo_origem, destino):
    try:
        # Certifique-se de que o diretório de destino exista
        if not os.path.exists(destino):
            os.makedirs(destino)
        
        caminho, nome_arquivo = os.path.split(arquivo_origem)
        novo_nome = "COPIA " + nome_arquivo
        arquivo_copia = os.path.join(destino, novo_nome)
        shutil.copy(arquivo_origem, arquivo_copia)  # Copia o arquivo original para a pasta destino com o novo nome
        print(f'{arquivo_origem} copiado com sucesso para {arquivo_copia}...')
        return arquivo_copia
    except Exception as e:
        print(f"Não foi possível copiar o arquivo {arquivo_origem}: {e}")       

def Excluir_linhas(arquivo):
    try:
        # Carrega o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Remove linhas onde a coluna 'SIAFI' tem valores '-' ou está em branco
        df = df[df['SIAFI'].notna() & (df['SIAFI'] != '-')]
        
        # Remove linhas onde a coluna 'Estado Atual' tem o valor 'Comprovado no SIAFI'
        df = df[df['Estado Atual'] != 'Comprovado no SIAFI.']

        # Salva o arquivo limpo de volta
        df.to_excel(arquivo, index=False)
        
        print(f'Linhas apagadas com sucesso no arquivo {arquivo}...')
    except Exception as e:
        print(f"Não foi possível apagar as linhas no arquivo {arquivo}: {e}")
 
def criar_coluna_siafi(arquivo):
    try:
        # Carrega o arquivo Excel
        df = pd.read_excel(arquivo, skiprows=[1])  # Pula a segunda linha
        
        # Cria a coluna 'SIAFI' com base na coluna 'Conta Corrente'
        df['SIAFI'] = df['Conta Corrente'].str.replace(r'^ED', '', regex=True)
        
        # Renomeia o cabeçalho das colunas
        df.columns = ['UG Executora', 'Descrição UG', 'Conta Corrente', 
                      'Transferência - Dia Final Vigência', 'Valores Firmados', 
                      'A Repassar', 'A Comprovar', 'Comprovado', 
                      'Valor Não Repassado/devolvido', 'SIAFI']
        
        # Salva o arquivo modificado de volta
        df.to_excel(arquivo, index=False)
        
        print(f'Coluna SIAFI criada com sucesso no arquivo {arquivo}...')
    except Exception as e:
        print(f"Não foi possível criar a coluna SIAFI no arquivo {arquivo}: {e}")

def tabela_dinamica(arquivo_origem_conformidade, arquivo_destino_conformidade):
    try:
        # Carregando o arquivo de origem
        df = pd.read_excel(arquivo_origem_conformidade)

        # Convertendo 'TED' para string e limpando espaços em branco
        df['TED'] = df['TED'].astype(str).str.strip()

        # Criando a tabela dinâmica agrupando por 'TED' e somando 'Valor Orçamentário (R$)'
        pivot_table = pd.pivot_table(df, 
                                     index=['Descentralizadora', 'Estado Atual', 'Fim da Vigência', 'SIAFI', 'TED'],
                                     values=['Valor Orçamentário (R$)'],
                                     aggfunc='sum',
                                     margins=True,
                                     margins_name='Total Geral').reset_index()

        # Reordenando as colunas na ordem desejada
        pivot_table = pivot_table[['Descentralizadora', 'Estado Atual', 'Fim da Vigência', 'SIAFI', 'TED', 'Valor Orçamentário (R$)']]

        # Salvando a tabela dinâmica no arquivo destino
        with pd.ExcelWriter(arquivo_destino_conformidade, mode='w', engine='openpyxl') as writer:
            pivot_table.to_excel(writer, sheet_name='Tabela Dinâmica', index=False)

        print(f"Tabela dinâmica criada com sucesso em {arquivo_destino_conformidade}...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def criar_colunas_personalizadas(arquivo):
    try:
        # Carrega o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Cria as colunas desejadas com valores iniciais vazios
        df['Vigência SIAFI'] = ''
        df['Valor Firmado SIAFI'] = ''
        df['Equivalência Data'] = ''
        df['Equivalência Valor'] = ''
        
        # Salva o arquivo modificado de volta
        df.to_excel(arquivo, index=False)
        
        print(f'Colunas criadas com sucesso no arquivo {arquivo}...')
    except Exception as e:
        print(f"Não foi possível criar as colunas no arquivo {arquivo}: {e}")

def preencher_colunas_siafi(arquivo_conformidade, arquivo_contas_cadastro):
    try:
        # Carrega o arquivo de conformidade
        df_conformidade = pd.read_excel(arquivo_conformidade)
        
        # Carrega o arquivo de contas cadastro
        df_contas_cadastro = pd.read_excel(arquivo_contas_cadastro)
        
        # Convertendo colunas SIAFI para string em ambos os DataFrames
        df_conformidade['SIAFI'] = df_conformidade['SIAFI'].astype(str)
        df_contas_cadastro['SIAFI'] = df_contas_cadastro['SIAFI'].astype(str)
        
        # Itera sobre cada linha do DataFrame de conformidade
        for index_conformidade, row_conformidade in df_conformidade.iterrows():
            siafi_alvo = row_conformidade['SIAFI']  # Obtém o valor de SIAFI para filtrar no arquivo de contas cadastro
            
            # Procura o valor de SIAFI no arquivo de contas cadastro
            row_contas_cadastro = df_contas_cadastro[df_contas_cadastro['SIAFI'] == siafi_alvo]
            
            if not row_contas_cadastro.empty:
                # Preenche as colunas desejadas com os valores correspondentes
                df_conformidade.at[index_conformidade, 'Vigência SIAFI'] = row_contas_cadastro['Transferência - Dia Final Vigência'].values[0]
                df_conformidade.at[index_conformidade, 'Valor Firmado SIAFI'] = row_contas_cadastro['Valores Firmados'].values[0]
        
        # Salva o arquivo de conformidade modificado de volta
        df_conformidade.to_excel(arquivo_conformidade, index=False)
        
        print(f'Colunas preenchidas com sucesso no arquivo {arquivo_conformidade} com base em {arquivo_contas_cadastro}...')
    except Exception as e:
        print(f"Não foi possível preencher as colunas no arquivo {arquivo_conformidade} com base em {arquivo_contas_cadastro}: {e}")    

def converter_data(data):
    """Converte uma string de data para o formato dd/mm/yyyy ou retorna 'código inválido'."""
    try:
        if 'código inválido' in data.lower():
            return 'código inválido'
        # Tenta converter data do formato yyyy-mm-dd para dd/mm/yyyy
        if '-' in data:
            partes = data.split('-')
            return f"{partes[2]}/{partes[1]}/{partes[0]}"
        # Tenta converter data do formato dd/mm/yyyy (não precisa fazer nada)
        elif '/' in data:
            return data
        else:
            return 'código inválido'
    except Exception as e:
        return 'código inválido'

def preencher_equivalencia(arquivo_conformidade):
    try:
        # Carrega o arquivo de conformidade
        df = pd.read_excel(arquivo_conformidade)

        # Converte as colunas 'Fim da Vigência' e 'Vigência SIAFI' para string
        df['Fim da Vigência'] = df['Fim da Vigência'].astype(str)
        df['Vigência SIAFI'] = df['Vigência SIAFI'].astype(str)

        # Aplica a função de conversão às colunas de data
        df['Fim da Vigência'] = df['Fim da Vigência'].apply(converter_data)
        df['Vigência SIAFI'] = df['Vigência SIAFI'].apply(converter_data)

        # Compara os valores das duas colunas de data e preenche a coluna 'Equivalência Data'
        df['Equivalência Data'] = df.apply(lambda row: 'Verdadeiro' if row['Fim da Vigência'] == row['Vigência SIAFI'] else 'Falso', axis=1)

        # Remove os caracteres não numéricos e converte para float as colunas de valor
        df['Valor Orçamentário (R$)'] = df['Valor Orçamentário (R$)'].replace('[^\d,.-]', '', regex=True).replace(',', '.', regex=True).astype(float)
        df['Valor Firmado SIAFI'] = df['Valor Firmado SIAFI'].replace('[^\d,.-]', '', regex=True).replace(',', '.', regex=True).astype(float)

        # Compara os valores das duas colunas de valor e preenche a coluna 'Equivalência Valor'
        df['Equivalência Valor'] = df.apply(lambda row: 'Verdadeiro' if abs(row['Valor Orçamentário (R$)'] - row['Valor Firmado SIAFI']) < 1e-2 else 'Falso', axis=1)

        # Formata as colunas de valores usando a função formatar_contabil
        df['Valor Orçamentário (R$)'] = df['Valor Orçamentário (R$)'].apply(formatar_contabil)
        df['Valor Firmado SIAFI'] = df['Valor Firmado SIAFI'].apply(formatar_contabil)

        # Salva o DataFrame modificado de volta ao Excel
        df.to_excel(arquivo_conformidade, index=False)

        # Carrega o arquivo Excel com openpyxl para aplicar formatação condicional
        wb = load_workbook(arquivo_conformidade)
        ws = wb.active

        # Define as cores para Verdadeiro e Falso
        fill_verdadeiro = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fill_falso = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Obtem o índice das colunas 'Equivalência Data' e 'Equivalência Valor'
        col_eq_data_idx = df.columns.get_loc('Equivalência Data') + 1
        col_eq_valor_idx = df.columns.get_loc('Equivalência Valor') + 1

        # Aplica a formatação condicional nas colunas 'Equivalência Data' e 'Equivalência Valor'
        for row in range(2, ws.max_row + 1):  # Ignorando o cabeçalho
            cell_equivalencia_data = ws.cell(row=row, column=col_eq_data_idx)
            cell_equivalencia_valor = ws.cell(row=row, column=col_eq_valor_idx)

            if cell_equivalencia_data.value == 'Verdadeiro':
                cell_equivalencia_data.fill = fill_verdadeiro
            elif cell_equivalencia_data.value == 'Falso':
                cell_equivalencia_data.fill = fill_falso

            if cell_equivalencia_valor.value == 'Verdadeiro':
                cell_equivalencia_valor.fill = fill_verdadeiro
            elif cell_equivalencia_valor.value == 'Falso':
                cell_equivalencia_valor.fill = fill_falso

        # Salva o arquivo Excel com as alterações de formatação
        wb.save(arquivo_conformidade)

        print(f'Colunas Equivalência Data e Equivalência Valor preenchidas e formatadas com sucesso no arquivo {arquivo_conformidade}...')
    except Exception as e:
        print(f"Não foi possível preencher e formatar as colunas no arquivo {arquivo_conformidade}: {e}")
      
def main():
    
    print('iniciando processamento...')
    
    # Arquivos a serem copiados
    conformidade = r'W:\B - TED\7 - AUTOMAÇÃO\Conformidade\Conformidade TEDs Adm Direta.xlsx'
    copia_conformidade = r'W:\B - TED\7 - AUTOMAÇÃO\Conformidade\COPIA Conformidade TEDs Adm Direta.xlsx'
    contas_cadastro = r'W:\B - TED\7 - AUTOMAÇÃO\Conformidade\TED - Contas Cadastro e Controle.xlsx'
    copia_contas_cadastro = r'W:\B - TED\7 - AUTOMAÇÃO\Conformidade\COPIA TED - Contas Cadastro e Controle.xlsx'
    
    # Destino para onde os arquivos serão copiados
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Conformidade'
    
    # Chama a função para copiar cada arquivo
    copiar_arquivo(conformidade, destino)
    copiar_arquivo(contas_cadastro, destino)
    Excluir_linhas(copia_conformidade)
    tabela_dinamica(copia_conformidade, copia_conformidade)
    criar_coluna_siafi(copia_contas_cadastro)
    criar_colunas_personalizadas(copia_conformidade)
    preencher_colunas_siafi(copia_conformidade, copia_contas_cadastro)
    converter_data(copia_conformidade)
    preencher_equivalencia(copia_conformidade)
    copiar_e_renomear_arquivo(copia_conformidade)
    
if __name__ == "__main__":
    main()
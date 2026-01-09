import openpyxl
import shutil
import os
import pandas as pd
import math
import openpyxl
import shutil
import datetime
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

print('Iniciando processamento do Painel...')
# Defina as variáveis globais aqui
col_ano_execucao = None
col_unidade = None
col_regiao = None

def processar_planilha(sheet):
    
    # Crie uma cópia do conjunto de células mescladas
    merged_ranges = sheet.merged_cells.ranges.copy()

    # Remover quebra de texto e mescla de células da linha de cabeçalho
    for merged_range in merged_ranges:
        if merged_range.min_row == 1:
            for row in sheet.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrapText=False)

        sheet.unmerge_cells(merged_range.coord)
    print('Quebra de texto executada')

    # Adicionar novas colunas no final
    col_ano_execucao = sheet.max_column + 1
    col_unidade = sheet.max_column + 2
    col_regiao = sheet.max_column + 3

    print('Colunas Criadas...')

    sheet.cell(row=1, column=col_ano_execucao, value="Ano Execução")
    sheet.cell(row=1, column=col_unidade, value="Unidade Descentralizadora")
    sheet.cell(row=1, column=col_regiao, value="Região")

    print('Colunas Criadas Nomeadas como: Ano Execução, Unidade Descentralizadora e Região...')

    # Sobrescrever o cabeçalho com as novas informações
    new_header = ["Órgão UGE - Código", "Órgão UGE - Descrição Nome", "UG Executora - Código",
                  "UG Executora - Nome Completo", "UG Executora - Sigla TG", "Ação Governo - Código",
                  "PTRES", "Categoria Econômica", "Dotação Atualizada",
                  "Destaque Recebido", "Crédito Disponível", "Despesas Empenhadas", "Despesas Pagas", "Ano Execução",
                  "Unidade Descentralizadora", "Região"]

    for col_idx, new_value in enumerate(new_header, start=1):
        sheet.cell(row=1, column=col_idx, value=new_value)

    print('Cabeçalho organizado na ordem correta...')
    
    # Formatar colunas "Dotação Atualizada", "Destaque Recebido", "Crédito Disponível", "Despesas Empenhadas" como contábeis e preencher células vazias com 0
    columns_to_format = [9, 10, 11, 12, 13]  # Colunas: Dotação Atualizada, Destaque Recebido, Crédito Disponível, Despesas Empenhadas
    for col_idx in columns_to_format:
        for row in sheet.iter_rows(min_row=2):
            cell = row[col_idx - 1]
            cell.number_format = '(* #,##0.00);(* (#,##0.00);(* "-"??);(@_)'
            if cell.value is None:
                cell.value = 0
                
    print('Formatação dos valores Numéricos executada...')
    
    # Preencher coluna "Ano Execução" com o ano atual
    ano_atual = datetime.now().year
    for row in sheet.iter_rows(min_row=2, min_col=col_ano_execucao, max_col=col_ano_execucao):
        cell = row[0]
        cell.value = ano_atual
        
    print('Coluna Ano Execução Preenchida...')
        
    # Carrega os dados da planilha "1- Apoio PTRES" do arquivo "Base Painel.xlsx"
    base_painel_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Base Painel.xlsx'
    base_painel_workbook = openpyxl.load_workbook(base_painel_file_path)
    base_painel_sheet = base_painel_workbook['1. Apoio PTRES']

    print('Base Consolidada Carregada para preenchimento das Colunas: Unidade Descentralizadora e Região...')
    print('Preenchimento em andamento aguarde...')
    
    # Preencher coluna "Unidade Descentralizada" usando a coluna PTRES
    for row in sheet.iter_rows(min_row=2):
        ptres_cell = row[6]  # Coluna PTRES
        unidade_cell = row[14]  # Coluna Unidade Descentralizadora
        for base_row in base_painel_sheet.iter_rows(min_row=2, max_row=base_painel_sheet.max_row, min_col=0, max_col=7):
            if str(base_row[0].value) == str(ptres_cell.value):
                unidade_cell.value = base_row[5].value
                break
    print('Unidade Descentralizadora Preenchida...')
            
    # Carrega os dados da planilha "2. Apoio UG" do arquivo "Base Painel.xlsx"
    base_painel_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Base Painel.xlsx'
    base_painel_workbook = openpyxl.load_workbook(base_painel_file_path)
    base_painel_sheet = base_painel_workbook['2. Apoio UG ']

    # Preencher coluna "Região" usando a coluna UG Executora - Código
    for row in sheet.iter_rows(min_row=2):
        executor_code = row[2]  # Coluna UG Executora - Código
        for base_row in base_painel_sheet.iter_rows(min_row=2, max_row=base_painel_sheet.max_row, min_col=1, max_col=8):
            if str(base_row[0].value) == str(executor_code.value):
                region_value = base_row[7]  # Coluna Região
                sheet.cell(row=row[0].row, column=col_regiao, value=region_value.value)
                break
    print('Coluna Região Preenchida...')

def carregar_processar_planilha(original_file_path, copy_file_path, target_uge_code, target_acao_governo):
    # Crie uma cópia do arquivo original para trabalhar
    shutil.copyfile(original_file_path, copy_file_path)

    # Abre o arquivo Excel
    workbook = openpyxl.load_workbook(copy_file_path)

    # Seleciona a planilha pelo nome
    sheet_name = 'Novo Painel SPO-TED'
    sheet = workbook[sheet_name]

    # Processa a planilha
    processar_planilha(sheet)

    # Salvar as alterações nos arquivos Excel
    workbook.save(copy_file_path)

    print("Preenchimento Executado com Sucesso...")

    # Filtrar pelo Orgão UGE - Código
    def filter_and_count_rows(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet_name = 'Novo Painel SPO-TED'
        sheet = workbook[sheet_name]
        
        print('Filtrando pela Coluna Órgão UGE - Código...')
        
        # Encontrar o índice da coluna "Órgão UGE - Código"
        org_uge_col_index = None
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Órgão UGE - Código":
                org_uge_col_index = col_idx
                break

        if org_uge_col_index is None:
            raise ValueError("Coluna 'Órgão UGE - Código' não encontrada.")

        # Converter a coluna para formato numérico
        for row in sheet.iter_rows(min_row=2, min_col=org_uge_col_index, max_col=org_uge_col_index):
            cell = row[0]
            cell.value = int(cell.value) if cell.value is not None else None

        print('Conversão para formato numérico realizada...')
        
        
        # Filtrar a coluna "Órgão UGE - Código" por target_uge_code
        filtered_rows = [row for row in sheet.iter_rows(min_row=2) if row[org_uge_col_index - 1].value == target_uge_code]

        # Excluir as linhas filtradas que também têm as ações de governo desejadas
        rows_to_delete = []
        for row in filtered_rows:
            if row[5].value in target_acao_governo:
                rows_to_delete.append(row[0].row)

        # Excluir as linhas do final para o começo para evitar problemas de referência
        for row_idx in reversed(rows_to_delete):
            sheet.delete_rows(row_idx)

        
        # Formatar colunas do cabeçalho
        header_format = Font(name='Calibri', size=11, bold=True, color="00000000")
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        # Formatar a linha 1 com fundo cinza claro e texto em negrito
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_format
            cell.fill = header_fill
            cell.border = header_border

        # Formatar colunas de A a N com fundo azul claro e texto preto
        for col_idx in range(1, 14):
            col_letter = get_column_letter(col_idx)
            col_range = f"{col_letter}2:{col_letter}{sheet.max_row}"
            
            for cell in sheet[col_range]:
                cell[0].font = header_format
                cell[0].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                cell[0].border = header_border

        # Formatar as 3 últimas colunas com fundo amarelo claro e texto preto
        for col_idx in range(sheet.max_column - 2, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_range = f"{col_letter}2:{col_letter}{sheet.max_row}"
            
            for cell in sheet[col_range]:
                cell[0].font = header_format
                cell[0].fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                cell[0].border = header_border

        workbook.save(file_path)

        print('Formatação Aplicada...')

    # Chama a função para filtrar e formatar
    filter_and_count_rows(copy_file_path)

print('Executando...')

# Caminho do arquivo Excel original
original_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Novo Painel SPO-TED.xlsx'

# Caminho do arquivo Excel copiado
copied_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Novo Painel SPO-TED.xlsx'

# Código do Órgão UGE para filtragem
target_uge_code = 26000

# Ações de governo desejadas
target_acao_governo = ['20RH','0A12']

print('Pegando as ações desejadas...')

# Processar a planilha e aplicar filtragens
carregar_processar_planilha(original_file_path, copied_file_path, target_uge_code, target_acao_governo)

# Caminho do arquivo original e de cópia
caminho_arquivo_original = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Repasse TED.xlsx'
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Repasse TED.xlsx'

def processar_arquivo(nome_arquivo):
    try:
        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(nome_arquivo)
        
        # Excluir a linha 2
        df = df.iloc[2:].reset_index(drop=True)
        
        # Excluir a última linha (iloc[-1])
        df = df.iloc[:-1]

        # Reordenar as colunas
        df = df[['TED', 'Descrição do Termo', 'Total Repassado', 'Estado Atual']]
        df.columns = ['TED', 'Descrição do Termo', 'Total Repassado', 'Estado Atual']

        # Substituir o texto na coluna 'Estado Atual'
        df['Estado Atual'] = df['Estado Atual'].str.replace('Relatório de cumprimento do objeto','RCO')
        
        # Substituir o texto na coluna 'Estado Atual'
        df['Estado Atual'] = df['Estado Atual'].str.replace('Relatório de cumprimento','RCO')

        # Definir as palavras-chave
        palavras_chave = ['Arquivado', 'Comprovado no SIAFI.', 'Termos Finalizados']

        # Filtrar e excluir as linhas que contenham as palavras-chave na coluna 'Estado Atual'
        df = df[~df['Estado Atual'].str.contains('|'.join(palavras_chave), case=False, na=False, regex=True)]

        print('Excluindo as ações indesejadas...')
        
        # Aplicar formatação numérica à coluna "Total Repassado" (formato brasileiro)
        df['Total Repassado'] = df['Total Repassado'].apply(lambda x:'{:,.2f}'.format(float(x)).replace(",", "_").replace(".", ",").replace("_", "."))

        # Salvar o DataFrame de volta no arquivo Excel
        df.to_excel(nome_arquivo, index=False)

        print("Executando...")
    except Exception as e:
        print(f"Ocorreu um erro: {str(e)}")

# Copiar o arquivo original para o arquivo de cópia
shutil.copy(caminho_arquivo_original, caminho_arquivo_copia)

# Chamar a função para processar o arquivo
processar_arquivo(caminho_arquivo_copia)

print('Preenchendo as vigências, aguarde...')

def calcular_situacao_vigencia(row):
    # Calcular a diferença entre a DATA HOJE e a Fim da Vigência em dias
    data_hoje = datetime.now()
    fim_vigencia = pd.to_datetime(row["Fim da Vigência"], format="%d/%m/%Y")
    dias_vencidos = (data_hoje - fim_vigencia).days

    # Definir a Situação Vigência com base nos dias vencidos
    if dias_vencidos == 0:
        return "VIGENTE", ""
    elif 0 < dias_vencidos <= 120:
        return "VENCIDO - 120", str(dias_vencidos)
    elif dias_vencidos > 120:
        return "VENCIDO + 120", str(dias_vencidos)
    else:
        return "VIGENTE", str(dias_vencidos)

def calcular_equivalencia_menos_120_dias(row):
    # Converter a coluna "DIAS VENCIDOS" para numérica (ignorando NaN)
    dias_vencidos = pd.to_numeric(row["DIAS VENCIDOS"], errors="coerce")
    
    # Verificar se os dias vencidos são menores que 120
    if pd.notna(dias_vencidos) and dias_vencidos < 120:
        return "VERDADEIRA"
    else:
        return "FALSO"

def copiar_dados_relatorio_entrega_rco(wb):
    # Caminho para o arquivo XLSX do Relatório de Entrega do RCO
    xlsx_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Relatório de Entrega do RCO.xlsx'

    # Carregar o arquivo XLSX em um DataFrame
    df_relatorio_rco = pd.read_excel(xlsx_file_path)

    # Criar uma nova aba no arquivo Excel
    ws = wb.create_sheet("Relatório Entrega RCO")

    # Copiar os dados do DataFrame para a nova aba, incluindo o cabeçalho
    for r_idx, row in enumerate(dataframe_to_rows(df_relatorio_rco, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

def adicionar_colunas_relatorio_entrega_rco(wb):
    # Carregar a segunda aba (Relatório Entrega RCO)
    sheet = wb["Relatório Entrega RCO"]
    
    # Adicionar as colunas "Situação", "DATA" e "DIAS" ao final da planilha
    sheet["D1"] = "Situação"
    sheet["E1"] = "DATA"
    sheet["F1"] = "DIAS"
    
    # Preencher a coluna "Situação" com "ENTREGUE"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.value = "ENTREGUE"
    
    # Preencher a coluna "DATA" com a data atual no formato DD/MM/AAAA
    data_atual = datetime.now().strftime("%d/%m/%Y")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.value = data_atual
    
    # Converter a coluna "Quando fez" para o formato de data "DD/MM/AAAA HH:MM:SS"
    for cell in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in cell:
            data_quando_fez = pd.to_datetime(cell.value, format="%Y-%m-%d %H:%M:%S", errors="coerce")
            if pd.notna(data_quando_fez):
                cell.value = data_quando_fez.strftime("%d/%m/%Y %H:%M:%S")
    
    # Calcular a coluna "DIAS" com base na diferença entre "DATA" e "Quando Fez", ignorando a hora
    data_hoje = datetime.now()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            data_quando_fez = pd.to_datetime(cell.value, format="%d/%m/%Y %H:%M:%S", errors="coerce")
            if pd.notna(data_quando_fez):
                dias_entregues = (data_hoje - data_quando_fez).days
                cell.offset(column=4).value = dias_entregues
                
def excluir_linhas_por_termos(df):
    # Termos a serem excluídos
    termos_a_excluir = ["comprovado no siafi.", "arquivado", "termo finalizado"]
    
    # Converter a coluna "Estado Atual" para minúsculas para fazer a comparação
    df["Estado Atual"] = df["Estado Atual"].str.lower()
    
    # Filtrar as linhas com base nos termos especificados na coluna "Estado Atual"
    df_filtrado = df[~df["Estado Atual"].isin(termos_a_excluir)]
    
    # Excluir as linhas onde a coluna "SIAFI" seja "-" ou vazio
    df_filtrado = df_filtrado[~df_filtrado["SIAFI"].isin(["", "-"])]
    
    return df_filtrado

      
def processarBaseCota():
    # Caminho para o arquivo original
    original_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Base Cota.xlsx'

    # Caminho para o novo arquivo em branco
    new_file_path = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

    # Carregar o arquivo original usando pandas
    df_base_cota = pd.read_excel(original_file_path)

    # Remover linhas vazias (aquelas em que todos os valores são nulos)
    df_base_cota.dropna(how='all', inplace=True)
    
    # Excluir as linhas com base nos termos especificados
    df_base_cota = excluir_linhas_por_termos(df_base_cota)

    # Criar um novo arquivo em branco do Excel
    wb = Workbook()
    ws = wb.active
    
    # Filtrar as linhas com base nos termos especificados na coluna "Estado Atual"
    termos_a_excluir = ["Comprovado no SIAFI.", "Arquivados", "Termo Finalizado"]
    df_base_cota = df_base_cota[~df_base_cota["Estado Atual"].str.lower().isin(termos_a_excluir)]

    # Criar a coluna "DATA HOJE" no DataFrame e preenchê-la com a data atual no formato DD/MM/AAAA
    df_base_cota["DATA HOJE"] = datetime.now().strftime("%d/%m/%Y")

    # Formatar a coluna "Fim da Vigência" para o formato de data brasileiro (DD/MM/AAAA)
    df_base_cota["Fim da Vigência"] = pd.to_datetime(df_base_cota["Fim da Vigência"], errors="coerce", format="%Y-%m-%d").dt.strftime("%d/%m/%Y")
    
    # Carregar a segunda aba (Relatório Entrega RCO)
    df_relatorio_rco = pd.read_excel(r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Relatório de Entrega do RCO.xlsx')

    # Calcular a coluna "Situação Vigência" e "DIAS VENCIDOS" com base nos dias vencidos
    df_base_cota[["Situação Vigência", "DIAS VENCIDOS"]] = df_base_cota.apply(calcular_situacao_vigencia, axis=1, result_type="expand")
    
    # Converter a coluna "DIAS VENCIDOS" para string
    df_base_cota["DIAS VENCIDOS"] = df_base_cota["DIAS VENCIDOS"].astype(str)
    
    # Aplicar a função ao DataFrame e criar a coluna "EQUIVALÊNCIA < 120 DIAS"
    df_base_cota["EQUIVALÊNCIA < 120 DIAS"] = df_base_cota.apply(calcular_equivalencia_menos_120_dias, axis=1)
    
    # Copiar os dados da primeira planilha para a nova aba
    for r_idx, row in enumerate(dataframe_to_rows(df_base_cota, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Chamar a função para adicionar o Relatório de Entrega do RCO como uma nova aba
    copiar_dados_relatorio_entrega_rco(wb)
    
    # Chamar a função para adicionar as colunas e fazer as alterações
    adicionar_colunas_relatorio_entrega_rco(wb)
    
    # Salvar o novo arquivo Excel com as novas abas
    wb.save(new_file_path)

# Chame a função para executar o código
processarBaseCota()

print('Aplicando as formatações...')

def marcar_entrega_rcos(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)

    # Selecione a primeira aba (Sheet)
    sheet = wb['Sheet']

    # Selecione a segunda aba (Relatório Entrega RCO)
    sheet_rco = wb['Relatório Entrega RCO']

    # nova coluna na aba "Sheet" chamada "RCO entregue 1 vez"
    sheet['R1'] = "RCO entregue 1 vez"
    
    # nova coluna na aba "Sheet" chamada "DIAS DE ENTREGA DO RCO"
    sheet['S1'] = "DIAS DE ENTREGA DO RCO"
    
    # nova coluna na aba "Sheet" chamada "EQUIVALENCIA > 180 DIAS"
    sheet['T1'] = "EQUIVALENCIA > 180 DIAS"

    # Crie um conjunto dos valores da coluna TED na aba "Relatório Entrega RCO"
    valores_ted_rco = set()
    ted_dias_mapping = {}

    for row in sheet_rco.iter_rows(min_row=2, max_row=sheet_rco.max_row, min_col=1, max_col=6, values_only=True):
        valores_ted_rco.add(row[0])
        ted_dias_mapping[row[0]] = row[5]

    # Preencha a coluna "RCO entregue 1 vez" e "DIAS DE ENTREGA DO RCO" na aba "Sheet"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=12, max_col=12):
        for cell in row:
            # Verifique se o valor da célula na coluna TED da "Sheet" está na lista de valores da coluna TED da "Relatório Entrega RCO"
            if cell.value in valores_ted_rco:
                cell_offset_rco = cell.offset(column=6)  # Coluna "RCO entregue 1 vez"
                cell_offset_rco.value = "ENTREGUE"
                
                cell_offset_dias = cell.offset(column=7)  # Coluna "DIAS DE ENTREGA DO RCO"
                cell_offset_dias.value = ted_dias_mapping[cell.value]

                # Verifique se DIAS DE ENTREGA DO RCO é maior que 180 e preencha a coluna "EQUIVALENCIA > 180 DIAS" adequadamente
                cell_offset_equivalencia = cell.offset(column=8)  # Coluna "EQUIVALENCIA > 180 DIAS"
                if ted_dias_mapping[cell.value] > 180:
                    cell_offset_equivalencia.value = "VERDADEIRO"
                else:
                    cell_offset_equivalencia.value = "FALSO"
            else:
                cell_offset_rco = cell.offset(column=6)  # Coluna "RCO entregue 1 vez"
                cell_offset_rco.value = "NÃO ENTREGUE"
                
                cell_offset_dias = cell.offset(column=7)  # Coluna "DIAS DE ENTREGA DO RCO"
                cell_offset_dias.value = None
                
                cell_offset_equivalencia = cell.offset(column=8)  # Coluna "EQUIVALENCIA > 180 DIAS"
                cell_offset_equivalencia.value = None

                # Preencha com "0" (Zero) na coluna "DIAS DE ENTREGA DO RCO" e "FALSO" na coluna "EQUIVALENCIA > 180 DIAS"
                cell_offset_dias.value = 0
                cell_offset_equivalencia.value = "FALSO"

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função e passe o caminho do arquivo copia como argumento
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
marcar_entrega_rcos(caminho_arquivo_copia)

print('Preenchendo de acordo com as condicionais...')

def adicionar_colunas_e_preencher_situacao_rcos(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)
    
    # Selecione a primeira aba (Sheet)
    sheet = wb.active

    # Adicione as colunas "Situação RCO descentralizada" e "Situação RCO descentralizadora"
    sheet['U1'] = "Situação RCO descentralizada"
    sheet['V1'] = "Situação RCO descentralizadora"

    # Percorra as linhas da aba "Sheet" a partir da segunda linha (assumindo que a primeira linha contém os cabeçalhos)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=23):
        situacao_vigencia = row[14].value  # Coluna "Situação Vigência" (índice 3, correspondente à coluna U)
        rco_entregue_1_vez = row[17].value   # Coluna "RCO entregue 1 vez" (índice 4, correspondente à coluna V)
        estado_atual = row[7].value  # Coluna "Estado Atual" (índice 7, correspondente à coluna H)
        equivalencia_maior180 = row[19].value

        situacao_rco_descentralizada = ""
        situacao_rco_descentralizadora = ""

        if estado_atual not in ["termo em execução", "TERMO EM EXECUÇÃO"] and situacao_vigencia == "VIGENTE" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE"
            
        elif estado_atual in ["termo em execução", "TERMO EM EXECUÇÃO"] and situacao_vigencia == "VIGENTE" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "EM EXECUÇÃO"
            situacao_rco_descentralizadora = "EM EXECUÇÃO"
            
        elif equivalencia_maior180 in "VERDADEIRO" and situacao_vigencia == "VENCIDO - 120" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE + 180 DIAS"
            
        elif equivalencia_maior180 in "FALSO" and situacao_vigencia == "VENCIDO - 120" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE + 180 DIAS"
            
        elif equivalencia_maior180 in "VERDADEIRO" and situacao_vigencia == "VENCIDO + 120" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE + 180 DIAS"
        
        elif equivalencia_maior180 in "FALSO" and situacao_vigencia == "VENCIDO - 120" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE"    
                    
        elif situacao_vigencia == "VIGENTE" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE"
            
        elif situacao_vigencia == "VIGENTE" and rco_entregue_1_vez == "NÃO ENTREGUE":
            situacao_rco_descentralizada = "EM EXECUÇÃO"
            situacao_rco_descentralizadora = "EM EXECUÇÃO"
            
        elif situacao_vigencia == "VENCIDO - 120" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE"
            
        elif situacao_vigencia == "VENCIDO - 120" and rco_entregue_1_vez == "NÃO ENTREGUE":
            situacao_rco_descentralizada = "RCO NÃO ENTREGUE"
            situacao_rco_descentralizadora = "AGUARDANDO RCO"
        
        elif situacao_vigencia == "VENCIDO + 120" and rco_entregue_1_vez == "NÃO ENTREGUE":
            situacao_rco_descentralizada = "VENCIDO"
            situacao_rco_descentralizadora = "RCO NÃO ENTREGUE"
        
        elif situacao_vigencia == "VENCIDO + 120" and rco_entregue_1_vez == "ENTREGUE":
            situacao_rco_descentralizada = "RCO ENTREGUE"
            situacao_rco_descentralizadora = "RCO EM ANÁLISE"        

        # Preencher as novas colunas com os valores calculados
        sheet.cell(row=row[0].row, column=21, value=situacao_rco_descentralizada)
        sheet.cell(row=row[0].row, column=22, value=situacao_rco_descentralizadora)

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função para adicionar as colunas e preencher as condições na aba "Sheet"
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
adicionar_colunas_e_preencher_situacao_rcos(caminho_arquivo_copia)

print('Vigências Preenchidas...')

def formatar_valores_negativos(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)

    # Selecione a primeira aba (Sheet)
    sheet = wb.active

    # Percorra as linhas da coluna "Valor Descentralizado (R$)" e formate os valores negativos
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=13, max_col=13):
        for cell in row:
            # Verifique se o valor da célula não é nulo e começa com um sinal de menos após remover espaços em branco
            if cell.value is not None and str(cell.value).strip().startswith('-'):
               # Formate o valor mantendo o sinal de menos, com centavos separados por vírgula e outras casas decimais por ponto
                formatted_value = '-' + '{:,.2f}'.format(abs(cell.value)).replace(',', '|').replace('.', ',').replace('|', '.')
                # Atualize o valor da célula
                cell.value = formatted_value

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função para formatar os valores negativos da coluna "Valor Descentralizado (R$)"
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
formatar_valores_negativos(caminho_arquivo_copia)

def extrair_numeros_e_salvar(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)

    # Acesse a aba "Sheet"
    sheet = wb['Sheet']

    # Percorra as linhas das colunas "Concedente" e "Proponente"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=10):
        # Verifique se os valores das células não são nulos
        concedente = row[0].value
        proponente = row[1].value

        if concedente and '/' in concedente:
            # Extraia os 6 dígitos antes da barra
            numeros_concedente = concedente.split('/')[0].strip()[:6]
            # Atualize o valor da célula
            sheet.cell(row=row[0].row, column=6, value=numeros_concedente)

        if proponente and '/' in proponente:
            # Extraia os 6 dígitos antes da barra
            numeros_proponente = proponente.split('/')[0].strip()[:6]
            # Atualize o valor da célula
            sheet.cell(row=row[1].row, column=7, value=numeros_proponente)

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função para executar a extração e salvar as duas abas
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
extrair_numeros_e_salvar(caminho_arquivo_copia)

print('Números Extraidos...')

def criar_aba_base_cota(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)

    # Acesse a aba "Sheet"
    sheet = wb['Sheet']

    # Crie uma nova aba chamada "BASE COTA"
    nova_aba = wb.create_sheet('BASE COTA')

    # Defina o cabeçalho personalizado da nova aba "BASE COTA"
    cabecalho = ["TED", "Descentralizada", "Descentralizadora", "Exercício (Orçamentário)" ,"Ação (Orçamentário)" , "PTRES", "Valor Descentralizado (R$) (Orçamentário)",
                "Situação Vigência", "Situação RCO descentralizada", "Situação RCO descentralizadora", "Fim da Vigência"]
    
    # Copie o cabeçalho para a nova aba "BASE COTA"
    nova_aba.append(cabecalho)

    # Copie os dados das colunas específicas da aba "Sheet" para a nova aba "BASE COTA"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        dados_filtrados = [row[i - 1] for i in [12, 6, 7, 2, 1, 5, 13, 15, 21, 22, 10,]]
        nova_aba.append(dados_filtrados)

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função para criar a nova aba "BASE COTA" com as colunas desejadas
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
criar_aba_base_cota(caminho_arquivo_copia)

print('Criando a aba BASE COTA...')

def criar_colunas_personalizadas(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)

    # Acesse a aba "BASE COTA"
    aba_base_cota = wb['BASE COTA']

    # Crie os cabeçalhos das colunas personalizadas
    cabecalhos = ['Unidade Responsável - SIGLA', 'AÇÃO - Código + Nome Reduzido', 'UG Descentralizadora - Código + Nome', 'GRUPO']
    for col_num, cabecalho in enumerate(cabecalhos, 12):  # Começando na coluna M (13)
        aba_base_cota.cell(row=1, column=col_num, value=cabecalho)

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função para criar as colunas personalizadas
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
criar_colunas_personalizadas(caminho_arquivo_copia)


def formatar_coluna_ptres(file_path):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(file_path)

    # Acesse a aba "BASE COTA"
    aba_base_cota = wb['BASE COTA']

    # Percorra as células da coluna PTRES (coluna F)
    for row in aba_base_cota.iter_rows(min_row=2, max_row=aba_base_cota.max_row, min_col=6, max_col=6):
        for cell in row:
            # Obtenha o valor da célula como string
            valor_ptres = str(cell.value)

            # Adicione zeros à esquerda para garantir que a string tenha 6 caracteres
            valor_ptres = valor_ptres.zfill(6)

            # Atualize o valor da célula
            cell.value = valor_ptres

    # Salve as alterações no arquivo Excel
    wb.save(file_path)

# Chame a função para formatar a coluna PTRES
caminho_arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
formatar_coluna_ptres(caminho_arquivo_copia)

print('Formatando a Coluna PTRES com 6 digitos...')

def copiar_abas_base_painel_para_copia_base_cota(base_painel_file, copia_base_cota_file):
    # Abra os arquivos Excel
    wb_base_painel = openpyxl.load_workbook(base_painel_file)
    wb_copia_base_cota = openpyxl.load_workbook(copia_base_cota_file)

    # Acesse as abas relevantes em ambos os arquivos
    aba_apoio_ptres_base_painel = wb_base_painel[wb_base_painel.sheetnames[0]]  # Primeira aba
    aba_apoio_ug_base_painel = wb_base_painel[wb_base_painel.sheetnames[1]]    # Segunda aba

    # Crie abas "APOIO PTRES" e "APOIO UG" no arquivo "COPIA Base Cota"
    if "APOIO PTRES" not in wb_copia_base_cota.sheetnames:
        wb_copia_base_cota.create_sheet("APOIO PTRES")
    if "APOIO UG" not in wb_copia_base_cota.sheetnames:
        wb_copia_base_cota.create_sheet("APOIO UG")

    # Acesse as abas criadas no arquivo "COPIA Base Cota"
    aba_apoio_ptres_copia_base_cota = wb_copia_base_cota["APOIO PTRES"]
    aba_apoio_ug_copia_base_cota = wb_copia_base_cota["APOIO UG"]

    # Copie os dados das abas "APOIO PTRES" e "APOIO UG" do arquivo "Base Painel.xlsx" para as abas correspondentes do arquivo "COPIA Base Cota.xlsx"
    for row in aba_apoio_ptres_base_painel.iter_rows(min_row=1, max_row=aba_apoio_ptres_base_painel.max_row):
        aba_apoio_ptres_copia_base_cota.append([cell.value for cell in row])

    for row in aba_apoio_ug_base_painel.iter_rows(min_row=1, max_row=aba_apoio_ug_base_painel.max_row):
        aba_apoio_ug_copia_base_cota.append([cell.value for cell in row])

    # Salve as alterações no arquivo "COPIA Base Cota.xlsx"
    wb_copia_base_cota.save(copia_base_cota_file)

# Caminhos para os arquivos Excel
caminho_arquivo_painel = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\Base Painel.xlsx'
caminho_arquivo_copia_base_cota = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

# Chame a função para copiar as abas
copiar_abas_base_painel_para_copia_base_cota(caminho_arquivo_painel, caminho_arquivo_copia_base_cota)

print('Copiando APOIO UG e APOIO PTRES...')

def preencher_unidade_responsavel_sigla(base_cota_file):
    # Abra o arquivo Excel
    wb_base_cota = openpyxl.load_workbook(base_cota_file)
    
    # Acesse as abas relevantes
    aba_base_cota = wb_base_cota['BASE COTA']
    aba_apoio_ug = wb_base_cota['APOIO UG']

    # Crie um dicionário para mapear os códigos da UG Executora às siglas
    ug_executora_mapping = {}
    for row in aba_apoio_ug.iter_rows(min_row=2, max_row=aba_apoio_ug.max_row, min_col=1, max_col=3):
        ug_codigo = str(row[0].value)
        ug_sigla = row[2].value
        ug_executora_mapping[ug_codigo] = ug_sigla

    # Preencha a coluna "Unidade Responsável - SIGLA" na aba "BASE COTA"
    for row in aba_base_cota.iter_rows(min_row=2, max_row=aba_base_cota.max_row, min_col=3, max_col=3):
        for cell in row:
            # Obtenha o valor da célula da coluna "Concedente" como string
            concedente = str(cell.value)

            # Verifique se o código está no mapeamento (correspondência parcial)
            ug_sigla = None
            for ug_codigo, sigla in ug_executora_mapping.items():
                if ug_codigo in concedente:
                    ug_sigla = sigla
                    break
            
            if ug_sigla:
                cell_offset = 9  # Offset para a coluna "Unidade Responsável - SIGLA"
                cell_offset = cell_offset if cell_offset <= aba_base_cota.max_column else aba_base_cota.max_column
                cell_offset = cell_offset if cell_offset > 0 else 1
                cell_to_update = row[0].offset(column=cell_offset)
                cell_to_update.value = ug_sigla

    # Salve as alterações no arquivo Excel
    wb_base_cota.save(base_cota_file)

# Caminho para o arquivo Excel
caminho_arquivo_cota = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

# Chame a função para preencher a coluna "Unidade Responsável - SIGLA"
preencher_unidade_responsavel_sigla(caminho_arquivo_cota)

print('Preenchendo Unidade Responsável...')

def preencher_ug_descentralizada_grupo(base_cota_file):
    # Abra o arquivo Excel
    wb_base_cota = openpyxl.load_workbook(base_cota_file)
    
    # Acesse as abas relevantes
    aba_base_cota = wb_base_cota['BASE COTA']
    aba_apoio_ug = wb_base_cota['APOIO UG']

    # Crie um dicionário para mapear os códigos da UG Executora com informações relevantes
    ug_executora_mapping = {}
    for row in aba_apoio_ug.iter_rows(min_row=2, max_row=aba_apoio_ug.max_row, min_col=1, max_col=6):
        ug_codigo = str(row[0].value)
        ug_nome_codigo = row[5].value
        ug_grupo = row[4].value
        ug_executora_mapping[ug_codigo] = (ug_nome_codigo, ug_grupo)

    # Preencha as colunas "UG Descentralizadora - Código + Nome" (Coluna N) e "GRUPO" (Coluna O) na aba "BASE COTA"
    for row in aba_base_cota.iter_rows(min_row=2, max_row=aba_base_cota.max_row, min_col=2, max_col=2):
        for cell in row:
            # Obtenha o valor da célula da coluna "Proponente" como string
            proponente = str(cell.value)
            
            # Verifique se o código está no mapeamento (correspondência parcial)
            if proponente in ug_executora_mapping:
                ug_nome_codigo, ug_grupo = ug_executora_mapping[proponente]

                # Preencha a coluna "UG Descentralizadora - Código + Nome" (Coluna N)
                cell_offset = 12  # Offset para a coluna "UG Descentralizadora - Código + Nome"
                cell_offset = cell_offset if cell_offset <= aba_base_cota.max_column else aba_base_cota.max_column
                cell_offset = cell_offset if cell_offset > 0 else 1
                cell_to_update = row[0].offset(column=cell_offset)
                cell_to_update.value = ug_nome_codigo

                # Preencha a coluna "GRUPO" (Coluna O)
                cell_offset = 13  # Offset para a coluna "GRUPO"
                cell_offset = cell_offset if cell_offset <= aba_base_cota.max_column else aba_base_cota.max_column
                cell_offset = cell_offset if cell_offset > 0 else 1
                cell_to_update = row[0].offset(column=cell_offset)
                cell_to_update.value = ug_grupo

    # Salve as alterações no arquivo Excel
    wb_base_cota.save(base_cota_file)

# Caminho para o arquivo Excel
caminho_arquivo_cota = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

# Chame a função para preencher as colunas "UG Descentralizada - Código + Nome" e "GRUPO"
preencher_ug_descentralizada_grupo(caminho_arquivo_cota)

print('Preenchendo UG Descentralizadora...')

def preencher_acao_nome_reduzido(base_cota_file):
    # Abra o arquivo Excel
    wb_base_cota = openpyxl.load_workbook(base_cota_file)

    # Acesse as abas relevantes
    aba_base_cota = wb_base_cota['BASE COTA']
    aba_apoio_ptres = wb_base_cota['APOIO PTRES']

    # Crie um dicionário para mapear os códigos do PTRES para a Ação - Código + Nome Reduzido
    ptres_mapping = {}
    for row in aba_apoio_ptres.iter_rows(min_row=2, max_row=aba_apoio_ptres.max_row, min_col=1, max_col=3):
        ptres_cod = str(row[0].value)
        acao_cod = str(row[2].value)
        ptres_mapping[ptres_cod] = acao_cod

    # Pegando o parâmetro do PTRES
    for row in aba_base_cota.iter_rows(min_row=2, max_row=aba_base_cota.max_row, min_col=6, max_col=6):
        for cell in row:
            # Obter o valor do PTRES da aba BASE COTA como string
            ptres = str(cell.value)
            
            # Verifique se o código está no mapeamento (correspondência parcial)
            acao_cod = None
            for ptres_cod, acao in ptres_mapping.items():
                if ptres_cod in ptres:
                    acao_cod = acao
                    break
        
            # Preencher a coluna "AÇÃO - Código + Nome Reduzido"
            cell_offset = 7
            cell_offset = cell_offset if cell_offset <= aba_base_cota.max_column else aba_base_cota.max_column
            cell_offset = cell_offset if cell_offset > 0 else 1
            cell_to_update = cell.offset(column=cell_offset)
            cell_to_update.value = acao_cod

    # Salve as alterações no arquivo Excel
    wb_base_cota.save(base_cota_file)

# Caminho para o arquivo Excel
caminho_arquivo_cota = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

print('Preenchendo Ação...')

# Chame a função para preencher a coluna "AÇÃO - Código + Nome Reduzido"
preencher_acao_nome_reduzido(caminho_arquivo_cota)


def apagar_linhas(base_cota_file):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(base_cota_file)
    
    # Acesse a aba "BASE COTA"
    aba_base_cota = wb['BASE COTA']
    
    # Descobrir a última linha com dados
    ultima_linha = aba_base_cota.max_row
    
    # Coletar linhas a serem deletadas
    linhas_para_deletar = []
    
    # Percorrer as linhas para identificar as linhas com "00none"
    for row in range(2, ultima_linha + 1):
        valor_ptres = aba_base_cota.cell(row=row, column=6).value
        if valor_ptres == "00None":
            linhas_para_deletar.append(row)
    
    # Deletar as linhas coletadas de baixo para cima
    for row in reversed(linhas_para_deletar):
        aba_base_cota.delete_rows(row)
    
    # Remover a última linha
    aba_base_cota.delete_rows(aba_base_cota.max_row)
    
    # Salvar as alterações no arquivo Excel
    wb.save(base_cota_file)

# Caminho para o arquivo Excel
caminho_arquivo_cota = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

# Chame a função para apagar as linhas
apagar_linhas(caminho_arquivo_cota)

print('Criando tabela dinamica...')


def criar_tabela_dinamica(base_cota_file):
    # Ler o arquivo Excel
    df = pd.read_excel(base_cota_file, sheet_name='BASE COTA')
    
    # Converter a coluna de totais para numérico, substituindo erros por 0
    df["Valor Descentralizado (R$) (Orçamentário)"] = pd.to_numeric(df["Valor Descentralizado (R$) (Orçamentário)"], errors='coerce').fillna(0)
    
    # Definir o cabeçalho personalizado
    cabecalho = ["TED", "Descentralizada", "Descentralizadora", "Exercício (Orçamentário)", "Ação (Orçamentário)", "PTRES", "Valor Descentralizado (R$) (Orçamentário)",
                 "Situação Vigência", "Situação RCO descentralizada", "Situação RCO descentralizadora", "Fim da Vigência",'Unidade Responsável - SIGLA', 'AÇÃO - Código + Nome Reduzido', 'UG Descentralizadora - Código + Nome',	'GRUPO'
                 ]

    # Verificar se todas as colunas do cabeçalho estão presentes no dataframe
    for col in cabecalho:
        if col not in df.columns:
            df[col] = None  # Adicionar a coluna se estiver faltando
    
    # Preencher células vazias com um valor único para garantir que sejam incluídas no agrupamento
    df.fillna('NULL', inplace=True)        

    # Agrupar as linhas idênticas e somar os totais
    colunas_para_agrupamento = [col for col in df.columns if col != "Valor Descentralizado (R$) (Orçamentário)"]
    df_agrupado = df.groupby(colunas_para_agrupamento)["Valor Descentralizado (R$) (Orçamentário)"].sum().reset_index()
    
    # Reverter os valores 'NULL' para células vazias
    df_agrupado.replace('NULL', '', inplace=True)
    
    # Reordenar as colunas de acordo com o cabeçalho personalizado
    df_agrupado = df_agrupado[cabecalho]
    
    # Escrever a tabela dinâmica resultante de volta no arquivo Excel
    with pd.ExcelWriter(base_cota_file, engine='openpyxl', mode='a') as writer:
        df_agrupado.to_excel(writer, sheet_name='BASE COTA 2', index=False)

# Caminho para o arquivo Excel
caminho_arquivo_cota = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'

# Chame a função para criar a tabela dinâmica
criar_tabela_dinamica(caminho_arquivo_cota)

print('Gerando os arquivos finais...')


def copiar_arquivo_base_cota():
    origem = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Base Cota.xlsx'
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\PAINEL ARQUIVO FINAL.xlsx'  # Nome do arquivo de destino

    try:
        shutil.copy(origem, destino)
        print(f'Arquivo copiado com sucesso para {destino}')
    except Exception as e:
        print(f'Ocorreu um erro ao copiar o arquivo: {e}')

# Chame a função para copiar o arquivo
copiar_arquivo_base_cota()

def copiar_arquivo_repasse_ted():
    origem = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Repasse TED.xlsx'
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\REPASSE ARQUIVO FINAL.xlsx'  # Nome do arquivo de destino

    try:
        shutil.copy(origem, destino)
        print(f'Arquivo copiado com sucesso para {destino}')
    except Exception as e:
        print(f'Ocorreu um erro ao copiar o arquivo: {e}')

# Chame a função para copiar o arquivo
copiar_arquivo_repasse_ted()

def copiar_arquivo_base_consolidada():
    origem = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\COPIA Novo Painel SPO-TED.xlsx'
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Painel\BASE CONSOLIDADA ARQUIVO FINAL.xlsx'  # Nome do arquivo de destino

    try:
        shutil.copy(origem, destino)
        print(f'Arquivo copiado com sucesso para {destino}')
    except Exception as e:
        print(f'Ocorreu um erro ao copiar o arquivo: {e}')

# Chame a função para copiar o arquivo
copiar_arquivo_base_consolidada()

print("Relatórios do painel gerados com Sucesso. os Arquivos se encontram em: W:\B - TED\7 - AUTOMAÇÃO\Painel\PAINEL ARQUIVO FINAL.xlsx, W:\B - TED\7 - AUTOMAÇÃO\Painel\BASE CONSOLIDADA ARQUIVO FINAL.xlsx e W:\B - TED\7 - AUTOMAÇÃO\Painel\REPASSE ARQUIVO FINAL.xlsx")
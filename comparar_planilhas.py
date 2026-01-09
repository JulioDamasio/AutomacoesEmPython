import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def comparar_planilhas():
    # Caminho da pasta onde estão os arquivos
    pasta = r"W:\B - TED\7 - AUTOMAÇÃO\PF E NC comparativo"

    # Nomes dos arquivos e colunas correspondentes
    arquivo_base = r"W:\B - TED\7 - AUTOMAÇÃO\PF E NC comparativo\PF Legado - Exercício 2025.xlsx"
    arquivo_comparar = r"W:\B - TED\7 - AUTOMAÇÃO\PF E NC comparativo\PFS 2025.xlsx"

    coluna_base = "Emissao - Dia"
    coluna_comparar = "Data de Emissão DOC. PF"

    # Carregar as planilhas (considera que as datas estão na primeira planilha de cada arquivo)
    df_base = pd.read_excel(arquivo_base)
    df_comp = pd.read_excel(arquivo_comparar)

    # Converter as colunas de data para o mesmo formato (garante comparação correta)
    df_base[coluna_base] = pd.to_datetime(df_base[coluna_base], errors='coerce')
    df_comp[coluna_comparar] = pd.to_datetime(df_comp[coluna_comparar], errors='coerce')

    # Identificar datas que estão na base mas não na comparação
    datas_faltantes = set(df_base[coluna_base]) - set(df_comp[coluna_comparar])

    print(f"Datas que estão na planilha base e não estão na PFS 2025: {len(datas_faltantes)} encontradas.")

    # Reabrir com openpyxl para colorir as células faltantes em vermelho
    wb = load_workbook(arquivo_base)
    ws = wb.active

    # Criar cor de preenchimento
    vermelho = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # Pinta as células da coluna onde a data estiver ausente
    for row in range(2, ws.max_row + 1):  # começa na linha 2 (ignorando o cabeçalho)
        cell = ws.cell(row=row, column=1)  # 1 = primeira coluna
        try:
            data = pd.to_datetime(cell.value)
            if data in datas_faltantes:
                cell.fill = vermelho
        except Exception:
            pass

    # Salvar nova planilha destacada
    arquivo_saida = f"{pasta}\\PF_Legado_com_Faltas_Marcadas.xlsx"
    wb.save(arquivo_saida)

    print(f"✅ Comparação concluída! Arquivo gerado: {arquivo_saida}")

# Chama a função para rodar o script
if __name__ == "__main__":
    comparar_planilhas()

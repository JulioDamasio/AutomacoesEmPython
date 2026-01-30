import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
import pandas as pd
import numpy as np
from datetime import datetime
import gc, time

def copiar_arquivos(arquivo_origem, destino):
    origens = [
        r'W:\B - TED\7 - AUTOMAÇÃO\NC e PF\NC funcionando - EXERCÍCIO 2024.xlsx',
        r'W:\B - TED\7 - AUTOMAÇÃO\NC e PF\PF Legado - Exercício 2024.xlsx'
    ]
    for origem in origens:
        caminho, nome_arquivo = os.path.split(origem)
        shutil.copy(origem, os.path.join(destino, nome_arquivo))  # Copia para a pasta destino
    caminho, nome_arquivo = os.path.split(arquivo_origem)
    novo_nome = "COPIA " + nome_arquivo
    arquivo_copia = os.path.join(caminho, novo_nome)
    shutil.copy(arquivo_origem, os.path.join(destino, novo_nome))  # Copia o arquivo original para a pasta destino com o novo nome
    return arquivo_copia

def criar_tabela_dinamica(arquivo_origem_ted, arquivo_destino_ted):
    try:
        # Fazendo uma cópia do arquivo de origem
        df = pd.read_excel(arquivo_origem_ted, header=[0, 1])  # Lendo as duas primeiras linhas como cabeçalho
        df.to_excel(arquivo_destino_ted, index=False)

        # Lendo o arquivo recém-criado
        df_copia = pd.read_excel(arquivo_destino_ted)

        # Criando a tabela dinâmica manualmente
        pivot_table = df_copia.groupby(level=0, axis=1).sum()

        # Salvando a tabela dinâmica sem índice multinível
        with pd.ExcelWriter(arquivo_destino_ted, mode='a', engine='openpyxl') as writer:
            pivot_table.to_excel(writer, sheet_name='TEDS')

        print("Tabela dinâmica criada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        
def criar_tabela_dinamica(arquivo_origem_ted, arquivo_destino_ted):
    try:
        # Fazendo uma cópia do arquivo de origem
        df = pd.read_excel(arquivo_origem_ted, header=[0, 1])  # Lendo as duas primeiras linhas como cabeçalho

        # Reestruturando o DataFrame para ter apenas um nível de colunas
        df.columns = ['_'.join(col[:1]).strip() for col in df.columns.values]

        # Criando a tabela dinâmica manualmente
        pivot_table = df.groupby(level=0, axis=1).sum()

        # Salvando a tabela dinâmica sem índice multinível
        with pd.ExcelWriter(arquivo_destino_ted, mode='w', engine='openpyxl') as writer:
            pivot_table.to_excel(writer, sheet_name='TEDS', index=False)

        print("Tabela dinâmica criada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        
def limpar_planilhaTEDS_e_filtrar_siafi(arquivo_destino_ted):
    try:
        print("Carregando o arquivo Excel e carregando a aba 'TEDS'...")
        df_TEDS = pd.read_excel(arquivo_destino_ted, sheet_name='TEDS')
        print("Arquivo Excel carregado e aba 'TEDS' selecionada com sucesso.")

        print("Removendo a última linha...")
        df_TEDS.drop(df_TEDS.tail(1).index, inplace=True)
        print("Última linha removida com sucesso...")

        print("Filtrando as linhas...")

        # Filtrando as linhas onde o SIAFI não é igual a '-'
        df_TEDS = df_TEDS[df_TEDS['SIAFI'] != "-"]

        print("Salvando as alterações no arquivo Excel...")

        # Sobrescrevendo completamente a aba 'Planilha2' com os dados filtrados
        with pd.ExcelWriter(arquivo_destino_ted, engine='openpyxl') as writer:
            df_TEDS.to_excel(writer, sheet_name='TEDS', index=False)

        print("Planilha TEDS limpa e filtrada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        
def dividir_coluna_ug_concedente_e_Descentralizada(arquivo_destino_ted):
    try:
        # Carregando o arquivo Excel
        df = pd.read_excel(arquivo_destino_ted)

        # Dividindo a coluna 'Concedente' em três novas colunas
        pattern = r'^(\d+) \/ ([^-]+) - (.+)$'
        df[['UG Descentralizadora', 'Descrição Descentralizadora', 'Sigla Descentralizadora']] = df['Descentralizadora'].str.extract(pattern)

        # Removendo a coluna original 'Concedente'
        df.drop(columns=['Descentralizadora'], inplace=True)
        
        # Dividindo a coluna 'Descentralizada' em três novas colunas
        pattern = r'^(\d+) \/ (\d+) \/ (.+)$'
        df[['UG Descentralizada', 'Gestão Descentralizada', 'Sigla Descentralizada']] = df['Descentralizada'].str.extract(pattern)
        
        # Preenchendo os valores da coluna 'Gestão Descentralizada' com 5 caracteres
        df['Gestão Descentralizada'] = df['Gestão Descentralizada'].apply(lambda x: str(x).zfill(5))

        # Removendo a coluna original 'Descentralizada'
        df.drop(columns=['Descentralizada'], inplace=True)

        # Salvando as alterações no arquivo Excel
        with pd.ExcelWriter(arquivo_destino_ted, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='TEDS', index=False)

        print("Colunas 'Descentralizadora' e 'Descentralizada' divididas em três colunas cada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def criar_coluna_id_ug_gestao(arquivo_destino_ted):
    try:
        # Carregando o arquivo Excel
        df = pd.read_excel(arquivo_destino_ted)
        
        # Preenchendo os valores da coluna 'Gestão Descentralizada' com 5 caracteres
        df['Gestão Descentralizada'] = df['Gestão Descentralizada'].apply(lambda x: str(x).zfill(5))

        # Preenchendo a coluna 'ID UG Gestão' com a concatenação de 'UG Descentralizada' e 'Gestão Descentralizada'
        df['ID UG Gestão'] = df['Descentralizada'].astype(str).str.zfill(5) + df['Gestão Descentralizada'].astype(str).str.zfill(5)

        # Salvando os dados em uma nova planilha temporária
        arquivo_temporario = os.path.splitext(arquivo_destino_ted)[0] + '_temp.xlsx'
        with pd.ExcelWriter(arquivo_temporario) as writer:
            df.to_excel(writer, index=False, sheet_name='TEDS')  # Definindo o nome da planilha como 'TEDS'

        # Substituindo o arquivo original pelo temporário
        os.replace(arquivo_temporario, arquivo_destino_ted)

        print("Coluna 'ID UG Gestão' criada e preenchida com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def copiar_e_sobrescrever_arquivo(arquivo_destino_ted, arquivo_ted_final):
    try:
        # Copiando o arquivo de origem para o destino
        shutil.copyfile(arquivo_destino_ted, arquivo_ted_final)

        print("Arquivo copiado com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def inserir_sinal_negativo(arquivo_origem, arquivo_destino):
    # Ler o arquivo Excel
    df = pd.read_excel(arquivo_origem)

    # Aplicar a condição onde a coluna "Operação" esteja preenchida com "( - )"
    linhas_filtradas = df[df['Operação'] == '( - )']

    # Colocar um sinal de negativo "-" na frente do valor na coluna "Valor Total NC"
    linhas_filtradas['Valor Total NC'] = -linhas_filtradas['Valor Total NC']

    # Substituir as linhas filtradas no DataFrame original
    df.loc[df['Operação'] == '( - )', 'Valor Total NC'] = linhas_filtradas['Valor Total NC']

    # Salvar o DataFrame de volta para o arquivo Excel
    df.to_excel(arquivo_destino, index=False)
    
def inserir_sinal_negativo_pf(arquivo_origem, arquivo_destino):
    # Ler o arquivo Excel
    df = pd.read_excel(arquivo_origem)

    # Aplicar a condição onde a coluna "Operação" esteja preenchida com "( - )"
    linhas_filtradas = df[df['Operação'] == '(-)']

    # Colocar um sinal de negativo "-" na frente do valor na coluna "Valor Total NC"
    linhas_filtradas['Valor Doc. PF (R$)'] = -linhas_filtradas['Valor Doc. PF (R$)']

    # Substituir as linhas filtradas no DataFrame original
    df.loc[df['Operação'] == '(-)', 'Valor Doc. PF (R$)'] = linhas_filtradas['Valor Doc. PF (R$)']

    # Salvar o DataFrame de volta para o arquivo Excel
    df.to_excel(arquivo_destino, index=False)

def apagar_segunda_e_ultima_linha(arquivo_excel):
    try:
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(arquivo_excel)
        sheet = wb.active

        # Apagar a segunda linha
        sheet.delete_rows(2)

        # Apagar a última linha
        sheet.delete_rows(sheet.max_row)

        # Salvar as alterações
        wb.save(arquivo_excel)
        wb.close()

        print("Segunda e última linha apagadas com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao apagar as linhas: {e}")

def copiar_e_apagar_linhas(arquivo_origem, arquivo_destino): 
    try: 
        # Copiar o arquivo para o destino 
        shutil.copyfile(arquivo_origem, arquivo_destino) 
        
        # Carregar o arquivo copiado com o pandas 
        df = pd.read_excel(arquivo_destino) 
        
        # 🔥 CORTA para 30 colunas
        df = df.iloc[:, :30]
        
        # Renomear as colunas 
        novo_cabecalho = [ "Resultado EOF", "Descrição EOF", "NE CCor - Ano Emissão", "Órgão UGE", "Descrição UGE", "UG Executora", "Descrição UG", "UGE - UG Setorial Financeira", "Descrição Setorial", "Ação Governo", "Descrição Ação", "Função PO", "SubFunção PO", "Programa PO", "Cod PO", "Descrição PO", "PTRES", "PI", "NE CCor", "Grupo Despesa", "Natureza Despesa Detalhada", "Descrição Natureza", "Elemento Despesa", "Descrição Elemento", "Fonte Recursos Detalhada", "DESTAQUE RECEBIDO", "CREDITO DISPONIVEL", "DESPESAS EMPENHADAS (CONTROLE EMPENHO)", "DESPESAS PAGAs (CONTROLE EMPENHO)", "Total" ] 
        
        df.columns = novo_cabecalho 
        
        # Salvar o DataFrame de volta no arquivo Excel 
        df.to_excel(arquivo_destino, index=False) 
        
        print("cabeçalho renomeado com sucesso...") 
    except Exception as e: print(f"Ocorreu um erro ao copiar e apagar linhas: {e}")
        
def atualizar_arquivo_copia(origem, destino):
    try:
        shutil.copyfile(origem, destino)
        print("Arquivo copiado com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao copiar o arquivo: {e}")

def adicionar_coluna_tipo_resultado(arquivo_excel):
    try:
        # Carregar o arquivo Excel com o pandas
        df = pd.read_excel(arquivo_excel)

        # Definir as condições para preencher a coluna Tipo Resultado
        condicoes = [
            (df['Resultado EOF'] == 0),
            (df['Resultado EOF'] == 1),
            (df['Resultado EOF'] == 2),
            (df['Resultado EOF'] == 3),
            (df['Resultado EOF'] == 7),
            (df['Resultado EOF'] == 8),
            (df['Resultado EOF'] == 6)
        ]

        # Definir os valores correspondentes às condições
        valores = ['Financeiro', 'Primário Obrigatório', 'Primário Discricionário', 'PAC', 'Bancada Impositiva (RP7)', 'Emenda de Comissão (RP8)', 'Emenda Individual (RP6)']

        # Adicionar a coluna Tipo Resultado com base nas condições
        df['Tipo Resultado'] = np.select(condicoes, valores, default='Outro')

        # Salvar o DataFrame de volta no arquivo Excel
        df.to_excel(arquivo_excel, index=False)

        print("Coluna Tipo Resultado adicionada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao adicionar a coluna Tipo Resultado: {e}")        

def adicionar_coluna_tipo_resultado(arquivo_excel):
    try:
        # Carregar o arquivo Excel com o pandas
        df = pd.read_excel(arquivo_excel)

        # Converter a coluna 'Resultado EOF' para texto
        df['Resultado EOF'] = df['Resultado EOF'].astype(str)

        # Definir as condições para preencher a coluna Tipo Resultado quando Resultado EOF for igual a 3, 7, 8 ou 6
        condicoes1 = [
            (df['Resultado EOF'] == '0'),
            (df['Resultado EOF'] == '1'),
            (df['Resultado EOF'] == '2'),
            (df['Resultado EOF'] == '3'),
            (df['Resultado EOF'] == '7'),
            (df['Resultado EOF'] == '8'),
            (df['Resultado EOF'] == '6'),
            (df['Resultado EOF'] == '9')
        ]
        valores1 = ['Financeiro (RP0)','Primário Obrigatório (RP1)','Primário Discricionário (RP2)','PAC (RP3)', 'Bancada Impositiva (RP7)', 'Emenda de Comissão (RP8)', 'Emenda Individual (RP6)','Emenda de Relator (RP9)']

        # Adicionar a coluna Tipo Resultado com base nas condições1
        df['Tipo Resultado'] = np.select(condicoes1, valores1, default='')

        # Definir as condições para preencher a coluna Tipo Resultado quando Resultado EOF for igual a 2 e verificar a coluna Cod PO
        condicoes2 = [
            (df['Resultado EOF'] == '2') & (df['Cod PO'].str.startswith('0') | df['Cod PO'].isin(['RO06', 'CV20', 'CV21', 'RO0B'])),
            (df['Resultado EOF'] == '2') & (df['Cod PO'].str.startswith('EB')),
            (df['Resultado EOF'] == '2') & (df['Cod PO'].str.startswith('EC')),
            (df['Resultado EOF'] == '2') & (df['Cod PO'].str.startswith('EI')),
            (df['Resultado EOF'] == '2') & (df['Cod PO'].str.startswith('ER')),
            (df['Resultado EOF'] == '2') & (df['Cod PO'].str.startswith('CB')),
            (df['Resultado EOF'] == '2') & (df['Cod PO'] == 'RO00')
        ]
        valores2 = [
            'Primário Discricionário (RP2)', 'Emenda de Bancada (RP2)', 'Emenda de Comissão (RP2)',
            'Emenda Individual (RP2)', 'Emenda de Relator (RP2)', 'Comissão de Bancada (RP2)', 'Primário Obrigatório (RP2)'
        ]

        # Adicionar a coluna Tipo Resultado com base nas condições2
        df['Tipo Resultado'] = np.select(condicoes2, valores2, default=df['Tipo Resultado'])

        # Salvar o DataFrame de volta no arquivo Excel
        df.to_excel(arquivo_excel, index=False)

        print("Coluna Tipo Resultado adicionada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao adicionar a coluna Tipo Resultado: {e}")

def adicionar_coluna_tipo_resultado_resumido(arquivo_excel):
    try:
        # Carregar o arquivo Excel com o pandas
        df = pd.read_excel(arquivo_excel)

        # Mapear as categorias da coluna Tipo Resultado para o Tipo Resultado Resumido
        mapeamento = {
            'Primário Obrigatório (RP1)': 'Obrigatório',
            'Financeiro (RP0)': 'Financeiro',
            'Primário Discricionário (RP2)': 'Discricionárias',
            'Primário Obrigatório (RP2)': 'Discricionárias',
            'Emenda de Bancada (RP2)': 'Emenda RP2',
            'Emenda de Comissão (RP2)': 'Emenda RP2',
            'Emenda Individual (RP2)': 'Emenda RP2',
            'Emenda de Relator (RP2)': 'Emenda RP2',
            'Comissão de Bancada (RP2)': 'Emenda RP2',
            'PAC (RP3)': 'PAC (RP3)',
            'Bancada Impositiva (RP7)': 'Emendas',
            'Emenda de Comissão (RP8)': 'Emendas',
            'Emenda Individual (RP6)': 'Emendas',
            'Emenda de Relator (RP9)': 'Emendas'
        }

        # Adicionar a coluna Tipo Resultado Resumido com base no mapeamento
        df['Tipo Resultado Resumido'] = df['Tipo Resultado'].map(mapeamento)

        # Salvar o DataFrame de volta no arquivo Excel
        df.to_excel(arquivo_excel, index=False)

        print("Coluna Tipo Resultado Resumido adicionada com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao adicionar a coluna Tipo Resultado Resumido: {e}")         

def encontrar_coluna(sheet, nome_coluna):
    for cell in sheet[1]:
        if cell.value == nome_coluna:
            return cell.column_letter
    return None

def atualizar_data_de_atualizacao(arquivo_excel, nome_coluna):
    try:
        # Abrir o arquivo Excel
        wb = openpyxl.load_workbook(arquivo_excel)
        sheet = wb.active

        # Encontrar o número da coluna correspondente ao nome da coluna
        coluna_atualizacao = encontrar_coluna(sheet, nome_coluna)
        if coluna_atualizacao is None:
            print(f"Coluna '{nome_coluna}' não encontrada.")
            return

        # Preencher a data e hora atual na célula 2 da coluna 'Data de Atualização'
        data_atualizacao = datetime.now().strftime("%d/%m/%Y")
        sheet[coluna_atualizacao + '2'] = data_atualizacao

        # Salvar as alterações
        wb.save(arquivo_excel)
        wb.close()

        print(f"Data de Atualização atualizada com sucesso em {arquivo_excel}.")

    except Exception as e:
        print(f"Ocorreu um erro ao atualizar a data de atualização: {e}")

def renomear_cabecalho_painel_execucao(arquivo_origem, arquivo_destino):
    try:
        # Copiar o arquivo para o destino
        shutil.copyfile(arquivo_origem, arquivo_destino)
        
        # Carregar o arquivo copiado com o pandas
        df = pd.read_excel(arquivo_destino)

        # Renomear as colunas
        novo_cabecalho = [
            "Resultado EOF", "Descrição EOF", "Orgão UGE","Descrição UGE",
            "UG Executora", "Descrição UG", "Orgão UG", "Ação Governo", "Descrição Ação", "PTRES", "Fonte Recursos Detalhada", "NE - Mês Emissão ", "Função PO", "SubFunção PO", "Programa PO", "Cod PO", "Descrição PO", "Categoria Econômica",  "Dotação Atualizada", "DESTAQUE RECEBIDO", "CREDITO DISPONIVEL", 
            "DESPESAS EMPENHADAS (CONTROLE EMPENHO)", "DESPESAS PAGAs (CONTROLE EMPENHO)"
        ]
        df.columns = novo_cabecalho

        # Salvar o DataFrame de volta no arquivo Excel
        df.to_excel(arquivo_destino, index=False)

        print("cabeçalho renomeado com sucesso...")

    except Exception as e:
        print(f"Ocorreu um erro ao copiar e apagar linhas: {e}")        

gc.collect()
time.sleep(2)

def atualizar_credito_disponivel_por_linha_fixa(
    arquivo_2026,
    arquivo_destino,
    linha_base_fixa=85152
):
    try:
        # 🔹 Lê o arquivo de 2026 a partir da linha 14
        df_2026 = pd.read_excel(
            arquivo_2026,
            skiprows=13
        )

        # 🔥 Garante no máximo 30 colunas
        df_2026 = df_2026.iloc[:, :30]

        novo_cabecalho = [
            "Resultado EOF", "Descrição EOF", "NE CCor - Ano Emissão",
            "Órgão UGE", "Descrição UGE", "UG Executora", "Descrição UG",
            "UGE - UG Setorial Financeira", "Descrição Setorial",
            "Ação Governo", "Descrição Ação", "Função PO", "SubFunção PO",
            "Programa PO", "Cod PO", "Descrição PO", "PTRES", "PI",
            "NE CCor", "Grupo Despesa", "Natureza Despesa Detalhada",
            "Descrição Natureza", "Elemento Despesa",
            "Descrição Elemento", "Fonte Recursos Detalhada",
            "DESTAQUE RECEBIDO", "CREDITO DISPONIVEL",
            "DESPESAS EMPENHADAS (CONTROLE EMPENHO)",
            "DESPESAS PAGAs (CONTROLE EMPENHO)", "Total"
        ]

        df_2026.columns = novo_cabecalho

        # 🔹 Abre o arquivo destino com openpyxl
        wb = load_workbook(arquivo_destino)
        ws = wb.active

        # 🔥 Apaga tudo abaixo da linha fixa
        max_linha = ws.max_row
        if max_linha > linha_base_fixa:
            ws.delete_rows(linha_base_fixa + 1, max_linha - linha_base_fixa)

        # 🔹 Descobre a próxima linha vazia
        start_row = ws.max_row + 1

        # 🔹 Adiciona as linhas NOVAS (verticalmente!)
        for _, row in df_2026.iterrows():
            ws.append(row.tolist())

        wb.save(arquivo_destino)

        print("✔ Dados de 2026 atualizados com sucesso.")

    except Exception as e:
        print(f"❌ Erro ao atualizar crédito disponível: {e}")
                                     
def main():
    
    arquivo_origem_ted = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Ted Simec.xlsx'
    arquivo_destino_ted = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\COPIA Ted Simec.xlsx'
    arquivo_ted_final = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Teds Firmados.xlsx'
    arquivo_ted_ação = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Ted Ação.xlsx'
    arquivo_ted_ação_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\COPIA Ted Ação.xlsx'
    arquivo_ted_ação_final = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Ted Ação.xlsx'
    arquivo_ncs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Ncs desde 2013.xlsx'
    arquivo_ncs_simec_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\COPIA Ncs desde 2013.xlsx'
    arquivo_ncs_simec_final = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Ncs desde 2013.xlsx'
    arquivo_pfs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Pfs desde 2013.xlsx'
    arquivo_copia_pfs_simec = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\COPIA Pfs desde 2013.xlsx'
    arquivo_pfs_final = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Pfs desde 2013.xlsx'
    arquivo_credito_disponivel = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Crédito Disponivel Geral.xlsx'
    arquivo_credito_disponivel_2026 =r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Crédito Disponivel Geral 2026.xlsx'
    arquivo_credito_disponivel_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\COPIA Crédito Disponivel Geral.xlsx'
    arquivo_credito_disponivel_final = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Crédito Disponivel Geral.xlsx'
    arquivo_painel_execução = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\Painel Execução.xlsx'
    arquivo_painel_execução_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Relatório Orçamentário e Financeiro\COPIA Painel Execução.xlsx'
    arquivo_painel_execução_final = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Painel Execução.xlsx'
    arquivo_data_atualizacao = r'W:\B - TED\7 - AUTOMAÇÃO\Orçamentário e Financeiro desde 2013\Data de Atualização.xlsx'
    nome_coluna_atualizacao = 'Data de Atualização'
    shutil.copy(arquivo_ncs_simec, arquivo_ncs_simec_copia)
    shutil.copy(arquivo_pfs_simec, arquivo_copia_pfs_simec)
    shutil.copyfile(arquivo_ted_ação, arquivo_ted_ação_copia)
    shutil.copyfile(arquivo_painel_execução, arquivo_painel_execução_copia)
    
    print("Iniciando processamento aguarde...")
    
    atualizar_arquivo_copia(arquivo_credito_disponivel, arquivo_credito_disponivel_copia)
    criar_tabela_dinamica(arquivo_origem_ted, arquivo_destino_ted)
    limpar_planilhaTEDS_e_filtrar_siafi(arquivo_destino_ted)
    dividir_coluna_ug_concedente_e_Descentralizada(arquivo_destino_ted)
    criar_coluna_id_ug_gestao(arquivo_destino_ted)
    inserir_sinal_negativo(arquivo_ncs_simec_copia, arquivo_ncs_simec_final)
    inserir_sinal_negativo_pf(arquivo_copia_pfs_simec, arquivo_pfs_final)
    dividir_coluna_ug_concedente_e_Descentralizada(arquivo_ted_ação_copia)
    criar_coluna_id_ug_gestao(arquivo_ted_ação_copia)
    apagar_segunda_e_ultima_linha(arquivo_ted_ação_copia)
    apagar_segunda_e_ultima_linha(arquivo_ncs_simec_final)
    apagar_segunda_e_ultima_linha(arquivo_pfs_final)
    copiar_e_apagar_linhas(arquivo_credito_disponivel, arquivo_credito_disponivel_copia)
    gc.collect()
    time.sleep(2)
    atualizar_credito_disponivel_por_linha_fixa(arquivo_credito_disponivel_2026, arquivo_credito_disponivel_copia)
    gc.collect()
    time.sleep(2)
    adicionar_coluna_tipo_resultado(arquivo_credito_disponivel_copia)
    adicionar_coluna_tipo_resultado_resumido(arquivo_credito_disponivel_copia)
    renomear_cabecalho_painel_execucao(arquivo_painel_execução, arquivo_painel_execução_copia)
    adicionar_coluna_tipo_resultado(arquivo_painel_execução_copia)
    adicionar_coluna_tipo_resultado_resumido(arquivo_painel_execução_copia)
    atualizar_data_de_atualizacao(arquivo_data_atualizacao, nome_coluna_atualizacao)
    copiar_e_sobrescrever_arquivo(arquivo_destino_ted, arquivo_ted_final)
    copiar_e_sobrescrever_arquivo(arquivo_ted_ação_copia, arquivo_ted_ação_final)
    copiar_e_sobrescrever_arquivo(arquivo_credito_disponivel_copia, arquivo_credito_disponivel_final)
    copiar_e_sobrescrever_arquivo(arquivo_painel_execução_copia, arquivo_painel_execução_final)
    print("Processo finalizado com sucesso!")
    
if __name__ == "__main__":
    main()
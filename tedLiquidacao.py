import pandas as pd
import shutil
import locale
import openpyxl
import datetime
import re
import os
import xlwings as xw
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
import sys
from tqdm import tqdm
import time
from copy import copy

print('Iniciando Processamento aguarde...')

def formatar_gestao_emitente(gestao):
        if pd.notnull(gestao):
            gestao_str = str(int(gestao)) # Converte o valor para string
            zeros_esquerda = '0' * (5 - len(gestao_str))
            return zeros_esquerda + gestao_str
        return ''

def remove_pontos_zeros(valor):
    if pd.notnull(valor):
        return str(int(float(valor))).rstrip('.')
    return ''

def formatar_contabil(value):
    if pd.notnull(value):
        if isinstance(value, (int, float)):
            return "{:,.2f}".format(float(value)).replace(",", "_").replace(".", ",").replace("_", ".")
        else:
            return value  # Mantém o cabeçalho ou outros valores não numéricos
    return None

def formatar_data(data):
    if pd.notnull(data):
        # Verifique se a data não é '-' antes de tentar formatá-la
        if data != '-':
            # Converte a data para o formato desejado (DD/MM/AAAA)
            data_formatada = datetime.strptime(str(data), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
            return data_formatada
    return None

def substituir_valores_grupo_despesa(value):
    if pd.notnull(value):
        # Converta o valor para inteiro antes de comparar
        try:
            value = int(value)
            if value == 3:
                return 'C'
            elif value == 4:
                return 'D'
            elif value == 5:
                return 'E'
        except ValueError:
            pass
    return value

def criar_aba_repassar(wb):
    nome_aba = 'REPASSAR'
    if nome_aba in wb.sheetnames:
        # Excluir a aba 'REPASSAR' se ela já existir
        wb.remove(wb[nome_aba])
    return wb.create_sheet(nome_aba)

# Função para formatar um valor monetário para o formato consistente
def formatar_valor_monetario(valor):
    if isinstance(valor, (float, int)):
        return valor  # Não é necessário formatação
    # Remova os pontos de milhares, substitua a vírgula por ponto e converta para float
    valor = valor.replace(".", "").replace(",", ".")
    return float(valor)

def processar_exeSimec():
    # Caminho do arquivo original
    arquivo_original = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\Execução SIMEC DOC NE.xlsx'

    # Caminho do arquivo de cópia
    arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA Execução SIMEC DOC NE.xlsx'

    # Faça uma cópia do arquivo original
    shutil.copy(arquivo_original, arquivo_copia)

    # Carregue o arquivo de cópia em um DataFrame
    df = pd.read_excel(arquivo_copia)

    # Remova a linha com todos os zeros (linha 1, índice 0)
    df = df.drop(0)
    
    # Aplica a função à coluna "Gestão Emitente - NE"
    df['Gestão Emitente - NE'] = df['Gestão Emitente - NE'].apply(formatar_gestao_emitente)
    
    # Converta todas as colunas relevantes para string
    df['Gestão Emitente - NE'] = df['Gestão Emitente - NE'].astype(str)
    
    # Use a função personalizada para remover pontos e zeros à direita na coluna "UG Executora Emitente - NE"
    df['UG Executora Emitente - NE'] = df['UG Executora Emitente - NE'].apply(remove_pontos_zeros)

    # Preencha a coluna "Dados do Empenho" com a concatenação das colunas desejadas e remova os pontos
    df['Dados do Empenho'] = (df['UG Executora Emitente - NE'] +
                              df['Gestão Emitente - NE'] +
                              df['Número do Empenho']).str.replace('.', '')

    # Reformatar todas as colunas relevantes para o formato desejado
    # Converta a coluna "Gestão Emitente - NE" para string e preencha com zeros à esquerda para ter sempre 5 dígitos
    df['Gestão Emitente - NE'] = df['Gestão Emitente - NE'].astype(str).str.replace('.', '').str.zfill(5)
    df['Número do Empenho'] = df['Número do Empenho'].str.zfill(12)
    
    # Mova a coluna "Dados do Empenho" para a posição desejada (índice 5)
    coluna_dados_empenho = df.pop('Dados do Empenho')
    df.insert(5, 'Dados do Empenho', coluna_dados_empenho)

    # Converte as três últimas colunas em números
    for coluna in df.columns[-1:]:
        if df[coluna].dtype != 'float64':  # Verifique se a coluna não é do tipo float64
            df[coluna] = pd.to_numeric(df[coluna], errors='coerce')  # Converte para números, tratando erros como NaN
    print("Converteu as três últimas colunas em números.")
    
    # Formate as três últimas colunas em formato contábil usando a função formatar_contabil
    for coluna in df.columns[-1:]:
        if df[coluna].dtype == 'float64':  # Verifique se a coluna contém números reais
            df[coluna] = df[coluna].apply(formatar_contabil)
    print("Formatou as três últimas colunas em formato contábil.")
    
    # Formate a coluna "Fim da Vigência" para o formato DD/MM/AAAA
    df['Fim da Vigência'] = df['Fim da Vigência'].apply(formatar_data)
    
    # Salve o DataFrame de volta no arquivo de cópia
    df.to_excel(arquivo_copia, index=False)

processar_exeSimec()


def processar_arquivo():
    # Defina a localização para o formato contábil desejado (centavos separados por vírgula)
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

    # Caminho do arquivo original
    arquivo_original = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

    # Caminho do arquivo de cópia
    arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

    # Faça uma cópia do arquivo original
    shutil.copy(arquivo_original, arquivo_copia)

    # Carregue o arquivo de cópia em um DataFrame
    df = pd.read_excel(arquivo_copia, header=None)  # Não considera a primeira linha como cabeçalho
    print("Carregou o arquivo de cópia em um DataFrame.")
    
    
    # Exclua as linhas de 1 a 12
    df = df.iloc[12:]  # Mantém as linhas após a 13ª linha
    print("Excluiu as linhas de 1 a 12.")
    
    # Converte as três últimas colunas em números
    for coluna in df.columns[-4:]:
        if df[coluna].dtype != 'float64':  # Verifique se a coluna não é do tipo float64
            df[coluna] = pd.to_numeric(df[coluna], errors='coerce')  # Converte para números, tratando erros como NaN
    print("Converteu as três últimas colunas em números...")
    
    # Formate as três últimas colunas em formato contábil usando a função formatar_contabil
    for coluna in df.columns[-4:]:
        if df[coluna].dtype == 'float64':  # Verifique se a coluna contém números reais
            df[coluna] = df[coluna].apply(formatar_contabil)
    print("Formatou as três últimas colunas em formato contábil...")
    
    # Substitua os valores na coluna "Grupo Despesa" (coluna 13 no índice base 0)
    df.iloc[:, 13] = df.iloc[:, 13].apply(substituir_valores_grupo_despesa)
    print("Substituiu os valores na coluna 'Grupo Despesa...")
    
    # Adicione uma nova coluna chamada "SITUAÇÃO" após a última coluna existente
    df['SITUAÇÃO'] = ''
    
    def determinar_situacao(row):
        ano_atual = datetime.now().year
        valor_coluna = row.iloc[2]
        if pd.notnull(valor_coluna):
            try:
                valor_inteiro = int(valor_coluna)
                if valor_inteiro == ano_atual:
                    return 'TRF003'
            except ValueError:
                pass
        return 'TRF004'
    
    # Preencha a coluna "SITUAÇÃO" com base na lógica desejada
    df['SITUAÇÃO'] = df.apply(determinar_situacao, axis=1)
    print("Preencheu a coluna 'SITUAÇÃO...")
    
    # Renomeie o cabeçalho com os nomes do array
    nomes_colunas = [
        'Resultado EOF', 'DESCRIÇÃO EOF', 'NE CCor - Ano Emissão', 'Órgão UGE',
        'DESCRIÇÃO UGE', 'UG Executora', 'DESCRIÇÃO EXECUTORA', 'UGE - UG Setorial Financeira',
        'DESCRIÇÃO FINANCEIRA', 'Ação Governo', 'PTRES', 'PI', 'NE CCor', 'Grupo Despesa',
        'Natureza Despesa Detalhada', 'NATUREZA', 'Elemento Despesa', 'ND', 'Fonte Recursos Detalhada',
        'DESPESAS LIQUIDADAS A PAGAR(CONTROLE EMPENHO)', 'RESTOS A PAGAR PROCESSADOS A PAGAR',
        'RESTOS A PAGAR NAO PROCES. LIQUIDADOS A PAGAR', 'Total', 'SITUAÇÃO'
    ]
    df.iloc[0] = nomes_colunas  # Substitui a primeira linha pelo cabeçalho renomeado
    print("Renomeou o cabeçalho com os nomes das colunas...")
    
    # Itere sobre a coluna "Total" e remova as linhas onde o valor é zero
    df = df[df[22] != 0]
    print("Removeu as linhas em que o valor na coluna 'Total' é zero...")
    
    # Reescreva o arquivo de cópia com as linhas excluídas, o cabeçalho renomeado
    df.to_excel(arquivo_copia, index=False, header=False)  # Não escreve cabeçalho
    print("Reescreveu o arquivo de cópia com as alterações...")
    
# Chame a função para processar o arquivo de cópia
processar_arquivo()

print("Executando...")

def excluir_linhas_total_zero(arquivo_copia):
    # Carregue o arquivo de cópia em um DataFrame
    df = pd.read_excel(arquivo_copia, header=None)  # Não considera a primeira linha como cabeçalho
    
    # Converte a coluna "Total" para texto (string)
    df[22] = df[22].astype(str)
    
    # Crie um filtro para as linhas onde o valor na coluna "Total" seja igual a '0,00'
    filtro_total_zero = (df[22] == '0,00')
    
    # Aplique o filtro para excluir as linhas correspondentes
    df = df[~filtro_total_zero]
    print('Exclusão das linhas com o total igual a 0,00...')
    
    # Reescreva o arquivo de cópia com as linhas excluídas
    df.to_excel(arquivo_copia, index=False, header=False)  # Não escreve cabeçalho

# Chame a função para excluir as linhas com "Resultado EOF" igual a "6", "7" ou "8", e "Total" igual a 0
arquivo_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

excluir_linhas_total_zero(arquivo_copia)

print("Executando...")

def copyData(arquivo_copia):

    # Abrir o arquivo "W:\B - TED\5 - Liquidação, lotes financeiro e macros\Liquidação\Automação Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2023.xlsx"
    arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'
    wb = openpyxl.load_workbook(arquivo_principal)

    # Selecionar a aba "Sheet1" (ou outra aba se for diferente)
    planilha_principal = wb["Sheet1"]

    # Adicionar as colunas "TED", "SIAFI", "Vigência" e "Estado Atual" nas colunas Y, Z, AA e AB
    colunas = ["TED", "SIAFI", "Vigência", "Estado Atual", "Vinculação Pagamento"]
    coluna_indices = ["Y", "Z", "AA", "AB", "AC"]
    for coluna, titulo in zip(coluna_indices, colunas):
        planilha_principal[f"{coluna}1"] = titulo

    print('Adicionando as colunas TED, SIAFI, Vigência, Estado atual e Vinculação pagamento...')
    print("Executando...")
    
    # Criar uma nova aba chamada "DOC NE"
    planilha_doc_ne = wb.create_sheet("DOC NE")
    print('Aba DOC NE criada...')
    print("Copiando os dados para a planilha TED Liquidação Geral...")
    
    # Abrir o arquivo COPIA Execução SIMEC DOC NE
    arquivo_doc_ne = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA Execução SIMEC DOC NE.xlsx'
    wb_doc_ne = openpyxl.load_workbook(arquivo_doc_ne)

    # Selecionar a aba com os dados (você pode precisar ajustar o nome da aba)
    planilha_dados_doc_ne = wb_doc_ne.active

    # Copiar os dados de "DOC NE" para a nova aba "DOC NE" no arquivo principal
    for linha in planilha_dados_doc_ne.iter_rows():
        nova_linha = [celula.value for celula in linha]
        planilha_doc_ne.append(nova_linha)

    # Salvar as alterações no arquivo principal
    wb.save(arquivo_principal)
copyData(arquivo_copia)

print('Copia sendo Executada...')

def preencher_colunas_siafi_vigencia_estado_atual(arquivo_principal, data_atual):
    wb_principal = openpyxl.load_workbook(arquivo_principal)

    # Acessar as abas "Sheet1" e "DOC NE"
    planilha_principal = wb_principal["Sheet1"]
    planilha_doc_ne = wb_principal["DOC NE"]
    
    print('Mapeando os dados das Colunas NE CCor e DOC NE...')

    # Criar dicionários para mapear os valores da coluna "NE CCor" na aba "DOC NE"
    ted_dict = {}
    siafi_dict = {}
    vigencia_dict = {}
    estado_atual_dict = {}
    for linha in planilha_doc_ne.iter_rows(min_row=2):  # Ignorar cabeçalho
        ne_ccor = linha[5].value   # Dados do Empenho
        estado_atual = linha[6].value  # Estado Atual
        vigencia = linha[7].value  # Fim da Vigência
        siafi = linha[8].value     # SIAFI
        ted = linha[9].value       # TED

        if ne_ccor:
            ted_dict[ne_ccor] = ted
            siafi_dict[ne_ccor] = siafi
            vigencia_dict[ne_ccor] = vigencia
            estado_atual_dict[ne_ccor] = estado_atual

    # Definir formatações de cor
    red_fill = PatternFill(start_color="FFFA5E46", end_color="FFFA5E46", fill_type="solid")
    green_fill = PatternFill(start_color="FF78E6A3", end_color="FF78E6A3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")

    print('Formatação das cores sendo executada...')

    # Preencher na aba "Sheet1"
    for linha in planilha_principal.iter_rows(min_row=2):
        ne_ccor_sheet1 = linha[12].value  # NE CCor na aba "Sheet1"

        if ne_ccor_sheet1:
            # TED
            if ne_ccor_sheet1 in ted_dict:
                linha[24].value = ted_dict[ne_ccor_sheet1]

            # SIAFI
            if ne_ccor_sheet1 in siafi_dict:
                linha[25].value = siafi_dict[ne_ccor_sheet1]

            # Vigência
            if ne_ccor_sheet1 in vigencia_dict:
                valor_vigencia = vigencia_dict[ne_ccor_sheet1]
                if isinstance(valor_vigencia, str):
                    try:
                        vigencia = datetime.strptime(valor_vigencia, "%d/%m/%Y").date()
                        linha[26].value = vigencia
                        if vigencia < data_atual:
                            linha[26].fill = red_fill
                        elif vigencia > data_atual:
                            linha[26].fill = green_fill
                    except ValueError:
                        linha[26].value = valor_vigencia  # mantém texto ("Termo em Execução", etc.)
                        linha[26].fill = yellow_fill
                else:
                    linha[26].value = valor_vigencia

            # Estado Atual
            if ne_ccor_sheet1 in estado_atual_dict:
                linha[27].value = estado_atual_dict[ne_ccor_sheet1]

    # Nova funcionalidade
    preencher_coluna_vinculacao_pagamento(planilha_principal)

    # Salvar
    wb_principal.save(arquivo_principal)
    print('Alterações salvas no arquivo principal...')


def preencher_coluna_vinculacao_pagamento(planilha_principal):
    print('Preenchendo a coluna Vinculação Pagamento...')
    for linha in planilha_principal.iter_rows(min_row=2):  # Ignore o cabeçalho
        resultado_eof = linha[0].value  # Valor da coluna "Resultado EOF" (índice 0)

        if resultado_eof == "2":
            linha[28].value = "400"  # Preencha com "400"
        elif resultado_eof == "3":
            linha[28].value = "415"  # Preencha com "415"
        elif resultado_eof == "9":
            linha[28].value = "409"  # Preencha com "409"
        elif resultado_eof == "6":
            linha[28].value = "405"
        elif resultado_eof == "7":
            linha[28].value = "410"
        elif resultado_eof == "8":
            linha[28].value = "408"             

# Obtenha a data atual
data_atual = datetime.now().date()

# Chame a função com o caminho do arquivo principal e a data atual
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'
preencher_colunas_siafi_vigencia_estado_atual(arquivo_principal, data_atual)

# Função para criar uma tabela dinâmica em uma nova aba do arquivo existente
def criar_tabela(arquivo_principal):
    # Carregue o arquivo principal
    wb = load_workbook(arquivo_principal)

    # Nome da aba da tabela dinâmica
    nome_aba_tabela = "Tabela"

    # Se a aba da tabela dinâmica já existe, exclua-a
    if nome_aba_tabela in wb.sheetnames:
        wb.remove(wb[nome_aba_tabela])

    # Crie uma nova aba chamada "Tabela Dinâmica"
    planilha_tabela_dinamica = wb.create_sheet(nome_aba_tabela)

    # Selecione as colunas desejadas
    colunas_selecionadas = [
        "Resultado EOF", "Vinculação Pagamento", "UGE - UG Setorial Financeira", "UG Executora", "TED", "Ação Governo",
        "Vigência", "SITUAÇÃO", "SIAFI", "Fonte Recursos Detalhada", "Grupo Despesa", "Total"
    ]

    # Acesse a aba "Sheet1"
    planilha_principal = wb["Sheet1"]

    # Extraia os dados da aba "Sheet1" para um DataFrame
    df = pd.DataFrame(planilha_principal.values)
    df.columns = df.iloc[0]
    df = df[1:]

    # Selecione apenas as colunas desejadas
    df = df[colunas_selecionadas]

    # Obtém a data atual
    data_atual = datetime.now().date()
    
    # Converta a coluna "Vigência" para o formato de data
    df['Vigência'] = pd.to_datetime(df['Vigência'], format='%d/%m/%Y').dt.date
    
    # Aplique filtros
    filtro_vigente = df['Vigência'] > data_atual
    filtro_nao_nulos = df['Vigência'].notnull()
    df_filtrado = df[filtro_vigente & filtro_nao_nulos]

    # Preencha a nova aba com os dados do DataFrame filtrado
    for row in dataframe_to_rows(df_filtrado, index=False, header=True):
        planilha_tabela_dinamica.append(row)
    
    # Salve o arquivo com a nova aba e a tabela dinâmica
    wb.save(arquivo_principal)

    # Imprima uma mensagem de conclusão
    print(f"Tabela criada com sucesso na aba '{nome_aba_tabela}' do arquivo {arquivo_principal}...")
    print('Executando...')

# Substitua o caminho do arquivo principal pelo seu
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para criar a tabela dinâmica
criar_tabela(arquivo_principal)

print('Convertendo a coluna Total para valor contábil... ')

def converter_coluna_total_para_numero(arquivo_principal):
    # Abra o arquivo Excel
    wb = openpyxl.load_workbook(arquivo_principal)

    # Selecione a aba de trabalho
    sheet = wb['Tabela']  # Substitua pelo nome correto da aba se necessário

    # Selecione a coluna "Total" pelo cabeçalho
    coluna_total = sheet['L']  # Substitua pelo cabeçalho correto da coluna

    # Pule a primeira célula (cabeçalho) e converta os valores na coluna para números
    for cell in coluna_total[1:]:
        if cell.value:
            cell.value = float(cell.value.replace(".", "").replace(",", "."))

    # Salve o arquivo
    wb.save(arquivo_principal)

# Substitua o caminho do arquivo principal pelo seu
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para converter a coluna "Total" em números
converter_coluna_total_para_numero(arquivo_principal)

print('Agrupando as Linhas que são igual os valores nas colunas Resultado EOF, Vinculação Pagamento, UGE - UG Setorial Financeira, UG Executora, TED, Ação Governo, Vigência, SITUAÇÃO, SIAFI, Fonte Recursos Detalhada e Grupo Despesa e somando seus Valores...')

def agrupar_e_somar_total(arquivo_principal):
    # Carregue o arquivo principal
    wb = load_workbook(arquivo_principal)

    # Acesse a aba da tabela dinâmica
    planilha_tabela_dinamica = wb["Tabela"]

    # Crie um DataFrame diretamente a partir dos valores da planilha
    data = planilha_tabela_dinamica.values
    colunas = next(data)  # A primeira linha contém os nomes das colunas
    df = pd.DataFrame(data, columns=colunas)

    # Defina as colunas usadas para agrupar
    colunas_para_agrupar = ["Resultado EOF","Vinculação Pagamento", "UGE - UG Setorial Financeira", "UG Executora", "TED", "Ação Governo", "Vigência", "SITUAÇÃO", "SIAFI", "Fonte Recursos Detalhada", "Grupo Despesa"]
    
    # Agrupe os dados e some a coluna 'Total'
    df_agrupado = df.groupby(colunas_para_agrupar, as_index=False)['Total'].sum()
    
    # Limpe a planilha removendo todas as linhas, exceto o cabeçalho
    planilha_tabela_dinamica.delete_rows(1, planilha_tabela_dinamica.max_row)

    # Adicione as linhas agrupadas ao DataFrame da planilha
    for row in dataframe_to_rows(df_agrupado, index=False, header=True):
        planilha_tabela_dinamica.append(row)

    # Preencha a aba da tabela dinâmica com os dados agrupados
    for index, row in df_agrupado.iterrows():
        for col_idx, col_name in enumerate(df_agrupado.columns):
            if col_name == 'Total':
                planilha_tabela_dinamica.cell(row=index + 2, column=col_idx + 1, value=formatar_contabil(row[col_name]))
            else:
                planilha_tabela_dinamica.cell(row=index + 2, column=col_idx + 1, value=row[col_name])

    # Salve o arquivo com as alterações
    wb.save(arquivo_principal)

    # Imprima uma mensagem de conclusão
    print("Agrupamento, soma e formatação contábil dos Totais concluídos...")

# Substitua o caminho do arquivo principal pelo seu
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para agrupar, somar e formatar os Totais
agrupar_e_somar_total(arquivo_principal)

print('Processamento da Execução SIMEC e TED Liquidação Geral Executadas...')

def processar_limite_saque(caminho_arquivo_original):
    # Obter o nome do arquivo original
    nome_arquivo_original = os.path.basename(caminho_arquivo_original)

    # Adicionar "COPIA" antes do nome do arquivo
    nome_arquivo_copia = "COPIA " + nome_arquivo_original

    # Caminho do arquivo de cópia
    caminho_arquivo_copia = os.path.join(os.path.dirname(caminho_arquivo_original), nome_arquivo_copia)

    # Copiar o arquivo Excel para o arquivo de cópia
    shutil.copy2(caminho_arquivo_original, caminho_arquivo_copia)

    # Carregar o arquivo Excel de cópia
    df = pd.read_excel(caminho_arquivo_copia)

    # Excluir a última linha
    df = df.drop(df.index[-1])

    # Renomear as colunas com os nomes desejados
    df.columns = [
        'Órgão UGE', 
        'Descrição UGE', 
        'UG Executora', 
        'Descrição UG', 
        'Órgão UGE - Gestão', 
        'Vinculação Pagamento', 
        'Fonte Recursos Detalhada', 
        'LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)'
    ]
    
    # Use a função personalizada para remover pontos e zeros à direita e formatar como string
    df['UG Executora'] = df['UG Executora'].apply(remove_pontos_zeros).apply(formatar_gestao_emitente)
    
    # Converter as colunas para string, substituindo NaN por string vazia
    df['Vinculação Pagamento'] = df['Vinculação Pagamento'].astype(str).apply(remove_pontos_zeros).fillna("")
    df['UG Executora'] = df['UG Executora'].astype(str).fillna("")
    df['Fonte Recursos Detalhada'] = df['Fonte Recursos Detalhada'].astype(str).fillna("")
    
    # Adicionar a nova coluna "UG+FONTE" com a concatenação
    df['Vinculação+UG+FONTE'] = (
        df['Vinculação Pagamento'] + df['UG Executora'] + df['Fonte Recursos Detalhada']
    )
    
    df['Vinculação+UG+FONTE'] = df['Vinculação+UG+FONTE'].astype(str).fillna("")
    
    # Reordenar as colunas (trocar a posição de "LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)" e "UG+FONTE")
    df = df[['Órgão UGE', 
             'Descrição UGE', 
             'UG Executora', 
             'Descrição UG', 
             'Órgão UGE - Gestão', 
             'Vinculação Pagamento', 
             'Fonte Recursos Detalhada', 
             'Vinculação+UG+FONTE', 
             'LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)']]

    # Exibir as primeiras linhas do DataFrame
    print(df.head())

    # Salvar as alterações de volta no arquivo de cópia
    df.to_excel(caminho_arquivo_copia, index=False)

# Exemplo de uso da função
caminho_arquivo_original = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\Limite de Saque MEC.xlsx'
processar_limite_saque(caminho_arquivo_original)


def copiar_dados_para_limite_saque():
   # Abrir o arquivo "W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2023.xlsx"
    arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'
    wb = openpyxl.load_workbook(arquivo_principal)

    # Criar uma nova aba chamada "Limite Saque"
    planilha_limite_saque = wb.create_sheet("Limite Saque")
    
    # Abrir o arquivo COPIA Execução SIMEC DOC NE
    arquivo_limite_saque = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA Limite de Saque MEC.xlsx'
    wb_limite_saque = openpyxl.load_workbook(arquivo_limite_saque)
    
    # Selecionar a aba com os dados (você pode precisar ajustar o nome da aba)
    planilha_dados_limite_saque = wb_limite_saque.active
    
    # Copiar os dados de "DOC NE" para a nova aba "DOC NE" no arquivo principal
    for linha in planilha_dados_limite_saque.iter_rows():
        nova_linha = [celula.value for celula in linha]
        planilha_limite_saque.append(nova_linha)

    # Salvar as alterações no arquivo principal
    wb.save(arquivo_principal)

# Exemplo de uso da função para copiar os dados
copiar_dados_para_limite_saque()

print('Processando Contas Cadastro e Controle...')

def copiarArquivoSemMesclagem(caminho_arquivo_original):
    # Obter o nome do arquivo original
    nome_arquivo_original = os.path.basename(caminho_arquivo_original)

    # Adicionar "COPIA" antes do nome do arquivo
    nome_arquivo_copia = "COPIA " + nome_arquivo_original

    # Caminho do arquivo de cópia
    caminho_arquivo_copia = os.path.join(os.path.dirname(caminho_arquivo_original), nome_arquivo_copia)

    # Abra o arquivo original usando openpyxl
    wb_original = load_workbook(caminho_arquivo_original)

    # Remova a mesclagem de células em todas as planilhas
    for sheet in wb_original.worksheets:
        merged_cells_copy = copy(sheet.merged_cells.ranges)  # Crie uma cópia das células mescladas
        for merged_cell_range in merged_cells_copy:
            sheet.unmerge_cells(merged_cell_range.coord)

    # Salve as alterações no arquivo original
    wb_original.save(caminho_arquivo_original)

    # Copiar o arquivo Excel para o arquivo de cópia
    shutil.copy2(caminho_arquivo_original, caminho_arquivo_copia)

    # Abra o arquivo de cópia usando openpyxl
    wb_copia = load_workbook(caminho_arquivo_copia)

    # Remova a mesclagem de células na cópia também
    for sheet in wb_copia.worksheets:
        merged_cells_copy = copy(sheet.merged_cells.ranges)  # Crie uma cópia das células mescladas
        for merged_cell_range in merged_cells_copy:
            sheet.unmerge_cells(merged_cell_range.coord)

    # Aplicar a operação de remover os dois primeiros caracteres da coluna "Conta Corrente" na cópia (coluna C)
    for sheet in wb_copia.worksheets:
        for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):  # Apenas coluna C
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value)[2:]

    # Salvar as alterações na cópia
    wb_copia.save(caminho_arquivo_copia)

# Exemplo de uso da função
caminho_arquivo_original = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\TED - Contas Cadastro e Controle.xlsx'
copiarArquivoSemMesclagem(caminho_arquivo_original)

print("Arquivo copiado sem mesclagem de células e com a operação aplicada.")

print("Copiando o Repasse para a Liquidação Geral...")

def copiarRepasse():
    # Caminho dos arquivos
    caminho_arquivo_origem = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED - Contas Cadastro e Controle.xlsx'
    caminho_arquivo_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

    # Carregar o arquivo de origem e destino
    arquivo_origem = openpyxl.load_workbook(caminho_arquivo_origem)
    arquivo_destino = openpyxl.load_workbook(caminho_arquivo_destino)

    # Nome da aba de destino
    nome_aba_destino = "Contas Cadastro e Controle"

    # Verificar se a aba "Contas Cadastro e Controle" já existe no arquivo de destino
    if nome_aba_destino in arquivo_destino.sheetnames:
        # Se a aba já existe, a remova para sobrescrever
        aba_existente = arquivo_destino[nome_aba_destino]
        arquivo_destino.remove(aba_existente)

    # Selecionar a aba de origem e criar uma nova aba de destino
    aba_origem = arquivo_origem.active  # ou selecione a aba desejada
    aba_destino = arquivo_destino.create_sheet(nome_aba_destino)

    # Copiar os dados da aba de origem para a nova aba de destino
    for row in aba_origem.iter_rows(min_row=1, values_only=True):
        nova_linha = []
        for valor in row:
            if valor is None:  # Verificar se a célula está vazia
                nova_linha.append(0)  # Preencher com 0
            else:
                nova_linha.append(valor)
        aba_destino.append(nova_linha)

    # Aplicar a formatação à linha 2 da aba de destino
    red_fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")  # Vermelho (ARGB hex)
    font = Font(bold=True, color="000000")  # Fonte em negrito e cor preta

    for row in aba_destino.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.fill = red_fill
            cell.font = font

    # Salvar as alterações no arquivo de destino
    arquivo_destino.save(caminho_arquivo_destino)

# Chame a função para copiar e formatar os dados
copiarRepasse()

print('Planilhas auxiliares sendo adicionadas...')

def processoFinal():
    # Caminho do arquivo principal
    arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

    # Carregue o arquivo principal
    wb = openpyxl.load_workbook(arquivo_principal)

    # Acesse a aba 'Tabela'
    planilha_tabela_dinamica = wb["Tabela"]
    
    # Acesse a aba 'REPASSAR' usando a função modificada
    planilha_repassar = criar_aba_repassar(wb)

    # Copie os dados da aba 'Tabela' para a aba 'REPASSAR'
    for row in planilha_tabela_dinamica.iter_rows(values_only=True):
        planilha_repassar.append(row)

    # Remover colunas existentes da 'L' à 'T' se já existirem
    for coluna in list(planilha_repassar.iter_cols(min_col=13, max_col=20)):
        for cell in coluna:
            cell.value = None

    # Inserir cabeçalhos das novas colunas após a coluna 'Total'
    cabecalhos = ["Vinculação+UG+FONTE", "LIMITE DE SAQUE", "SOMASE UG+FONTE", "CONDICIONAL LIMITE DE SAQUE", "SOMASE SIAFI", "CONTAS REPASSAR", "CONDICIONAL CONTAS A REPASSAR", "SIMEC (NC-PF)", "CONDICIONAL SIMEC"]
    for col_idx, coluna in enumerate(cabecalhos, start=13):  # Começando na coluna 'L'
        planilha_repassar.cell(row=1, column=col_idx, value=coluna)

    # Salve o arquivo com as alterações
    wb.save(arquivo_principal)

    # Imprima uma mensagem de conclusão
    print("Dados copiados para a aba 'REPASSAR' com sucesso...")

# Chame a função processoFinal
processoFinal()

def preencher_coluna_ug_fonte():
    # Caminho do arquivo principal
    arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

    # Carregar o arquivo principal
    wb = openpyxl.load_workbook(arquivo_principal)

    # Acesse a aba 'REPASSAR'
    planilha_repassar = wb['REPASSAR']

    # Carregar a aba 'REPASSAR' em um DataFrame
    data = planilha_repassar.values
    colunas = next(data)
    df = pd.DataFrame(data, columns=colunas)

    # Garantir que os valores da chave sejam strings
    df['UG Executora'] = df['UG Executora'].astype(str).str.strip()
    df['Fonte Recursos Detalhada'] = df['Fonte Recursos Detalhada'].astype(str).str.strip()
    df['Vinculação Pagamento'] = df['Vinculação Pagamento'].astype(str).str.strip()

    # Criar chave 'Vinculação+UG+FONTE'
    df['Vinculação+UG+FONTE'] = df['Vinculação Pagamento'] + df['UG Executora'] + df['Fonte Recursos Detalhada']

    # Preencher a coluna 'UG+FONTE' na aba 'REPASSAR' (coluna 13 ou 'L')
    for r_idx, ug_fonte in enumerate(df['Vinculação+UG+FONTE'], start=2):  
        planilha_repassar.cell(row=r_idx, column=13, value=ug_fonte)
    
    # ---- CORREÇÃO DA LEITURA DA ABA 'Limite Saque' ----
    
    # Carregar a aba 'Limite Saque'
    planilha_limite = wb['Limite Saque']
    data_limite = planilha_limite.values
    colunas_limite = next(data_limite)
    df_limite = pd.DataFrame(data_limite, columns=colunas_limite)

    # Garantir que os valores da chave sejam strings e remover espaços
    df_limite['UG Executora'] = df_limite['UG Executora'].astype(str).str.strip()
    df_limite['Fonte Recursos Detalhada'] = df_limite['Fonte Recursos Detalhada'].astype(str).str.strip()
    df_limite['Vinculação Pagamento'] = df_limite['Vinculação Pagamento'].astype(str).str.strip()

    # Criar chave na aba 'Limite Saque'
    df_limite['Chave_Limite'] = df_limite['Vinculação Pagamento'] + df_limite['UG Executora'] + df_limite['Fonte Recursos Detalhada']

    # Criar dicionário para busca otimizada
    dicionario_limite = df_limite.set_index('Chave_Limite')['LIMITES DE SAQUE (OFSS, DIVIDA, BACEN E PREV)'].to_dict()

    # ---- CORREÇÃO NO PREENCHIMENTO DA COLUNA ----
    
    # Preencher a coluna 'LIMITE DE SAQUE' na aba 'REPASSAR' (coluna 14 ou 'M')
    for r_idx in range(2, planilha_repassar.max_row + 1):
        ug_fonte = planilha_repassar.cell(row=r_idx, column=13).value  # Chave gerada

        # Verifica se a chave existe no dicionário
        limite_saque = dicionario_limite.get(ug_fonte, 0)

        # Preenche o valor correto
        planilha_repassar.cell(row=r_idx, column=14, value=limite_saque)

    # Salvar o arquivo com as alterações
    wb.save(arquivo_principal)

    # Imprimir uma mensagem de conclusão
    print("Coluna 'LIMITE DE SAQUE' corrigida e preenchida corretamente!")

# Chamar a função
preencher_coluna_ug_fonte()

print('preenchendo a coluna SOMASE UG+FONTE...')

def calcular_e_salvar_somase_ug_fonte(arquivo_path):
    # Carregue o arquivo principal
    wb = openpyxl.load_workbook(arquivo_path)

    # Acesse a aba 'REPASSAR' no arquivo original
    planilha_repassar = wb['REPASSAR']

    # Carregar a aba 'REPASSAR' em um DataFrame
    data = planilha_repassar.values
    colunas = next(data)
    df = pd.DataFrame(data, columns=colunas)

    # Substituir vírgulas por pontos em 'Total' na aba 'REPASSAR'
    for r_idx in range(2, planilha_repassar.max_row + 1):
        total_str = planilha_repassar.cell(row=r_idx, column=12).value  # 'Total'
        # Remover os pontos dos números
        total_str = total_str.replace('.', '')
        # Substituir a vírgula por ponto
        total_str = total_str.replace(',', '.')
        # Converter para float
        total = float(total_str)
        planilha_repassar.cell(row=r_idx, column=12, value=total)  # Atualizar a célula com o valor float

    # Calcular a coluna 'SOMASE UG+FONTE' manualmente
    ug_fonte_totals = {}
    for r_idx in range(2, planilha_repassar.max_row + 1):
        ug_fonte = planilha_repassar.cell(row=r_idx, column=13).value  # 'UG+FONTE'
        total = float(planilha_repassar.cell(row=r_idx, column=12).value)  # 'Total'
        if ug_fonte not in ug_fonte_totals:
            ug_fonte_totals[ug_fonte] = 0.0
        ug_fonte_totals[ug_fonte] += total

    # Atualizar a coluna 'SOMASE UG+FONTE' no arquivo do Excel
    for r_idx in range(2, planilha_repassar.max_row + 1):
        ug_fonte = planilha_repassar.cell(row=r_idx, column=13).value  # 'UG+FONTE'
        if ug_fonte in ug_fonte_totals:
            planilha_repassar.cell(row=r_idx, column=15, value=ug_fonte_totals[ug_fonte])

    # Calcular a coluna 'SOMASE SIAFI' manualmente
    siafi_totals = {}
    for r_idx in range(2, planilha_repassar.max_row + 1):
        siafi = planilha_repassar.cell(row=r_idx, column=9).value  # 'SIAFI' na coluna 'H'
        total = float(planilha_repassar.cell(row=r_idx, column=12).value)  # 'Total'
        if siafi not in siafi_totals:
            siafi_totals[siafi] = 0.0
        siafi_totals[siafi] += total

    # Atualizar a coluna 'SOMASE SIAFI' no arquivo do Excel
    for r_idx in range(2, planilha_repassar.max_row + 1):
        siafi = planilha_repassar.cell(row=r_idx, column=9).value  # 'SIAFI' na coluna 'H'
        if siafi in siafi_totals:
            planilha_repassar.cell(row=r_idx, column=17, value=siafi_totals[siafi])

    # Salvar o DataFrame de volta na aba 'REPASSAR' do arquivo original
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for c_idx, value in enumerate(row[1:], start=1):
            planilha_repassar.cell(row=r_idx, column=c_idx, value=value)

    # Salvar o arquivo com as alterações
    wb.save(arquivo_path)

    # Feche o arquivo
    wb.close()

# Caminho do arquivo principal
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para calcular e salvar a coluna 'SOMASE UG+FONTE'
calcular_e_salvar_somase_ug_fonte(arquivo_principal)

def preencherCondicionalLimite(arquivo_path):
    
    # Carregue o arquivo principal
    wb = openpyxl.load_workbook(arquivo_path)

    # Acesse a aba 'REPASSAR' no arquivo original
    planilha_repassar = wb['REPASSAR']

    # Carregue a aba 'REPASSAR' em um DataFrame
    data = planilha_repassar.values
    colunas = next(data)
    df = pd.DataFrame(data, columns=colunas)
    
    # Calcular a coluna 'CONDICIONAL LIMITE SAQUE'
    for r_idx in range(2, planilha_repassar.max_row + 1):
        limiteSaque = planilha_repassar.cell(row=r_idx, column=14).value  # Coluna "LIMITE DE SAQUE"
        somaseUg = planilha_repassar.cell(row=r_idx, column=15).value  # Coluna "SOMASE UG+FONTE"

    # Calcular a coluna 'CONDICIONAL LIMITE SAQUE'
    for r_idx in range(2, planilha_repassar.max_row + 1):
        limiteSaque = planilha_repassar.cell(row=r_idx, column=14).value  # Coluna "LIMITE DE SAQUE"
        somaseUg = planilha_repassar.cell(row=r_idx, column=15).value  # Coluna "SOMASE UG+FONTE"

        # Certifique-se de que ambos os valores não sejam None
        if limiteSaque is not None and somaseUg is not None:
            limiteSaque = float(limiteSaque)  
            somaseUg = float(somaseUg)

            if limiteSaque >= somaseUg:
                planilha_repassar.cell(row=r_idx, column=16, value="NÃO")
            else:
                planilha_repassar.cell(row=r_idx, column=16, value="REPASSAR")
           
    # Salvar o DataFrame de volta na aba 'REPASSAR' do arquivo original
    for r_idx, row in enumerate(df.itertuples(), start=2):
        for c_idx, value in enumerate(row[1:], start=1):
            planilha_repassar.cell(row=r_idx, column=c_idx, value=value)
                       
    # Salvar o arquivo com as alterações
    wb.save(arquivo_path)

    # Feche o arquivo
    wb.close()                
                
# Caminho do arquivo principal
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'      

# invocar a função
preencherCondicionalLimite(arquivo_principal)

print('Preenchendo a coluna Contas a repassar...')

def preencher_coluna_contas_repassar(arquivo_path):
    # Carregue o arquivo principal
    wb = openpyxl.load_workbook(arquivo_path)

    # Acesse a aba 'REPASSAR'
    planilha_repassar = wb['REPASSAR']

    # Carregue a aba 'Contas Cadastro e Controle'
    planilha_contas_cadastro = wb['Contas Cadastro e Controle']
    
    # Percorra a aba 'REPASSAR'
    for r_idx in range(2, planilha_repassar.max_row + 1):
        siafi = str(planilha_repassar.cell(row=r_idx, column=9).value)  # Força a leitura como string

        # Converta o valor da coluna SOMASE SIAFI para um número (remove formatação contábil)
        somase_siafi = float(str(planilha_repassar.cell(row=r_idx, column=17).value).replace(',', '.'))

        # Percorra a aba 'Contas Cadastro e Controle' para encontrar uma correspondência
        for row in planilha_contas_cadastro.iter_rows(min_row=3):
            conta_corrente = row[2].value  # Coluna "Conta Corrente"
            valores_firmados = row[5].value  # Coluna "VALORES FIRMADOS"

            if siafi == conta_corrente:
                planilha_repassar.cell(row=r_idx, column=18, value=valores_firmados)
                break
            
    # Salve o arquivo com as alterações
    wb.save(arquivo_path)

    # Feche o arquivo
    wb.close()

# Caminho do arquivo principal
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para preencher a coluna "CONTAS REPASSAR"
preencher_coluna_contas_repassar(arquivo_principal)

print('Executando...')

def preencherCondicional(arquivo_path):
    # Carregue o arquivo principal
    wb = openpyxl.load_workbook(arquivo_path)

    # Acesse a aba 'REPASSAR'
    planilha_repassar = wb['REPASSAR']
                
    # Preencha a coluna "CONDICIONAL CONTAS A REPASSAR" com base na lógica
    for r_idx in range(2, planilha_repassar.max_row + 1):
        P = planilha_repassar.cell(row=r_idx, column=17).value  # Coluna "SOMASE SIAFI"
        Q = planilha_repassar.cell(row=r_idx, column=18).value  # Coluna "CONTAS REPASSAR"
        
        # Verifica se Q (CONTAS REPASSAR) é None ou vazio
        if Q is None or Q == '':
            condicional = "SEM SALDO"
        else:
            # Verifica se P (SOMASE SIAFI) é None
            if P is None:
                condicional = "SEM SALDO"
            else:
                condicional = "COM SALDO" if P <= Q else "SEM SALDO"
        
        planilha_repassar.cell(row=r_idx, column=19, value=condicional)  # Coluna "CONDICIONAL CONTAS A REPASSAR"

    # Salve o arquivo com as alterações
    wb.save(arquivo_path)

    # Feche o arquivo
    wb.close()

# Caminho do arquivo principal
arquivo_principal = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para processar o arquivo
preencherCondicional(arquivo_principal)

print('buscando os dados da planilha NCPF...')

def processar_e_copiar_ncpf(file_path):
    # Crie um novo arquivo "COPIA NCPF.xlsx" no caminho desejado
    copia_ncpf_path = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA NCPF.xlsx'
    wb_copia = openpyxl.Workbook()
    sheet_copia = wb_copia.active

    # Copie os dados da planilha original para a nova planilha
    wb_ncpf = openpyxl.load_workbook(file_path, data_only=True)
    sheet_ncpf = wb_ncpf.active

    for row in sheet_ncpf.iter_rows(values_only=True):
        sheet_copia.append(row)

    # Encontre as células mescladas e as desmescla
    for merged_cell in sheet_copia.merged_cells.ranges:
        for row in merged_cell:
            sheet_copia.unmerge_cells(merged_cell.coord)

    # Quebra de Texto (somente se o valor for uma string)
    for row in sheet_copia.iter_rows(min_row=2, max_row=sheet_copia.max_row):
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.split('\n')

    # Deleta a segunda linha
    sheet_copia.delete_rows(2)

    # Adicionar cabeçalho NC-PF
    header_row = sheet_copia[1]
    header_cell = 'D1'
    sheet_copia[header_cell] = "NC-PF"

    # Faz o cálculo das colunas Valor descentralizado - Valor Repassado e preenche a coluna NC-PF
    for row in sheet_copia.iter_rows(min_row=2, max_row=sheet_copia.max_row):
        total_descentralizado = row[1].value
        total_repassado = row[2].value

        if total_descentralizado is not None and total_repassado is not None:
            nc_pf = total_descentralizado - total_repassado
            cell_nc_pf = 'D' + str(row[0].row)
            sheet_copia[cell_nc_pf] = nc_pf
        
    # Salve as alterações no arquivo copiado no caminho desejado
    wb_copia.save(copia_ncpf_path)
    wb_copia.close()

    # Agora você tem uma cópia do arquivo "NCPF.xlsx" processado e pronto para uso.

# Caminho do arquivo principal
arquivo_ncpf = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\NCPF.xlsx'
processar_e_copiar_ncpf(arquivo_ncpf)

print('copiando dados da planilha NCPF para a aba NC-PF da planinha COPIA TED Liquidação Geral - EXERCÍCIO 2026...')

def copiar_colunas_para_aba_nc_pf(arquivo_origem, arquivo_destino):
    # Abra o arquivo de origem (COPIA NCPF.xlsx)
    wb_origem = openpyxl.load_workbook(arquivo_origem)
    sheet_origem = wb_origem.active

    # Abra o arquivo de destino (COPIA TED Liquidação Geral - EXERCÍCIO 2023.xlsx)
    wb_destino = openpyxl.load_workbook(arquivo_destino)
    
    # Verifique se a aba "NC-PF" já existe no arquivo de destino
    if "NC-PF" in wb_destino.sheetnames:
        # Se a aba existir, exclua-a para sobrescrever os dados
        wb_destino.remove(wb_destino["NC-PF"])
    
    # Crie uma nova aba chamada "NC-PF" no arquivo de destino
    wb_destino.create_sheet("NC-PF")    
        
    # Selecione a aba "NC-PF" no arquivo de destino
    sheet_destino = wb_destino["NC-PF"]

    # Copie as colunas "TED" e "NC-PF" para a aba "NC-PF" do arquivo de destino
    for row in sheet_origem.iter_rows(min_row=1, max_row=sheet_origem.max_row, values_only=True):
        # Coluna "TED" está na coluna B e "NC-PF" na coluna D do arquivo de origem
        ted_value = row[0]  # Coluna A
        nc_pf_value = row[3]  # Coluna D
        sheet_destino.append([ted_value, nc_pf_value])

    # Salve as alterações no arquivo de destino
    wb_destino.save(arquivo_destino)

# Caminho do arquivo de origem
arquivo_ncpf = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA NCPF.xlsx'

# Caminho do arquivo de destino
arquivo_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chame a função para copiar as colunas para a aba "NC-PF"
copiar_colunas_para_aba_nc_pf(arquivo_ncpf, arquivo_destino)

print('Preenchendo a coluna SIMEC NC-PF...')

def preencher_coluna_simec(arquivo_destino, aba_repassar, aba_nc_pf):
    # Abra o arquivo de destino
    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_repassar = wb_destino[aba_repassar]
    sheet_nc_pf = wb_destino[aba_nc_pf]

    # Crie um dicionário para mapear os valores TED da aba "NC-PF" para os valores NC-PF
    ted_nc_pf_map = {}

    for row in sheet_nc_pf.iter_rows(min_row=2, values_only=True):
        ted_value = row[0]  # Coluna "TED" da aba "NC-PF" (coluna A)
        nc_pf_value = row[1]  # Coluna "NC-PF" da aba "NC-PF" (coluna B)
        ted_nc_pf_map[ted_value] = nc_pf_value

    # Percorra a coluna "TED" da aba "Repassar" (coluna D)
    for row in sheet_repassar.iter_rows(min_row=2):
        ted_repassar = row[4].value  # Coluna "TED" da aba "Repassar" (coluna D)

        # Verifique se o valor TED da aba "Repassar" existe no dicionário
        if ted_repassar in ted_nc_pf_map:
            nc_pf_value = ted_nc_pf_map[ted_repassar]

            # Preencha a coluna "SIMEC (NC-PF)" da aba "Repassar" (coluna S) com o valor NC-PF correspondente
            row[19].value = nc_pf_value
            
    # Salve as alterações no arquivo de destino
    wb_destino.save(arquivo_destino)

# Caminho do arquivo de destino
arquivo_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Nome das abas
aba_repassar = "REPASSAR"
aba_nc_pf = "NC-PF"

# Chame a função para preencher a coluna "SIMEC (NC-PF)" na aba "REPASSAR"
preencher_coluna_simec(arquivo_destino, aba_repassar, aba_nc_pf)

print('Preenchendo a coluna condicional SIMEC...')

def preencher_condicional_simec(arquivo_destino, aba_repassar):
    # Abra o arquivo de destino
    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_repassar = wb_destino[aba_repassar]

    # Define o número de casas decimais para arredondar
    casas_decimais = 2
    
    # Cor de preenchimento amarelo
    yellow_fill = PatternFill(start_color="E4BB02", end_color="E4BB02", fill_type="lightUp")

    # Função para converter string para float
    def converter_para_float(valor):
        if isinstance(valor, str):
            # Remove pontos e substitui vírgula por ponto
            valor = valor.replace('.', '').replace(',', '.')
            try:
                return float(valor)
            except ValueError:
                return None
        return valor

    # Percorra a coluna "SIMEC (NC-PF)" (coluna S) e a coluna "CONDICIONAL SIMEC" (coluna T)
    for row in sheet_repassar.iter_rows(min_row=2):
        simec_value = row[16].value  # Coluna S
        condicional_simec = row[19].value  # Coluna T
        valor_k = converter_para_float(row[11].value)  # Coluna K
        valor_m = converter_para_float(row[13].value)  # Coluna M

        # Verifique se os valores de K e M são numéricos
        if isinstance(valor_k, (int, float)) and isinstance(valor_m, (int, float)):
            # Verifique se valor_k < valor_m e aplique a cor amarela
            if valor_k < valor_m:
                print(f"Pintando linha {row[0].row} de amarelo: K={valor_k}, M={valor_m}")
                for cell in row:
                    cell.fill = yellow_fill
            else:
                print(f"Linha {row[0].row} não pintada: K={valor_k}, M={valor_m}")
        else:
            print(f"Linha {row[0].row} com valores não numéricos em K ou M: K={valor_k}, M={valor_m}")
                
        # Verifique se os valores de SIMEC e CONDICIONAL SIMEC são numéricos
        if isinstance(simec_value, (int, float)) and isinstance(condicional_simec, (int, float)):
            # Arredonde os valores para o número de casas decimais desejado
            simec_value = round(simec_value, casas_decimais)
            condicional_simec = round(condicional_simec, casas_decimais)

            if simec_value == condicional_simec:
                row[20].value = "REPASSAR"
            elif simec_value > condicional_simec:
                row[20].value = "NÃO"
            else:
                row[20].value = "REPASSAR"

    # Formate as colunas com os índices 10, 12, 13, 15, 16 e 18
    colunas_a_formatar = [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]  # Verifique os índices reais das colunas
    for coluna_index in colunas_a_formatar:
        for row in sheet_repassar.iter_rows(min_row=2):
            cell = row[coluna_index]
            value = cell.value

            # Verifique se o valor é numérico ou nulo
            if value is not None and isinstance(value, (int, float)):
                formatted_value = formatar_contabil(value)
                cell.value = formatted_value
            elif value is None:
                cell.value = None  # Deixe as células nulas como estão
        
        # Salve as alterações no arquivo de destino
        wb_destino.save(arquivo_destino)

# Caminho do arquivo de destino
arquivo_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Nome da aba "REPASSAR"
aba_repassar = "REPASSAR"

# Chame a função para preencher a coluna "CONDICIONAL SIMEC" na aba "REPASSAR"
preencher_condicional_simec(arquivo_destino, aba_repassar)

def mover_linhas_emendas(arquivo_destino):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(arquivo_destino)

    # Verificar se a aba "REPASSAR" existe
    if "REPASSAR" not in wb.sheetnames:
        print("Aba 'REPASSAR' não encontrada no arquivo!")
        return

    sheet_repassar = wb["REPASSAR"]

    # Criar a aba "EMENDAS" se não existir
    if "EMENDAS" not in wb.sheetnames:
        wb.create_sheet("EMENDAS")

    sheet_emendas = wb["EMENDAS"]

    # Carregar os dados da aba "REPASSAR" em um DataFrame
    df_repassar = pd.read_excel(arquivo_destino, sheet_name="REPASSAR")

    # Verificar se a coluna "Resultado EOF" existe
    if "Resultado EOF" not in df_repassar.columns:
        print("Coluna 'Resultado EOF' não encontrada na aba 'REPASSAR'!")
        return

    # Filtrar as linhas onde "Resultado EOF" é 6, 7 ou 8
    df_emendas = df_repassar[df_repassar["Resultado EOF"].isin([6, 7, 8])]
    
    if df_emendas.empty:
        print("Nenhuma linha para mover para 'EMENDAS'.")
        return

    # Descobrir a última linha preenchida na aba "EMENDAS"
    ultima_linha_emendas = sheet_emendas.max_row

    # Copiar cabeçalho se "EMENDAS" estiver vazia
    if ultima_linha_emendas == 1 and sheet_emendas.cell(row=1, column=1).value is None:
        for col_idx, col_name in enumerate(df_emendas.columns, start=1):
            sheet_emendas.cell(row=1, column=col_idx, value=col_name)

    # Copiar as linhas da aba "REPASSAR" para "EMENDAS", mantendo a formatação
    for _, row in df_emendas.iterrows():
        nova_linha_emendas = sheet_emendas.max_row + 1
        linha_original = row.name + 2  # Ajustar índice do Pandas para linha no Excel

        for col_idx, value in enumerate(row, start=1):
            cell_origem = sheet_repassar.cell(row=linha_original, column=col_idx)
            cell_destino = sheet_emendas.cell(row=nova_linha_emendas, column=col_idx, value=value)

            # Clonar o estilo da célula
            if cell_origem.fill:
                cell_destino.fill = PatternFill(
                    start_color=cell_origem.fill.start_color.rgb,
                    end_color=cell_origem.fill.end_color.rgb,
                    fill_type=cell_origem.fill.fill_type
                )

    print("Linhas copiadas para 'EMENDAS' com cores preservadas!")

    # Remover as linhas filtradas da aba "REPASSAR"
    linhas_para_remover = sorted([idx + 2 for idx in df_emendas.index], reverse=True)

    for linha in linhas_para_remover:
        sheet_repassar.delete_rows(linha)

    print("Linhas removidas de 'REPASSAR' mantendo formatação das outras linhas!")

    # Salvar alterações no arquivo Excel
    wb.save(arquivo_destino)

# Caminho do arquivo de destino
arquivo_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Chamar a função para mover as linhas para "EMENDAS" e limpar "REPASSAR"
mover_linhas_emendas(arquivo_destino)

def preencher_condicional_simec2(arquivo_destino, aba_emendas):
    # Abra o arquivo de destino
    wb_destino = openpyxl.load_workbook(arquivo_destino)
    sheet_repassar = wb_destino[aba_emendas]

    # Define o número de casas decimais para arredondar
    casas_decimais = 2
    
    # Cor de preenchimento amarelo
    yellow_fill = PatternFill(start_color="E4BB02", end_color="E4BB02", fill_type="lightUp")

    # Função para converter string para float
    def converter_para_float(valor):
        if isinstance(valor, str):
            # Remove pontos e substitui vírgula por ponto
            valor = valor.replace('.', '').replace(',', '.')
            try:
                return float(valor)
            except ValueError:
                return None
        return valor

    # Percorra a coluna "SIMEC (NC-PF)" (coluna S) e a coluna "CONDICIONAL SIMEC" (coluna T)
    for row in sheet_repassar.iter_rows(min_row=2):
        simec_value = row[16].value  # Coluna S
        condicional_simec = row[19].value  # Coluna T
        valor_k = converter_para_float(row[11].value)  # Coluna K
        valor_m = converter_para_float(row[13].value)  # Coluna M

        # Verifique se os valores de K e M são numéricos
        if isinstance(valor_k, (int, float)) and isinstance(valor_m, (int, float)):
            # Verifique se valor_k < valor_m e aplique a cor amarela
            if valor_k < valor_m:
                print(f"Pintando linha {row[0].row} de amarelo: K={valor_k}, M={valor_m}")
                for cell in row:
                    cell.fill = yellow_fill
            else:
                print(f"Linha {row[0].row} não pintada: K={valor_k}, M={valor_m}")
        else:
            print(f"Linha {row[0].row} com valores não numéricos em K ou M: K={valor_k}, M={valor_m}")
                
        # Verifique se os valores de SIMEC e CONDICIONAL SIMEC são numéricos
        if isinstance(simec_value, (int, float)) and isinstance(condicional_simec, (int, float)):
            # Arredonde os valores para o número de casas decimais desejado
            simec_value = round(simec_value, casas_decimais)
            condicional_simec = round(condicional_simec, casas_decimais)

            if simec_value == condicional_simec:
                row[20].value = "REPASSAR"
            elif simec_value > condicional_simec:
                row[20].value = "NÃO"
            else:
                row[20].value = "REPASSAR"

    # Formate as colunas com os índices 10, 12, 13, 15, 16 e 18
    colunas_a_formatar = [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]  # Verifique os índices reais das colunas
    for coluna_index in colunas_a_formatar:
        for row in sheet_repassar.iter_rows(min_row=2):
            cell = row[coluna_index]
            value = cell.value

            # Verifique se o valor é numérico ou nulo
            if value is not None and isinstance(value, (int, float)):
                formatted_value = formatar_contabil(value)
                cell.value = formatted_value
            elif value is None:
                cell.value = None  # Deixe as células nulas como estão
        
        # Salve as alterações no arquivo de destino
        wb_destino.save(arquivo_destino)

# Caminho do arquivo de destino
arquivo_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'

# Nome da aba "REPASSAR"
aba_emendas = "EMENDAS"

# Chame a função para preencher a coluna "CONDICIONAL SIMEC" na aba "REPASSAR"
preencher_condicional_simec2(arquivo_destino, aba_emendas)

def copiar_arquivo_liquidação():
    origem = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\COPIA TED Liquidação Geral - EXERCÍCIO 2026.xlsx'
    destino = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\LIQUIDAÇÃO A REPASSAR.xlsx'  # Nome do arquivo de destino

    try:
        shutil.copy(origem, destino)
        print(f'Arquivo copiado com sucesso para {destino}')
    except Exception as e:
        print(f'Ocorreu um erro ao copiar o arquivo: {e}')

# Chame a função para copiar o arquivo
copiar_arquivo_liquidação()

def copiar_arquivo_liquidação2():
    origem = r'W:\B - TED\7 - AUTOMAÇÃO\Liquidação\LIQUIDAÇÃO A REPASSAR.xlsx'
    destino_base = r'W:\B - TED\4 - FINANCEIRO\4.1 - LIQUIDAÇÃO\2026' # Pasta base de destino

    try:
        # Obtém a data atual
        data_atual = datetime.now()
        
        # Mapeia o número do mês para o nome do mês
        meses = {
            1: "01-Janeiro", 2: "02-Fevereiro", 3: "03-Março", 4: "04-Abril",
            5: "05-Maio", 6: "06-Junho", 7: "07-Julho", 8: "08-Agosto",
            9: "09-Setembro", 10: "10-Outubro", 11: "11-Novembro", 12: "12-Dezembro"
        }
        
        # Obtém o número e o nome do mês atual
        mes_atual_numero = data_atual.month
        mes_atual_nome = meses[mes_atual_numero]

        # Cria o caminho da pasta do mês
        pasta_mes = os.path.join(destino_base, mes_atual_nome)

        # Se a pasta do mês não existir, cria ela
        if not os.path.exists(pasta_mes):
            os.makedirs(pasta_mes)
            print(f"Pasta do mês criada: {pasta_mes}")

        # Define o nome do arquivo no formato "LIQUIDAÇÃO A REPASSAR DD-MM-AA.xlsx"
        nome_arquivo = f"LIQUIDAÇÃO A REPASSAR {data_atual.strftime('%d-%m-%y')}.xlsx"
        destino_final = os.path.join(pasta_mes, nome_arquivo)

        # Copia o arquivo para o destino final
        shutil.copy(origem, destino_final)
        print(f'Arquivo copiado com sucesso para {destino_final}')
    except Exception as e:
        print(f'Ocorreu um erro ao copiar o arquivo: {e}')

# Chame a função para copiar o arquivo
copiar_arquivo_liquidação2()

print('Processo Totalmente Finalizado. O seu arquivo final se encontra em "W:\B - TED\7 - AUTOMAÇÃO\Liquidação\LIQUIDAÇÃO A REPASSAR.xlsx"')
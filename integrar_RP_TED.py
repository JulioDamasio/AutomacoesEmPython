import pandas as pd
import shutil
import os
import xlwings as xw
from datetime import datetime
from openpyxl import load_workbook

def formatar_contabil(value):
    if pd.notnull(value):
        if isinstance(value, (int, float)):
            return "{:,.2f}".format(float(value)).replace(",", "_").replace(".", ",").replace("_", ".")
        else:
            return value  # Mantém o cabeçalho ou outros valores não numéricos
    return None

def integrar_rp_ted():
    # Caminho base
    pasta_base = r"W:\B - TED\7 - AUTOMAÇÃO\TEDS UG-RP"

    # Caminhos originais e modelo
    caminho_teds = os.path.join(pasta_base, "TEDs na UG intermediaria.xlsx")
    caminho_relatorio = os.path.join(pasta_base, "RELATORIO ANALITICO-2025.xlsx")
    caminho_modelo = os.path.join(pasta_base, "Modelo TEDs.xlsx")

    # Cria cópias para não alterar os originais
    copia_teds = os.path.join(pasta_base, "COPIA_TEDs_intermediaria.xlsx")
    copia_relatorio = os.path.join(pasta_base, "COPIA_RELATORIO_ANALITICO.xlsx")

    shutil.copy(caminho_teds, copia_teds)
    shutil.copy(caminho_relatorio, copia_relatorio)

    print("✅ Cópias criadas com sucesso. Manipulando apenas as cópias...")

    # Lê as planilhas
    teds = pd.read_excel(copia_teds)
    relatorio = pd.read_excel(copia_relatorio)

    # Diagnóstico de colunas (útil para conferir nomes)
    print("\n🧭 Colunas na planilha TEDs na UG intermediaria:")
    print(teds.columns.tolist())
    print("\n🧭 Colunas na planilha RELATORIO ANALITICO-2025:")
    print(relatorio.columns.tolist())

    # Ajuste de nomes (confirme se esses são os nomes exatos)
    chave_teds = "PTRES (Orçamentário)"   # Coluna na TEDs
    chave_relatorio = "PTRES"             # Coluna na RELATÓRIO
    coluna_rp = "RP"                      # Coluna a ser trazida
    coluna_valor = "Valor Autorizado (R$)"    # Coluna usada pra formatação

    # Remove duplicatas no relatório (para evitar múltiplas combinações no merge)
    relatorio_unico = relatorio.drop_duplicates(subset=[chave_relatorio], keep="first")

    # Faz o merge (integração)
    resultado = pd.merge(
        teds,
        relatorio_unico[[chave_relatorio, coluna_rp]],
        left_on=chave_teds,
        right_on=chave_relatorio,
        how="left"
    )

    # Remove coluna duplicada PTRES, caso tenha vindo no merge
    if chave_relatorio in resultado.columns and chave_relatorio != chave_teds:
        resultado.drop(columns=[chave_relatorio], inplace=True)

    # Verifica PTRES não encontrados
    nao_encontrados = resultado[resultado[coluna_rp].isna()]
    if not nao_encontrados.empty:
        print(f"\n⚠️ {len(nao_encontrados)} PTRES não foram encontrados no relatório.")
        print("Exemplos:", nao_encontrados[chave_teds].head().tolist())
    else:
        print("\n✅ Todos os PTRES foram encontrados com sucesso.")

    # 🔹 NÃO remove mais linhas com Valor Autorizado vazio ou 0
    # Apenas avisa caso a coluna não exista
    if coluna_valor not in resultado.columns:
        print(f"⚠️ Coluna '{coluna_valor}' não encontrada — confirme o nome exato acima.")

    # Aplica formatação contábil se a coluna existir
    if coluna_valor in resultado.columns:
        resultado[coluna_valor] = resultado[coluna_valor].apply(formatar_contabil)
        print(f"💰 Coluna '{coluna_valor}' formatada com sucesso.")
    else:
        print(f"⚠️ Coluna '{coluna_valor}' não encontrada para formatação.")

    # Se não existir o modelo, salva apenas o resultado normal
    data_hoje = datetime.now().strftime("%d-%m-%Y")
    nome_final = f"TEDs_integradas_{data_hoje}.xlsx"
    caminho_final = os.path.join(pasta_base, nome_final)

    if not os.path.exists(caminho_modelo):
        resultado.to_excel(caminho_final, index=False)
        print(f"💾 Modelo não encontrado — planilha final salva em: {caminho_final}")
        return caminho_final

    # --- Carrega o modelo e escreve os dados a partir da linha 2 (abaixo do cabeçalho) ---
    wb = load_workbook(caminho_modelo)
    ws = wb.active  # usa a primeira aba do modelo

    # Define a linha inicial fixa (1 = cabeçalho, começa a escrever na 2)
    linha_inicio = 2

    # Helper: escreve valor tratando células mescladas corretamente
    def set_cell_value_safe(ws, row, col, value, linha_dados_inicio):
        coord = ws.cell(row=row, column=col).coordinate
        for mr in ws.merged_cells.ranges:
            if coord in mr:
                if mr.min_row >= linha_dados_inicio:
                    if row == mr.min_row and col == mr.min_col:
                        ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
        ws.cell(row=row, column=col).value = value

    # Converte DataFrame em lista de listas (somente valores, sem cabeçalho)
    linhas_dados = resultado.values.tolist()

    # Escreve linha a linha (começando em linha_inicio)
    for i, linha in enumerate(linhas_dados, start=linha_inicio):
        for j, valor in enumerate(linha, start=1):
            set_cell_value_safe(ws, i, j, valor, linha_inicio)

    # Salva o workbook usando o modelo como base
    wb.save(caminho_final)
    print(f"\n💾 Planilha final salva com sucesso usando o modelo em: {caminho_final}")
    return caminho_final

if __name__ == "__main__":
    integrar_rp_ted()

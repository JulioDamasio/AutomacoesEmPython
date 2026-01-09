import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil
import os
import pandas as pd
import numpy as np
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings


def copiar_arquivos(pasta_destino):
    origens = [
       r'W:\B - TED\7 - AUTOMAÇÃO\Residência\Residência Acompanhamento.xlsx',
       r'W:\B - TED\7 - AUTOMAÇÃO\Residência\Limite de Saque Residência.xlsx',
       r'W:\B - TED\7 - AUTOMAÇÃO\Residência\TED - Contas Cadastro e Controle.xlsx'
    ]
    for origem in origens:
        caminho, nome_arquivo = os.path.split(origem)
        novo_nome = "COPIA " + nome_arquivo
        shutil.copy(origem, os.path.join(pasta_destino, novo_nome))  # Copia para a pasta destino com o novo nome
        print(f"Arquivo {nome_arquivo} copiado como {novo_nome} para {pasta_destino}")

def renomear_colunas_excel(arquivo, novos_nomes_colunas):
    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo)

        # Verificar se o número de novos nomes é igual ao número de colunas existentes
        if len(novos_nomes_colunas) != len(df.columns):
            raise ValueError("O número de novos nomes de colunas não corresponde ao número de colunas existentes.")

        # Renomear as colunas
        df.columns = novos_nomes_colunas

        # Salvar o DataFrame com os novos nomes de colunas no arquivo Excel
        df.to_excel(arquivo, index=False)

        print(f"As colunas do arquivo '{arquivo}' foram renomeadas com sucesso!")

    except Exception as e:
        print(f"Erro ao renomear as colunas: {e}")

def remover_colunas_excel(arquivo, colunas_para_remover, salvar_como=None):

    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Remover colunas especificadas
        df.drop(df.columns[colunas_para_remover], axis=1, inplace=True)
        
        # Salvar o resultado
        if salvar_como:
            df.to_excel(salvar_como, index=False)
        else:
            df.to_excel(arquivo, index=False)
        
        print("Colunas removidas com sucesso!")
        
    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")

def adicionar_coluna_excel(arquivo, indice_coluna, nome_coluna, valor_padrao=None, salvar_como=None):
    try:
        df = pd.read_excel(arquivo)
        df.insert(indice_coluna, nome_coluna, valor_padrao)
        
        if salvar_como:
            df.to_excel(salvar_como, index=False)
        else:
            df.to_excel(arquivo, index=False)
        
        print(f"Coluna '{nome_coluna}' adicionada com sucesso!")
    except Exception as e:
        print(f"Erro: {e}")

def substituir_valores_coluna(caminho_arquivo, coluna_idx, novo_valor):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active  # ou: ws = wb['Acompanhamento Residência'] se quiser garantir a aba

    for row in ws.iter_rows(min_row=2, min_col=coluna_idx+1, max_col=coluna_idx+1):
        for cell in row:
            cell.value = novo_valor

    wb.save(caminho_arquivo)        

def concatenar_colunas_excel(arquivo, coluna1_idx, coluna2_idx, coluna_resultado_idx, salvar_como=None):

    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Garantir que os índices das colunas estejam no intervalo correto
        if coluna1_idx >= len(df.columns) or coluna2_idx >= len(df.columns):
            raise IndexError("Índice de coluna para concatenar fora do intervalo válido.")
        
        # Garantir que o índice da coluna de resultado esteja no intervalo
        if coluna_resultado_idx >= len(df.columns):
            raise IndexError("Índice de coluna de resultado fora do intervalo válido.")
        
        # Concatenar as colunas especificadas sem separador
        df.iloc[:, coluna_resultado_idx] = df.iloc[:, coluna1_idx].astype(str) + df.iloc[:, coluna2_idx].astype(str)
        
        # Salvar o DataFrame atualizado
        if salvar_como:
            df.to_excel(salvar_como, index=False)
        else:
            df.to_excel(arquivo, index=False)

        print(f"Colunas {coluna1_idx} e {coluna2_idx} concatenadas com sucesso no índice {coluna_resultado_idx} no arquivo {arquivo}!")
        
    except Exception as e:
        print(f"Erro ao processar o arquivo {arquivo}: {e}")

def concatenar_colunas_excel_por_aba(arquivo, aba, coluna1_idx, coluna2_idx, coluna_resultado_idx, salvar_como=None):
    try:
        # Carregar todas as abas do arquivo Excel
        sheets = pd.read_excel(arquivo, sheet_name=None)

        # Verificar se a aba desejada está presente no arquivo
        if aba not in sheets:
            raise ValueError(f"A aba '{aba}' não foi encontrada no arquivo.")

        # Carregar o DataFrame da aba específica
        df = sheets[aba]

        # Garantir que os índices das colunas estejam no intervalo correto
        if coluna1_idx >= len(df.columns) or coluna2_idx >= len(df.columns):
            raise IndexError("Índice de coluna para concatenar fora do intervalo válido.")
        
        # Garantir que o índice da coluna de resultado esteja no intervalo
        if coluna_resultado_idx >= len(df.columns):
            raise IndexError("Índice de coluna de resultado fora do intervalo válido.")

        # Concatenar as colunas especificadas sem separador
        df.iloc[:, coluna_resultado_idx] = df.iloc[:, coluna1_idx].astype(str) + df.iloc[:, coluna2_idx].astype(str)

        # Atualizar o DataFrame modificado na aba
        sheets[aba] = df

        # Salvar todas as abas no arquivo Excel (sobrescrevendo ou salvando como novo arquivo)
        with pd.ExcelWriter(salvar_como or arquivo, engine='openpyxl') as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Colunas {coluna1_idx} e {coluna2_idx} concatenadas com sucesso na aba '{aba}' no arquivo {arquivo}!")

    except Exception as e:
        print(f"Erro ao processar o arquivo {arquivo}: {e}")    
    
def calcular_expressao_saldo(arquivo, col_a_idx, col_b_idx, col_c_idx, col_d_idx, col_resultado_idx, salvar_como=None):
    
    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Garantir que os índices das colunas estejam no intervalo correto
        num_colunas = len(df.columns)
        if any(idx >= num_colunas for idx in [col_a_idx, col_b_idx, col_c_idx, col_d_idx, col_resultado_idx]):
            raise IndexError("Um dos índices fornecidos está fora do intervalo de colunas disponíveis.")
        
        # Calcular a expressão A - (B - C) + D
        df.iloc[:, col_resultado_idx] = (
            df.iloc[:, col_a_idx] - df.iloc[:, col_b_idx] - df.iloc[:, col_c_idx] + df.iloc[:, col_d_idx])
        
        # Salvar o DataFrame atualizado
        if salvar_como:
            df.to_excel(salvar_como, index=False)
        else:
            df.to_excel(arquivo, index=False)
        
        print(f"Expressão A - (B - C) + D calculada com sucesso e salva na coluna de índice {col_resultado_idx} no arquivo {arquivo}!")
    
    except Exception as e:
        print(f"Erro ao processar o arquivo {arquivo}: {e}")

def calcular_subtracao_SD(arquivo, col_a_idx, col_b_idx, col_resultado_idx, salvar_como=None):

    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo)
        
        # Garantir que os índices das colunas estejam no intervalo correto
        num_colunas = len(df.columns)
        if any(idx >= num_colunas for idx in [col_a_idx, col_b_idx, col_resultado_idx]):
            raise IndexError("Um dos índices fornecidos está fora do intervalo de colunas disponíveis.")
        
        # Substituir valores None/NaN por 0 para evitar erros durante a subtração
        df.iloc[:, col_a_idx] = df.iloc[:, col_a_idx].fillna(0)
        df.iloc[:, col_b_idx] = df.iloc[:, col_b_idx].fillna(0)
        
        # Calcular a expressão A - B
        resultado = df.iloc[:, col_a_idx] - df.iloc[:, col_b_idx]
        
        # Aplicar formatação para exibir valores negativos com sinal de "-"
        df.iloc[:, col_resultado_idx] = resultado.apply(lambda x: f"-{abs(x)}" if x < 0 else x)
        
        # Salvar o DataFrame atualizado
        if salvar_como:
            df.to_excel(salvar_como, index=False)
        else:
            df.to_excel(arquivo, index=False)
        
        print(f"Subtração A - B calculada com sucesso e salva na coluna de índice {col_resultado_idx} no arquivo {arquivo}!")
    
    except Exception as e:
        print(f"Erro ao processar o arquivo {arquivo}: {e}")
          
def tabela_dinamica(arquivo_origem, colunas_linhas, colunas_valores_idx, funcao_agregacao='sum', sheet_name='Tabela Dinâmica'):
    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo_origem)

        # Garantir que os índices das colunas estejam no intervalo correto
        num_colunas = len(df.columns)
        for idx in colunas_linhas + colunas_valores_idx:
            if idx >= num_colunas:
                raise IndexError(f"O índice {idx} está fora do intervalo de colunas disponíveis.")
        
        # Obter os nomes das colunas com base nos índices
        colunas_linhas_nomes = [df.columns[idx] for idx in colunas_linhas]
        colunas_valores_nomes = [df.columns[idx] for idx in colunas_valores_idx]

        # Criar a tabela dinâmica
        tabela_dinamica = pd.pivot_table(
            df,
            index=colunas_linhas_nomes,
            values=colunas_valores_nomes,
            aggfunc=funcao_agregacao,
            fill_value=0
        )

        # Carregar o arquivo Excel existente para verificar se a aba já existe
        with pd.ExcelWriter(arquivo_origem, engine='openpyxl', mode='a') as writer:
            # Verificar se a aba já existe e, se necessário, removê-la
            workbook = writer.book
            if sheet_name in workbook.sheetnames:
                std = workbook[sheet_name]
                workbook.remove(std)
                print(f"A aba '{sheet_name}' já existia e foi removida.")

            # Adicionar a tabela dinâmica como uma nova aba
            tabela_dinamica.to_excel(writer, sheet_name=sheet_name)

        print(f"Tabela dinâmica criada com sucesso e salva em uma nova aba '{sheet_name}' no arquivo {arquivo_origem}!")

    except Exception as e:
        print(f"Erro ao criar a tabela dinâmica: {e}")

def td_acompanhamento(arquivo_origem, colunas_linhas, sheet_name='Residência Acompanhamento'):
    try:
        # Carregar o arquivo Excel
        df = pd.read_excel(arquivo_origem)

        # Garantir que os índices das colunas estejam no intervalo correto
        num_colunas = len(df.columns)
        for idx in colunas_linhas:
            if idx >= num_colunas:
                raise IndexError(f"O índice {idx} está fora do intervalo de colunas disponíveis.")
        
        # Obter os nomes das colunas com base nos índices
        colunas_linhas_nomes = [df.columns[idx] for idx in colunas_linhas]

        # Selecionar apenas as colunas desejadas
        df_selecionado = df[colunas_linhas_nomes]

        # Agrupar as linhas, sem aplicar agregação, apenas agrupando pelos rótulos
        tabela_dinamica = df_selecionado.groupby(colunas_linhas_nomes, as_index=False).first()

        # Carregar o arquivo Excel existente para verificar se a aba já existe
        with pd.ExcelWriter(arquivo_origem, engine='openpyxl', mode='a') as writer:
            # Verificar se a aba já existe e, se necessário, removê-la
            workbook = writer.book
            if sheet_name in workbook.sheetnames:
                std = workbook[sheet_name]
                workbook.remove(std)
                print(f"A aba '{sheet_name}' já existia e foi removida.")

            # Adicionar a tabela dinâmica como uma nova aba
            tabela_dinamica.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Tabela dinâmica criada com sucesso e salva em uma nova aba '{sheet_name}' no arquivo {arquivo_origem}!")
    
    except Exception as e:
        print(f"Erro ao criar a tabela dinâmica: {e}")

def preencher_colunas_UG_linha_a_linha(copia_ted_para_finalizar, copia_arquivo_valores, coluna_destino, chave_coluna_origem, chave_coluna_destino, coluna_siafi_valores):
    try:
        # Carregar os DataFrames
        df_ted = pd.read_excel(copia_ted_para_finalizar)
        df_valores = pd.read_excel(copia_arquivo_valores)
        
        # Corrigir possíveis espaços e caracteres indesejados nas colunas
        df_ted.columns = df_ted.columns.str.strip()
        df_valores.columns = df_valores.columns.str.strip()
        
        # Verificar se as colunas esperadas estão presentes
        if chave_coluna_origem not in df_ted.columns:
            raise ValueError(f"A coluna chave '{chave_coluna_origem}' não foi encontrada em {copia_ted_para_finalizar}.")
        if chave_coluna_destino not in df_valores.columns:
            raise ValueError(f"A coluna chave '{chave_coluna_destino}' não foi encontrada em {copia_arquivo_valores}.")
        if coluna_siafi_valores not in df_valores.columns:
            raise ValueError(f"A coluna 'SIAFI' '{coluna_siafi_valores}' não foi encontrada em {copia_arquivo_valores}.")
        
        # Corrigir possíveis espaços e caracteres indesejados nas chaves
        df_valores[chave_coluna_destino] = df_valores[chave_coluna_destino].astype(str).str.strip().str.split('.').str[0]
        df_ted[chave_coluna_origem] = df_ted[chave_coluna_origem].astype(str).str.strip().str.split('.').str[0]
        
        # Exibir os primeiros 10 valores da coluna "UG Proponente" para depuração
        print("Exibindo os 10 primeiros valores da coluna 'UG Proponente' do arquivo de valores:")
        print(df_valores[chave_coluna_destino].head(10).to_list())
        
        # Contador para limitar a exibição das mensagens de correspondência ausente
        limite_prints = 10
        contagem_nao_encontrados = 0

        # Preencher a coluna destino linha a linha
        for i, row in df_ted.iterrows():
            chave_origem = row[chave_coluna_origem]
            
            # Exibir para depuração (mostrar apenas os 10 primeiros caracteres da chave)
            print(f"\nProcessando linha {i}: chave origem = {str(chave_origem)[:10]}")
            
            # Garantir que a chave origem não seja NaN ou vazia
            if pd.isna(chave_origem) or chave_origem == '':
                print(f"Chave origem NaN ou vazia na linha {i}, pulando.")
                continue
            
            # Encontrar a linha correspondente no df_valores
            correspondencia = df_valores[df_valores[chave_coluna_destino] == chave_origem]
            
            # Exibir para depuração (mostrar apenas os 10 primeiros resultados)
            if not correspondencia.empty:
                print("Primeiros 10 resultados da correspondência:")
                print(correspondencia.head(10)[[chave_coluna_destino, coluna_siafi_valores]])
            
            # Verificar se encontramos correspondência
            if not correspondencia.empty:
                # Pegar o valor da coluna SIAFI para a chave encontrada
                siafi_value = correspondencia.iloc[0][coluna_siafi_valores]
                print(f"Preenchendo SIAFI com: {siafi_value}")
                df_ted.at[i, coluna_destino] = siafi_value
            else:
                # Caso não encontre, manter 0 ou outro valor e controlar a quantidade de mensagens
                if contagem_nao_encontrados < limite_prints:
                    print(f"Nenhuma correspondência encontrada para chave {str(chave_origem)[:10]}, preenchendo com 0.")
                contagem_nao_encontrados += 1
                df_ted.at[i, coluna_destino] = 0

        # Informar se houve mais correspondências ausentes do que o limite de mensagens
        if contagem_nao_encontrados > limite_prints:
            print(f"\n{contagem_nao_encontrados} linhas sem correspondência. Mostradas apenas as primeiras {limite_prints}.")

        # Salvar as alterações no arquivo
        df_ted.to_excel(copia_ted_para_finalizar, index=False)
        
        print(f"\nColuna {coluna_destino} preenchida com sucesso...")
    
    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna {coluna_destino}: {e}")

def preencher_colunas_UG_linha_a_linha2(copia_ted_para_finalizar, copia_arquivo_valores, 
                                       aba_destino, aba_origem, coluna_destino, 
                                       chave_coluna_origem, chave_coluna_destino, coluna_siafi_valores):
    try:
        # Carregar os DataFrames a partir das abas específicas
        df_ted = pd.read_excel(copia_ted_para_finalizar, sheet_name=aba_destino)
        df_valores = pd.read_excel(copia_arquivo_valores, sheet_name=aba_origem)
        
        # Corrigir possíveis espaços e caracteres indesejados nas colunas
        df_ted.columns = df_ted.columns.str.strip()
        df_valores.columns = df_valores.columns.str.strip()
        
        # Verificar se as colunas esperadas estão presentes
        if chave_coluna_origem not in df_ted.columns:
            raise ValueError(f"A coluna chave '{chave_coluna_origem}' não foi encontrada na aba '{aba_destino}'.")
        if chave_coluna_destino not in df_valores.columns:
            raise ValueError(f"A coluna chave '{chave_coluna_destino}' não foi encontrada na aba '{aba_origem}'.")
        if coluna_siafi_valores not in df_valores.columns:
            raise ValueError(f"A coluna 'SIAFI' '{coluna_siafi_valores}' não foi encontrada na aba '{aba_origem}'.")
        
        # Corrigir possíveis espaços e caracteres indesejados nas chaves
        df_valores[chave_coluna_destino] = df_valores[chave_coluna_destino].astype(str).str.strip().str.split('.').str[0]
        df_ted[chave_coluna_origem] = df_ted[chave_coluna_origem].astype(str).str.strip().str.split('.').str[0]
        
        # Exibir os primeiros 10 valores da coluna "UG Proponente" para depuração
        print("Exibindo os 10 primeiros valores da coluna 'UG Proponente' do arquivo de valores:")
        print(df_valores[chave_coluna_destino].head(10).to_list())
        
        # Contador para limitar a exibição das mensagens de correspondência ausente
        limite_prints = 10
        contagem_nao_encontrados = 0

        # Preencher a coluna destino linha a linha
        for i, row in df_ted.iterrows():
            chave_origem = row[chave_coluna_origem]
            
            # Exibir para depuração (mostrar apenas os 10 primeiros caracteres da chave)
            print(f"\nProcessando linha {i}: chave origem = {str(chave_origem)[:10]}")
            
            # Garantir que a chave origem não seja NaN ou vazia
            if pd.isna(chave_origem) or chave_origem == '':
                print(f"Chave origem NaN ou vazia na linha {i}, pulando.")
                continue
            
            # Encontrar a linha correspondente no df_valores
            correspondencia = df_valores[df_valores[chave_coluna_destino] == chave_origem]
            
            # Exibir para depuração (mostrar apenas os 10 primeiros resultados)
            if not correspondencia.empty:
                print("Primeiros 10 resultados da correspondência:")
                print(correspondencia.head(10)[[chave_coluna_destino, coluna_siafi_valores]])
            
            # Verificar se encontramos correspondência
            if not correspondencia.empty:
                # Pegar o valor da coluna SIAFI para a chave encontrada
                siafi_value = correspondencia.iloc[0][coluna_siafi_valores]
                print(f"Preenchendo SIAFI com: {siafi_value}")
                df_ted.at[i, coluna_destino] = siafi_value
            else:
                # Caso não encontre, manter 0 ou outro valor e controlar a quantidade de mensagens
                if contagem_nao_encontrados < limite_prints:
                    print(f"Nenhuma correspondência encontrada para chave {str(chave_origem)[:10]}, preenchendo com 0.")
                contagem_nao_encontrados += 1
                df_ted.at[i, coluna_destino] = 0

        # Informar se houve mais correspondências ausentes do que o limite de mensagens
        if contagem_nao_encontrados > limite_prints:
            print(f"\n{contagem_nao_encontrados} linhas sem correspondência. Mostradas apenas as primeiras {limite_prints}.")
        
        # Salvar as alterações no arquivo Excel na aba especificada
        with pd.ExcelWriter(copia_ted_para_finalizar, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df_ted.to_excel(writer, sheet_name=aba_destino, index=False)
        
        print(f"\nColuna {coluna_destino} preenchida com sucesso na aba '{aba_destino}'...")
    
    except Exception as e:
        print(f"Ocorreu um erro ao preencher a coluna {coluna_destino}: {e}")        


def adicionar_colunas_aba(arquivo, aba, indices_colunas, nomes_colunas, valores_padrao, salvar_como=None):
    try:
        df = pd.read_excel(arquivo, sheet_name=aba)

        # Verifica se as listas têm o mesmo tamanho
        if not (len(indices_colunas) == len(nomes_colunas) == len(valores_padrao)):
            raise ValueError("As listas devem ter o mesmo tamanho.")

        # Adiciona as colunas (no fim, temporariamente)
        for nome, valor in zip(nomes_colunas, valores_padrao):
            if nome not in df.columns:
                df[nome] = valor
                print(f"Coluna '{nome}' adicionada com sucesso.")
            else:
                print(f"A coluna '{nome}' já existe. Nenhuma alteração feita.")

        # Reordena colunas com base nos índices desejados
        colunas_atuais = df.columns.tolist()

        # Cria mapeamento nome → índice alvo
        mapeamento = {nome: idx for nome, idx in zip(nomes_colunas, indices_colunas)}

        # Reordena colunas conforme o mapeamento
        for nome, idx in sorted(mapeamento.items(), key=lambda x: x[1]):
            colunas_atuais.remove(nome)
            colunas_atuais.insert(idx, nome)

        df = df[colunas_atuais]

        # Salva o arquivo
        caminho_saida = salvar_como if salvar_como else arquivo
        with pd.ExcelWriter(caminho_saida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=aba, index=False)

        print("Todas as colunas foram adicionadas e reordenadas com sucesso.")

    except Exception as e:
        print(f"Erro ao adicionar colunas na aba '{aba}': {e}")

def excluir_linhas_por_valores(arquivo, aba, coluna_filtro, valores_excluir, salvar_como=None):
    try:
        # Carregar todas as abas do arquivo Excel
        excel = pd.ExcelFile(arquivo)
        sheets = pd.read_excel(arquivo, sheet_name=None)  # Carregar todas as abas em um dicionário
        
        # Verificar se a aba desejada está presente no arquivo
        if aba not in sheets:
            raise ValueError(f"A aba '{aba}' não foi encontrada no arquivo.")
        
        # Carregar o DataFrame da aba específica
        df = sheets[aba]
        
        # Remover espaços extras das colunas (se existirem)
        df.columns = df.columns.str.strip()
        
        # Exibir os primeiros valores da coluna para confirmação (opcional)
        print(f"Valores iniciais na coluna '{coluna_filtro}':")
        print(df[coluna_filtro].head(10))
        
        # Excluir linhas onde a coluna contém os valores indesejados
        df = df[~df[coluna_filtro].astype(str).isin(valores_excluir)]
        
        # Atualizar a aba modificada no dicionário
        sheets[aba] = df
        
        # Salvar todas as abas no arquivo Excel (sobrescrevendo ou salvando como novo arquivo)
        with pd.ExcelWriter(salvar_como or arquivo, engine='openpyxl') as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"Linhas com valores {valores_excluir} na coluna '{coluna_filtro}' foram excluídas com sucesso da aba '{aba}'.")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def garantir_fontes_por_uge(arquivo, aba, coluna_uge, coluna_fonte, fontes_necessarias, salvar_como=None):
    try:
        # Carregar todas as abas do arquivo Excel
        sheets = pd.read_excel(arquivo, sheet_name=None)
        
        # Verificar se a aba desejada está presente no arquivo
        if aba not in sheets:
            raise ValueError(f"A aba '{aba}' não foi encontrada no arquivo.")
        
        # Carregar o DataFrame da aba específica
        df = sheets[aba]
        
        # Remover espaços extras das colunas (se existirem)
        df.columns = df.columns.str.strip()
        
        # Exibir os primeiros valores para depuração (opcional)
        print(f"Valores iniciais da coluna '{coluna_uge}':")
        print(df[coluna_uge].dropna().unique())

        # Lista para armazenar novas linhas
        novas_linhas = []

        # Iterar sobre cada "Órgão UGE" único
        for uge in df[coluna_uge].dropna().unique():
            # Filtrar linhas com o "Órgão UGE" atual
            df_filtrado = df[df[coluna_uge] == uge]

            # Verificar se as fontes necessárias estão presentes
            fontes_existentes = df_filtrado[coluna_fonte].dropna().unique()
            
            for fonte in fontes_necessarias:
                if fonte not in fontes_existentes:
                    # Se a fonte estiver ausente, pegar a primeira linha como base
                    nova_linha = df_filtrado.iloc[0].copy()
                    nova_linha[coluna_fonte] = fonte
                    novas_linhas.append(nova_linha)
        
        # Adicionar as novas linhas ao DataFrame original
        if novas_linhas:
            df = pd.concat([df, pd.DataFrame(novas_linhas)], ignore_index=True)
        
        # Ordenar o DataFrame por 'Órgão UGE' e depois por 'Fonte Recursos Detalhada'
        # Para garantir que '1000A0008U' venha antes de '1012A00008V', usamos sort_values com o parâmetro 'ascending'
        df = df.sort_values(by=[coluna_uge, coluna_fonte], 
                            key=lambda col: col.map(lambda x: (x != '1000A0008U', x)),
                            ascending=[True, True]).reset_index(drop=True)
        
        # Atualizar a aba modificada no dicionário
        sheets[aba] = df
        
        # Salvar todas as abas no arquivo Excel (sobrescrevendo ou salvando como novo arquivo)
        with pd.ExcelWriter(salvar_como or arquivo, engine='openpyxl') as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"As linhas com fontes {fontes_necessarias} foram garantidas e ordenadas para cada 'Órgão UGE' na aba '{aba}'.")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def preencher_coluna_valor_bolsa_tributos(arquivo, aba, coluna_bolsa, coluna_bolsa_tributos, salvar_como=None):
    try:
        # Carregar todas as abas do arquivo Excel
        sheets = pd.read_excel(arquivo, sheet_name=None)

        # Verificar se a aba desejada está presente no arquivo
        if aba not in sheets:
            raise ValueError(f"A aba '{aba}' não foi encontrada no arquivo.")
        
        # Carregar o DataFrame da aba específica
        df = sheets[aba]
        
        # Remover espaços extras das colunas (se existirem)
        df.columns = df.columns.str.strip()

        # Verificar se a coluna 'Valor Bolsa' existe
        if coluna_bolsa not in df.columns:
            raise ValueError(f"A coluna '{coluna_bolsa}' não foi encontrada na aba '{aba}'.")
        
        # Preencher a coluna 'Valor Bolsa + Tributos' com 20% do valor da coluna 'Valor Bolsa'
        df[coluna_bolsa_tributos] = df[coluna_bolsa] * 0.20

        # Atualizar a aba modificada no dicionário
        sheets[aba] = df

        # Salvar todas as abas no arquivo Excel (sobrescrevendo ou salvando como novo arquivo)
        with pd.ExcelWriter(salvar_como or arquivo, engine='openpyxl') as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"A coluna '{coluna_bolsa_tributos}' foi preenchida com 20% a mais do valor da coluna '{coluna_bolsa}' na aba '{aba}'.")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def preencher_coluna_estimativa_folha(arquivo, aba, col_bolsa, col_bolsa_tributos, col_limite_saque, col_estimativa_folha, salvar_como=None):
    try:
        # Carrega o arquivo e a aba especificada
        df = pd.read_excel(arquivo, sheet_name=aba)

        # Garante que os nomes de colunas não tenham espaços extras
        df.columns = df.columns.str.strip()

        # Verifica se as colunas existem
        for col in [col_bolsa, col_bolsa_tributos]:
            if col not in df.columns:
                raise ValueError(f"A coluna '{col}' não foi encontrada.")

        # Preenche a coluna 'Estimativa Folha' com a soma dos valores + 0.60
        df[col_estimativa_folha] = df[col_bolsa].fillna(0) + df[col_bolsa_tributos].fillna(0) - df[col_limite_saque].fillna(0) + 0.60

        # Salva o resultado
        with pd.ExcelWriter(salvar_como if salvar_como else arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=aba, index=False)

        print(f"Coluna '{col_estimativa_folha}' preenchida com sucesso na aba '{aba}'.")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def preencher_somase_despesas_liquidadas(arquivo, aba, coluna_siafi, coluna_despesas, coluna_somase, salvar_como=None):
    try:
        # Carregar todas as abas do arquivo Excel
        sheets = pd.read_excel(arquivo, sheet_name=None)

        # Verificar se a aba desejada está presente no arquivo
        if aba not in sheets:
            raise ValueError(f"A aba '{aba}' não foi encontrada no arquivo.")

        # Carregar o DataFrame da aba específica
        df = sheets[aba]

        # Remover espaços extras das colunas (se existirem)
        df.columns = df.columns.str.strip()

        # Verificar se as colunas necessárias existem
        if coluna_siafi not in df.columns:
            raise ValueError(f"A coluna '{coluna_siafi}' não foi encontrada na aba '{aba}'.")
        if coluna_despesas not in df.columns:
            raise ValueError(f"A coluna '{coluna_despesas}' não foi encontrada na aba '{aba}'.")
        
        # Criar um dicionário com a soma dos valores de "Despesas liquidadas a Pagar" para cada "SIAFI"
        soma_por_siafi = df.groupby(coluna_siafi)[coluna_despesas].sum()

        # Preencher a coluna "SOMASE despesas Liquidadas" com a soma correspondente ao "SIAFI"
        df[coluna_somase] = df[coluna_siafi].map(soma_por_siafi)

        # Atualizar a aba modificada no dicionário
        sheets[aba] = df

        # Salvar todas as abas no arquivo Excel (sobrescrevendo ou salvando como novo arquivo)
        with pd.ExcelWriter(salvar_como or arquivo, engine='openpyxl') as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"A coluna '{coluna_somase}' foi preenchida com a soma das despesas liquidadas agrupadas por '{coluna_siafi}' na aba '{aba}'.")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def manipular_excel(arquivo, salvar_como=None):
    try:
        # Carregar o arquivo Excel sem considerar cabeçalho
        df = pd.read_excel(arquivo, header=None)
        
        # Verificar se a linha de índice 1 existe
        if len(df) > 1:
            # Apagar a linha de índice 1
            df = df.drop(index=1).reset_index(drop=True)
            print("Linha de índice 1 removida com sucesso.")
        else:
            print("A linha de índice 1 não existe para ser removida.")

        # Verificar se o índice 2 está presente nas colunas
        if 2 >= len(df.columns):
            raise IndexError("O índice 2 está fora do intervalo de colunas disponíveis.")

        # Criar uma nova coluna 'SIAFI' no índice 3
        # Extrair os 6 últimos dígitos da coluna de índice 2
        df.insert(3, 'SIAFI', df.iloc[:, 2].astype(str).str[-6:])

        # Salvar o DataFrame atualizado no arquivo Excel
        if salvar_como:
            df.to_excel(salvar_como, index=False, header=False)
            print(f"Arquivo salvo como '{salvar_como}'.")
        else:
            df.to_excel(arquivo, index=False, header=False)
            print(f"Arquivo original '{arquivo}' atualizado com sucesso.")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def preencher_estimativa_folha_limite(arquivo, aba, coluna_estimativa, coluna_limite, coluna_despesas, coluna_resultado, salvar_como=None):
    try:
        # Carregar todas as abas do arquivo Excel
        sheets = pd.read_excel(arquivo, sheet_name=None)
        
        # Verificar se a aba desejada está presente no arquivo
        if aba not in sheets:
            raise ValueError(f"A aba '{aba}' não foi encontrada no arquivo.")
        
        # Carregar o DataFrame da aba específica
        df = sheets[aba]
        
        # Remover espaços extras das colunas (se existirem)
        df.columns = df.columns.str.strip()
        
        # Verificar se as colunas necessárias existem
        colunas_necessarias = [coluna_estimativa, coluna_limite, coluna_despesas]
        for coluna in colunas_necessarias:
            if coluna not in df.columns:
                raise ValueError(f"A coluna '{coluna}' não foi encontrada na aba '{aba}'.")

        # Garantir que os dados sejam convertidos para números (tratando valores não numéricos como NaN)
        df[coluna_estimativa] = pd.to_numeric(df[coluna_estimativa], errors='coerce').fillna(0)
        df[coluna_limite] = pd.to_numeric(df[coluna_limite], errors='coerce').fillna(0)
        df[coluna_despesas] = pd.to_numeric(df[coluna_despesas], errors='coerce').fillna(0)
        
       # Correção:
        df[coluna_resultado] = (df[coluna_estimativa] + df[coluna_limite]) > df[coluna_despesas]
        df[coluna_resultado] = df[coluna_resultado].apply(lambda x: 'VERDADEIRO' if x else 'FALSO')
        
        # Atualizar a aba modificada no dicionário
        sheets[aba] = df
        
        # Salvar todas as abas no arquivo Excel (sobrescrevendo ou salvando como novo arquivo)
        with pd.ExcelWriter(salvar_como or arquivo, engine='openpyxl') as writer:
            for sheet_name, data in sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"A coluna '{coluna_resultado}' foi preenchida com sucesso na aba '{aba}'.")
    
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def preencher_coluna_comparacao(arquivo, aba, coluna_liquidadas, coluna_estimativa, nova_coluna, salvar_como=None):
    try:
        # Carregar o arquivo Excel, especificando a aba
        df = pd.read_excel(arquivo, sheet_name=aba)

        # Remover espaços extras das colunas (se existirem)
        df.columns = df.columns.str.strip()

        # Verificar se as colunas necessárias existem
        if coluna_liquidadas not in df.columns or coluna_estimativa not in df.columns:
            raise ValueError("Uma ou mais colunas especificadas não foram encontradas no arquivo.")

        # Preencher a nova coluna com base na comparação
        df[nova_coluna] = df[coluna_liquidadas] > df[coluna_estimativa]
        
        # Converter os valores booleanos para "VERDADEIRO" ou "FALSO"
        df[nova_coluna] = df[nova_coluna].replace({True: 'VERDADEIRO', False: 'FALSO'})

        # Salvar o DataFrame atualizado no arquivo Excel
        with pd.ExcelWriter(salvar_como if salvar_como else arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=aba, index=False)
            print(f"Arquivo atualizado com sucesso na aba '{aba}'.")
            
    except Exception as e:
        print(f"Ocorreu um erro: {e}")


def preencher_coluna_saldo_residencia(caminho_arquivo):
    # Carrega os dados da aba "Tabela Dinâmica"
    df_dinamica = pd.read_excel(caminho_arquivo, sheet_name='Tabela Dinâmica')
    df_dinamica.columns = df_dinamica.columns.str.strip()
    print("Colunas disponíveis na aba 'Tabela Dinâmica':")
    for col in df_dinamica.columns:
        print(f"'{col}'")

    # Calcula o saldo por Órgão UGE
    df_dinamica['Saldo Calculado'] = (
        df_dinamica['DESPESAS EMPENHADAS (CONTROLE EMPENHO)'].fillna(0)
        - df_dinamica['DESPESAS LIQUIDADAS A PAGAR(CONTROLE EMPENHO)'].fillna(0)
        - df_dinamica['DESPESAS PAGAS (CONTROLE EMPENHO)'].fillna(0)
        + df_dinamica['CREDITO DISPONIVEL'].fillna(0)
    )

    # Cria dicionário: Órgão UGE → Saldo
    saldos_por_uge = dict(zip(df_dinamica['Órgão UGE'], df_dinamica['Saldo Calculado']))

    # Carrega o workbook com openpyxl
    wb = load_workbook(caminho_arquivo)
    ws = wb['Acompanhamento Residência']

    # Obtém os nomes das colunas
    header = [cell.value for cell in ws[1]]

    # Localiza índice das colunas
    try:
        col_uge = header.index('Órgão UGE') + 1
    except ValueError:
        raise Exception("Coluna 'Órgão UGE' não encontrada na aba 'Acompanhamento Residência'.")

    try:
        col_saldo = header.index('Saldo') + 1
    except ValueError:
        col_saldo = len(header) + 1
        ws.cell(row=1, column=col_saldo, value='Saldo')

    # Preenche o saldo de acordo com o Órgão UGE
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valor_uge = row[col_uge - 1].value
        saldo = saldos_por_uge.get(valor_uge)

        if saldo is not None:
            row[col_saldo - 1].value = saldo

    # Salva o arquivo
    wb.save(caminho_arquivo)
    print("Coluna 'Saldo' preenchida com base no 'Órgão UGE'.")

def preencher_coluna_saldo_mais_liquidadas(caminho_arquivo):
    # Carrega o workbook
    wb = load_workbook(caminho_arquivo)
    ws = wb['Acompanhamento Residência']

    # Lê cabeçalho
    header = [cell.value for cell in ws[1]]

    # Localiza as colunas necessárias
    try:
        col_saldo = header.index('Saldo') + 1
    except ValueError:
        raise Exception("Coluna 'Saldo' não encontrada.")

    try:
        col_liquidadas = header.index('Despesas liquidadas a Pagar') + 1
    except ValueError:
        raise Exception("Coluna 'Despesas liquidadas a Pagar' não encontrada.")

    # Verifica ou cria a nova coluna
    nome_col_soma = 'Saldo + Despesas Liquidadas a pagar'
    try:
        col_soma = header.index(nome_col_soma) + 1
    except ValueError:
        col_soma = len(header) + 1
        ws.cell(row=1, column=col_soma, value=nome_col_soma)

    # Percorre as linhas preenchendo os valores
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valor_saldo = row[col_saldo - 1].value or 0
        valor_liquidadas = row[col_liquidadas - 1].value or 0

        try:
            soma = float(valor_saldo) + float(valor_liquidadas)
        except:
            soma = None

        row[col_soma - 1].value = soma

    # Salva o arquivo
    wb.save(caminho_arquivo)
    print(f"Coluna '{nome_col_soma}' preenchida com sucesso.")

def verificar_liquidadas_vs_estimativa(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb['Acompanhamento Residência']

    # Lê o cabeçalho
    header = [cell.value for cell in ws[1]]

    # Localiza as colunas
    try:
        col_liquidadas = header.index('Despesas liquidadas a Pagar') + 1
        col_estimativa = header.index('Estimativa Folha') + 1
        col_limite = header.index('Limite de saque') + 1
    except ValueError as e:
        raise Exception(f"Coluna não encontrada: {e}")

    # Criar/identificar nova coluna de resultado
    nome_col_resultado = 'Despesas Liquidadas >= Estimativa da Folha + Limite de saque?'
    try:
        col_resultado = header.index(nome_col_resultado) + 1
    except ValueError:
        col_resultado = len(header) + 1
        ws.cell(row=1, column=col_resultado, value=nome_col_resultado)

    # Preenche os valores
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valor_liquidadas = row[col_liquidadas - 1].value or 0
        valor_estimativa = row[col_estimativa - 1].value or 0
        valor_limite = row[col_limite - 1].value or 0

        try:
            resultado = (float(valor_liquidadas) >= (float(valor_estimativa) + float(valor_limite)))
        except:
            resultado = False  # trata erros com dados vazios ou inválidos

        row[col_resultado - 1].value = resultado

    # Salva o arquivo
    wb.save(caminho_arquivo)
    print(f"Coluna '{nome_col_resultado}' preenchida com sucesso.")

def preencher_coluna_comparacao_residencia(caminho_arquivo):
    # Abre o Excel com openpyxl
    wb = load_workbook(caminho_arquivo)
    ws = wb['Acompanhamento Residência']

    # Lê os dados da aba com pandas para facilitar a comparação
    df = pd.read_excel(caminho_arquivo, sheet_name='Acompanhamento Residência')

    # Corrige espaços extras nos nomes das colunas
    df.columns = df.columns.str.strip()

    # Nome da nova coluna
    nova_coluna = "Saldo + Despesas liquidadas >= Estimativa da Folha + Limite de saque?"

    # Garante que a nova coluna exista na planilha
    header = [cell.value for cell in ws[1]]
    try:
        col_resultado = header.index(nova_coluna) + 1
    except ValueError:
        col_resultado = len(header) + 1
        ws.cell(row=1, column=col_resultado, value=nova_coluna)

    # Preenche a nova coluna com a comparação
    for i, row in df.iterrows():
        valor_saldo_despesas = row.get("Saldo + Despesas Liquidadas a pagar", 0) or 0
        estimativa_folha = row.get("Estimativa Folha", 0) or 0
        limite_saque = row.get("Limite saque", 0) or 0

        resultado = (float(valor_saldo_despesas) >= (float(estimativa_folha) + float(limite_saque)))
        ws.cell(row=i+2, column=col_resultado, value="VERDADEIRO" if resultado else "FALSO")

    # Salva o arquivo
    wb.save(caminho_arquivo)


def preencher_total_repassar_consolidado(caminho_arquivo, nome_aba='Acompanhamento Residência'):
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]

    header = [cell.value for cell in ws[1]]

    col_saldo_liquidadas = header.index("Saldo + Despesas Liquidadas a pagar") + 1
    col_estimativa_folha = header.index("Estimativa Folha") + 1
    col_limite_saque = header.index("Limite de saque") + 1
    col_desp_liquidadas = header.index("Despesas liquidadas a Pagar") + 1
    col_saldo_a_repassar = header.index("Saldo a repassar") + 1

    try:
        col_total_repassar = header.index("Total a Repassar Consolidado") + 1
    except ValueError:
        col_total_repassar = len(header) + 1
        ws.cell(row=1, column=col_total_repassar, value="Total a Repassar Consolidado")

    # Criar ou localizar as colunas das condições
    nome_coluna_1 = "Saldo + Despesas liquidadas >= Valor a repassar consolidado"
    nome_coluna_2 = "Saldo a repassar >= Valor a repassar consolidado?"

    if nome_coluna_1 in header:
        col_condicao_1 = header.index(nome_coluna_1) + 1
    else:
        col_condicao_1 = len(header) + 1
        ws.cell(row=1, column=col_condicao_1, value=nome_coluna_1)

    if nome_coluna_2 in header:
        col_condicao_2 = header.index(nome_coluna_2) + 1
    else:
        col_condicao_2 = len(header) + 1
        ws.cell(row=1, column=col_condicao_2, value=nome_coluna_2)

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        saldo_liquidadas = row[col_saldo_liquidadas - 1].value or 0
        estimativa_folha = row[col_estimativa_folha - 1].value or 0
        limite_saque = row[col_limite_saque - 1].value or 0
        despesas_liquidadas = row[col_desp_liquidadas - 1].value or 0
        saldo_a_repassar = row[col_saldo_a_repassar - 1].value or 0

        if all(isinstance(val, (int, float)) for val in [saldo_liquidadas, estimativa_folha, limite_saque, despesas_liquidadas]):
            # Regra 1
            if saldo_liquidadas < estimativa_folha:
                total_repassar = saldo_liquidadas - limite_saque
            else:
                total_repassar = estimativa_folha

            # Regra 2
            if despesas_liquidadas > estimativa_folha:
                total_repassar = despesas_liquidadas - limite_saque

            ws.cell(row=i, column=col_total_repassar, value=total_repassar)

            # Preencher condição 1: Saldo + Despesas liquidadas >= Total a Repassar
            condicao_1 = (saldo_liquidadas >= total_repassar)
            ws.cell(row=i, column=col_condicao_1, value="VERDADEIRO" if condicao_1 else "FALSO")

            # Preencher condição 2: Saldo a Repassar >= Total a Repassar
            condicao_2 = (saldo_a_repassar >= total_repassar)
            ws.cell(row=i, column=col_condicao_2, value="VERDADEIRO" if condicao_2 else "FALSO")

    wb.save(caminho_arquivo)

def formatar_contabil(value):
    if pd.notnull(value):
        if isinstance(value, (int, float)):
            return "{:,.2f}".format(float(value)).replace(",", "_").replace(".", ",").replace("_", ".")
        else:
            return value
    return None

def aplicar_formatacao_contabil(caminho_arquivo, nome_aba='Acompanhamento Residência'):
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]

    colunas_para_formatar = [
        "Valor Bolsa",
        "Tributos",
        "Limite de saque",
        "Estimativa Folha",
        "Total a Repassar Consolidado",
        "Despesas liquidadas a Pagar",
        "Saldo",
        "Saldo + Despesas Liquidadas a pagar",
        "Saldo a repassar",
    ]

    # Mapeia os nomes das colunas para índices
    header = [cell.value for cell in ws[1]]
    col_indices = {
        nome: header.index(nome) + 1
        for nome in colunas_para_formatar
        if nome in header
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for nome_coluna, col_index in col_indices.items():
            cell = row[col_index - 1]
            valor_original = cell.value
            if isinstance(valor_original, (int, float)):
                cell.value = formatar_contabil(valor_original)

    wb.save(caminho_arquivo)
    
def main():
    
    pasta_destino = r'W:\B - TED\7 - AUTOMAÇÃO\Residência'
    arquivo_residencia_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx'
    limite_de_saque_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Limite de Saque Residência.xlsx'
    contas_cadastro_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA TED - Contas Cadastro e Controle.xlsx'
    
    #Copiar arquivos para a pasta de destino
    copiar_arquivos(pasta_destino)
    
    substituir_valores_coluna(
    caminho_arquivo=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
    coluna_idx=7,  # índice Python (base 0) da 8ª coluna
    novo_valor="1000A0008U"
    )
    
    arquivos_config = [
        # Adicionar e Remover Colunas
        {
            'arquivo': r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
            'colunas_a_remover': [11],
            'colunas_a_adicionar': [(11, 'ORGÃO+FONTE', ''), (18, 'SALDO', ''), (19, 'SALDO + DESPESAS LIQUIDADAS', ''), (20, 'SIAFI', ''), (21, 'TED', ''),(22, 'GESTÃO', ''),(23, 'UG Proponente', '')]
        },
    ]
    
    concatenar_config = [    
        # Concatenar
        {
            'arquivo': r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
            'coluna1_idx': 0,  # Índice da primeira coluna a ser concatenada
            'coluna2_idx': 7,  # Índice da segunda coluna a ser concatenada
            'coluna_resultado_idx': 11,
        },
    ]    
        
    expressao_config = [
    #Calcular Saldo com expressão A - (B - C) + D
        {
            'arquivo': r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
            'col_a_idx': 15,   # Índice da coluna A
            'col_b_idx': 16,   # Índice da coluna B
            'col_c_idx': 17,   # Índice da coluna C
            'col_d_idx': 13,   # Índice da coluna D
            'col_resultado_idx': 18,
        },
        
    ]
    
    # Lista de configurações para a subtração A - B
    subtracao_config = [
        {
            'arquivo': r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
            'col_a_idx': 18,   # Índice da coluna A
            'col_b_idx': 16,   # Índice da coluna B
            'col_resultado_idx': 19  # Índice da coluna onde o resultado será armazenado
        }
    ]
    
    # Processar cada arquivo de acordo com a configuração
    for config in arquivos_config:
        arquivo = config.get('arquivo')
        
        # Remover colunas (se aplicável)
        if config.get('colunas_a_remover'):
            remover_colunas_excel(arquivo, config['colunas_a_remover'])
        
        # Adicionar colunas (se aplicável)
        if config.get('colunas_a_adicionar'):
            for col in config['colunas_a_adicionar']:
                indice, nome, valor_padrao = col
                adicionar_coluna_excel(arquivo, indice, nome, valor_padrao)
    
    # Processar cada arquivo
    for config in concatenar_config:
        arquivo = config.get('arquivo')
        concatenar_colunas_excel(
            arquivo,
            config.get('coluna1_idx'),
            config.get('coluna2_idx'),
            config.get('coluna_resultado_idx'),
        )
    
    # Calcular expressão (se aplicável)
    for config in expressao_config:
        if all(key in config for key in ['col_a_idx', 'col_b_idx', 'col_c_idx', 'col_d_idx', 'col_resultado_idx']):
            calcular_expressao_saldo(
                config['arquivo'],
                config['col_a_idx'],
                config['col_b_idx'],
                config['col_c_idx'],
                config['col_d_idx'],
                config['col_resultado_idx']
            )
    
   # Executar a operação de subtração para cada arquivo
    for config in subtracao_config:
        calcular_subtracao_SD(
            config['arquivo'],
            config['col_a_idx'],
            config['col_b_idx'],
            config['col_resultado_idx']
        )
    
    # Array com os novos nomes das colunas
    novos_nomes_colunas = [
        'Órgão UGE', 'Descrição UGE', 'UG Executora', 'Descrição UG Executora', 'PTRES', 'PI', 'Descrição PI', 'Fonte Recursos Detalhada', 'Descrição Fonte', 'Natureza Despesa' , 'Descrição Natureza', 'Órgão + Fonte', 'Destaque Recebido', 'Crédito Disponível', 'Despesas Pré Empenhadas a empenhar', 'Despesas Empenhadas', 'Despesas Liquidadas a Pagar','Despesas Pagas', 'Saldo', 'Saldo + Despesas Liquidadas', 'SIAFI', 'TED', 'GESTÃO','UG Proponente'
    ]
    novos_nomes_colunas2 = [
        'Órgão UGE', 'Descrição UGE','Órgão UGE Gestão','Descrição Gestão','Fonte Recursos Detalhada', 'Descrição Fonte','Vinculação Pagamento','Descrição Vinculação','Item Informação','Descrição Informação','Limite de Saque'
    ]
    
    novos_nomes_colunas3 = [
        'UG Executora ', 'Descrição UG Executora','Conta corrente','SIAFI','Final Vigência','Valores Firmados', 'A Repassar','A Comprovar','Comprovado','Valor não repassado/devolvido'
    ]
    
    # Chamar a função para renomear as colunas
    renomear_colunas_excel(arquivo_residencia_copia, novos_nomes_colunas)

    preencher_colunas_UG_linha_a_linha(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        coluna_destino='SIAFI',
        chave_coluna_origem='UG Executora',
        chave_coluna_destino='UG Proponente',
        coluna_siafi_valores='SIAFI'
    )
    
    preencher_colunas_UG_linha_a_linha(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        coluna_destino='TED',
        chave_coluna_origem='UG Executora',
        chave_coluna_destino='UG Proponente',
        coluna_siafi_valores='TED'
    )
    
    preencher_colunas_UG_linha_a_linha(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        coluna_destino='GESTÃO',
        chave_coluna_origem='UG Executora',
        chave_coluna_destino='UG Proponente',
        coluna_siafi_valores='GESTÃO'
    )
    
    preencher_colunas_UG_linha_a_linha(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        coluna_destino='UG Proponente',
        chave_coluna_origem='Órgão UGE',
        chave_coluna_destino='Órgão UGE',
        coluna_siafi_valores='UG Proponente'
    ) 

    # Exemplo de uso da função
    tabela_dinamica(
        arquivo_origem=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        colunas_linhas=[0, 1, 11],  
        colunas_valores_idx=[12, 13, 14 ,15 ,16, 19],  
        funcao_agregacao='sum'
    )
    
    # Exemplo de uso da função
    td_acompanhamento(
        arquivo_origem=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        colunas_linhas=[0, 1, 11, 7],  # Índices das 4 colunas que você deseja agrupar
        sheet_name='Acompanhamento Residência'
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='PROCV RESIDENCIA 2024',
        coluna_destino='UG Proponente',
        chave_coluna_origem='Órgão UGE',
        chave_coluna_destino='Órgão UGE',
        coluna_siafi_valores='UG Proponente'
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='PROCV RESIDENCIA 2024',
        coluna_destino='GESTÃO',
        chave_coluna_origem='Órgão UGE',
        chave_coluna_destino='Órgão UGE',
        coluna_siafi_valores='GESTÃO'
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='PROCV RESIDENCIA 2024',
        coluna_destino='TED',
        chave_coluna_origem='Órgão UGE',
        chave_coluna_destino='Órgão UGE',
        coluna_siafi_valores='TED'
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\PROCV RESIDÊNCIA 2024.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='PROCV RESIDENCIA 2024',
        coluna_destino='SIAFI',
        chave_coluna_origem='Órgão UGE',
        chave_coluna_destino='Órgão UGE',
        coluna_siafi_valores='SIAFI'
    )
    
    # Exemplo de uso para adicionar múltiplas colunas na aba 'Dados'
    adicionar_colunas_aba(
        arquivo=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        aba='Acompanhamento Residência',
        indices_colunas=[8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,],  # Posições onde as colunas serão inseridas
        nomes_colunas=['Valor Bolsa', 
                       'Tributos',
                       'Limite de saque',
                       'Estimativa Folha',
                       'Total a Repassar Consolidado',
                       'Despesas liquidadas a Pagar',
                       'Saldo',
                       'Saldo + Despesas Liquidadas a pagar',
                       'Despesas Liquidadas >= Estimativa da Folha + Limite de saque?',
                       'Saldo + Despesas liquidadas >= Estimativa da Folha + Limite de saque?',
                       'Saldo + Despesas liquidadas >= Valor a repassar consolidado',
                       'Saldo a repassar',
                       'Saldo a repassar >= Valor a repassar consolidado?'],  # Nomes das colunas
        valores_padrao=['', '', '', '', '', '', '', '', '', '', '', '', '',],  # Valores padrão para cada coluna
        salvar_como= None # Deixe como None se quiser salvar no mesmo arquivo
)
    
    excluir_linhas_por_valores(
        arquivo=arquivo_residencia_copia,
        aba='Acompanhamento Residência',
        coluna_filtro='Fonte Recursos Detalhada',
        valores_excluir=['1000000000', '1012000000','1000A0008O'],
        salvar_como=None  # Deixe como None para salvar no mesmo arquivo
    )
    
    # Exemplo de uso
    arquivo_residencia_copia = r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx'
    garantir_fontes_por_uge(
        arquivo=arquivo_residencia_copia,
        aba='Acompanhamento Residência',
        coluna_uge='Órgão UGE',
        coluna_fonte='Fonte Recursos Detalhada',
        fontes_necessarias=['1000A0008U'],
        salvar_como=None  # Deixe como None para salvar no mesmo arquivo
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\Folha Residência.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='Plan1',
        coluna_destino='Valor Bolsa',
        chave_coluna_origem='Órgão UGE',
        chave_coluna_destino='ÓRGÃO',
        coluna_siafi_valores='RESIDÊNCIA MÉDICA'
    )
    
    preencher_coluna_valor_bolsa_tributos(
        arquivo=arquivo_residencia_copia,
        aba='Acompanhamento Residência',
        coluna_bolsa='Valor Bolsa',
        coluna_bolsa_tributos='Tributos',
        salvar_como=None  # Deixe como None para salvar no mesmo arquivo
    )
    
    # Chamar a função para renomear as colunas
    renomear_colunas_excel(limite_de_saque_copia, novos_nomes_colunas2)
    
    arquivos_config2 = [
        # Adicionar e Remover Colunas
        {
            'arquivo': r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Limite de Saque Residência.xlsx',
            'colunas_a_adicionar': [(6, 'ORGÃO+FONTE', '')]
        },
    ]
    
    concatenar_config2 = [    
        # Concatenar
        {
            'arquivo': r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Limite de Saque Residência.xlsx',
            'coluna1_idx': 0,  # Índice da primeira coluna a ser concatenada
            'coluna2_idx': 4,  # Índice da segunda coluna a ser concatenada
            'coluna_resultado_idx': 6,
        },
    ]
    
    # Processar cada arquivo de acordo com a configuração
    for config in arquivos_config2:
        arquivo = config.get('arquivo')
        if config.get('colunas_a_adicionar'):
            for col in config['colunas_a_adicionar']:
                indice, nome, valor_padrao = col
                adicionar_coluna_excel(arquivo, indice, nome, valor_padrao)    
    
    # Processar cada arquivo
    for config in concatenar_config2:
        arquivo = config.get('arquivo')
        concatenar_colunas_excel(
            arquivo,
            config.get('coluna1_idx'),
            config.get('coluna2_idx'),
            config.get('coluna_resultado_idx'),
        )
    
    # Exemplo de uso
    arquivo = r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx'
    concatenar_colunas_excel_por_aba(
        arquivo=arquivo,
        aba='Acompanhamento Residência',
        coluna1_idx=0,  # Índice da primeira coluna para concatenar
        coluna2_idx=3,  # Índice da segunda coluna para concatenar
        coluna_resultado_idx=2,  # Índice da coluna onde o resultado será salvo
        salvar_como=None  # Deixe como None para sobrescrever o arquivo original
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Limite de Saque Residência.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='Sheet1',
        coluna_destino='Limite de saque',
        chave_coluna_origem='ORGÃO+FONTE',
        chave_coluna_destino='ORGÃO+FONTE',
        coluna_siafi_valores='Limite de Saque'
    )

    preencher_coluna_estimativa_folha(
        arquivo=arquivo_residencia_copia,
        aba='Acompanhamento Residência',
        col_bolsa='Valor Bolsa',
        col_bolsa_tributos='Tributos',
        col_limite_saque='Limite de saque',
        col_estimativa_folha='Estimativa Folha',
        salvar_como=None  # Deixe como None para salvar no mesmo arquivo
    )
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='Tabela Dinâmica',
        coluna_destino='Despesas liquidadas a Pagar',
        chave_coluna_origem='ORGÃO+FONTE',
        chave_coluna_destino='ORGÃO+FONTE',
        coluna_siafi_valores='DESPESAS LIQUIDADAS A PAGAR(CONTROLE EMPENHO)'
    ) 
     
    manipular_excel(
        arquivo=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA TED - Contas Cadastro e Controle.xlsx',
        salvar_como=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA TED - Contas Cadastro e Controle.xlsx'
    )
    
    # Chamar a função para renomear as colunas
    renomear_colunas_excel(contas_cadastro_copia, novos_nomes_colunas3)
    
    preencher_colunas_UG_linha_a_linha2(
        copia_ted_para_finalizar=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx',
        copia_arquivo_valores=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA TED - Contas Cadastro e Controle.xlsx',
        aba_destino='Acompanhamento Residência',
        aba_origem='Sheet1',
        coluna_destino='Saldo a repassar',
        chave_coluna_origem='SIAFI',
        chave_coluna_destino='SIAFI',
        coluna_siafi_valores='A Repassar'
    )
      
    preencher_coluna_saldo_residencia(r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx')
    preencher_coluna_saldo_mais_liquidadas(r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx')
    verificar_liquidadas_vs_estimativa(r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx')
    preencher_coluna_comparacao_residencia(r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx')
    preencher_total_repassar_consolidado(caminho_arquivo=r'W:\B - TED\7 - AUTOMAÇÃO\Residência\COPIA Residência Acompanhamento.xlsx')
    aplicar_formatacao_contabil(arquivo_residencia_copia)

    print("Processo totalmente finalizado.")
    
if __name__ == "__main__":
    main()